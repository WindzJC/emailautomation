#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter
from email.utils import parseaddr
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

from openpyxl import load_workbook

APP_ROOT = Path(__file__).resolve().parent.parent
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))

import settings


OUTPUT_COLUMNS = [
    "first_name",
    "last_name",
    "email",
    "AuthorName",
    "AuthorEmail",
    "BookTitle",
    "PersonalizedOpeningLine",
    "SourceURL",
    "ConfidenceScore",
    "Website",
    "BookURL",
    "WhyAstraFit",
    "source_file",
    "source_sheet",
    "source_row",
]
OUTPUT_FILENAME = "lead_op_author_personalized_upload.csv"
REVIEW_FILENAME = "lead_op_author_personalized_review.csv"
REVIEW_COLUMNS = OUTPUT_COLUMNS + ["review_reason"]

EMAIL_RE = re.compile(r"^[^@\s<>]+@[^@\s<>]+\.[^@\s<>]+$")
URL_RE = re.compile(r"(?i)\b(?:https?://|www\.)\S+")
PHONE_RE = re.compile(r"^[+()\d\s.\-]{7,}$")
PERCENT_RE = re.compile(r"^\s*\d+(?:\.\d+)?\s*%\s*$")
DATE_RE = re.compile(r"^\s*\d{1,4}[-/]\d{1,2}(?:[-/]\d{1,4})?\s*$")
PLACEHOLDER_RE = re.compile(r"{[A-Za-z][A-Za-z0-9_]*}|\[[^\[\]\r\n]+\]|<<[^<>\r\n]+>>")
BRACE_PLACEHOLDER_RE = re.compile(r"{([A-Za-z][A-Za-z0-9_]*)}")
SQUARE_PLACEHOLDER_RE = re.compile(r"\[([^\[\]\r\n]+)\]")
ANGLE_PLACEHOLDER_RE = re.compile(r"<<([^<>\r\n]+)>>")

BAD_NAME_VALUES = {
    "admin",
    "approved",
    "archway",
    "authorhouse",
    "authorhouseuk",
    "balboa",
    "bookbaby",
    "canceled",
    "cancelled",
    "complete",
    "completed",
    "customer service",
    "in progress",
    "iuniverse",
    "lulu",
    "n/a",
    "na",
    "not started",
    "pending",
    "rejected",
    "resubmission",
    "submission",
    "support",
    "trafford",
    "unknown",
    "westbow",
    "xlibris",
}
BAD_BOOK_VALUES = BAD_NAME_VALUES | {
    "100%",
    "50%",
    "book trailer",
    "ebook",
    "hardcover",
    "illustration package",
    "launch package",
    "marketing package",
    "none",
    "paperback",
    "publishing package",
    "tbd",
    "untitled",
    "website",
    "website package",
}
PACKAGE_TERMS = ("package", "service", "services", "paperback", "hardcover", "ebook", "book trailer", "website")

HEADER_ALIASES = {
    "email": ("email", "e-mail", "e_mail", "mail"),
    "first_name": ("first_name", "firstname", "first name", "first"),
    "last_name": ("last_name", "lastname", "last name", "last"),
    "AuthorEmail": ("authoremail", "author email", "author_email", "email"),
    "AuthorName": ("authorname", "author name", "author_name", "full name", "fullname", "name", "author"),
    "BookTitle": ("booktitle", "book title", "book_title", "title", "publication title", "publicationtitle"),
    "PersonalizedOpeningLine": ("personalizedopeningline", "personalized opening line", "opening line"),
    "SourceURL": ("sourceurl", "source url", "source_url", "source", "url"),
    "ConfidenceScore": ("confidencescore", "confidence score", "confidence_score", "confidence"),
    "Website": ("website", "author website", "site"),
    "BookURL": ("bookurl", "book url", "book_url", "retailer link", "retailerlink"),
    "WhyAstraFit": ("whyastrafit", "why astra fit", "why_astra_fit"),
}


def clean_space(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").strip().split())


def normalize_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


BAD_NAME_KEYS = {normalize_key(item) for item in BAD_NAME_VALUES}
BAD_BOOK_KEYS = {normalize_key(item) for item in BAD_BOOK_VALUES}


def normalize_email(value: object) -> str:
    _, addr = parseaddr(clean_space(value))
    return addr.strip().lower()


def valid_email(value: str) -> bool:
    return bool(value and EMAIL_RE.match(value))


def looks_like_phone(value: str) -> bool:
    text = clean_space(value)
    digits = re.sub(r"\D+", "", text)
    return bool(PHONE_RE.match(text) and len(digits) >= 7)


def looks_like_percentage(value: str) -> bool:
    return bool(PERCENT_RE.match(clean_space(value)))


def unsafe_name(value: str) -> bool:
    text = clean_space(value).strip("\"'`“”‘’()[]{}:;,.")
    if not text:
        return True
    if normalize_key(text) in BAD_NAME_KEYS:
        return True
    if valid_email(normalize_email(text)) or URL_RE.search(text) or looks_like_phone(text) or looks_like_percentage(text):
        return True
    letters = re.sub(r"[^A-Za-zÀ-ÖØ-öø-ÿ]+", "", text)
    return len(letters) <= 1


def unsafe_book_title(book_title: str, author_name: str, first_name: str) -> bool:
    title = clean_space(book_title)
    if not title:
        return False
    key = normalize_key(title)
    if key in BAD_BOOK_KEYS:
        return True
    if key and key in {normalize_key(author_name), normalize_key(first_name)}:
        return True
    if valid_email(normalize_email(title)) or URL_RE.search(title) or looks_like_phone(title) or looks_like_percentage(title):
        return True
    if DATE_RE.match(title):
        return True
    lower = title.lower()
    return any(term in lower for term in PACKAGE_TERMS)


def normalize_placeholder_punctuation(value: str) -> tuple[str, bool]:
    raw = clean_space(value)
    cleaned = BRACE_PLACEHOLDER_RE.sub(lambda match: match.group(1), raw)
    cleaned = SQUARE_PLACEHOLDER_RE.sub(lambda match: f"({match.group(1).strip()})", cleaned)
    cleaned = ANGLE_PLACEHOLDER_RE.sub(lambda match: match.group(1).strip(), cleaned)
    return clean_space(cleaned), cleaned != raw


def confidence_score(value: str) -> str:
    text = clean_space(value)
    if text in {"2", "3"}:
        return text
    try:
        numeric = int(float(text))
    except Exception:
        return ""
    return str(numeric) if numeric in {2, 3} else ""


def split_author_name(author_name: str) -> tuple[str, str]:
    parts = [part.strip("\"'`“”‘’()[]{}:;,.") for part in clean_space(author_name).split()]
    if not parts:
        return "", ""
    return parts[0], " ".join(parts[1:])


def opening_line_for_book(book_title: str) -> str:
    title = clean_space(book_title)
    if not title:
        return ""
    return (
        f"Our team came across {title} and thought the book could benefit from a clearer, "
        "more polished online presentation for readers discovering your work."
    )


def build_header_map(headers: Sequence[str]) -> Dict[str, str]:
    by_key = {normalize_key(header): header for header in headers if clean_space(header)}
    out: Dict[str, str] = {}
    for target, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            match = by_key.get(normalize_key(alias))
            if match:
                out[target] = match
                break
    return out


def source_value(row: Dict[str, str], header_map: Dict[str, str], field: str) -> str:
    source_header = header_map.get(field, "")
    return clean_space(row.get(source_header, "")) if source_header else ""


def first_from_author(author_name: str) -> str:
    return split_author_name(author_name)[0]


def last_from_author(author_name: str) -> str:
    return split_author_name(author_name)[1]


def author_from_parts(first_name: str) -> str:
    return clean_space(first_name)


def iter_xlsx_rows(paths: Sequence[Path]) -> Iterable[Dict[str, object]]:
    for path in paths:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                if getattr(sheet, "sheet_state", "visible") != "visible":
                    continue
                rows = sheet.iter_rows(values_only=True)
                header_values: Sequence[object] | None = None
                header_row = 0
                for row_number, values in enumerate(rows, start=1):
                    if any(clean_space(value) for value in values):
                        header_values = values
                        header_row = row_number
                        break
                if not header_values:
                    continue
                headers = [clean_space(value) or f"column_{idx}" for idx, value in enumerate(header_values, start=1)]
                header_map = build_header_map(headers)
                for row_number, values in enumerate(rows, start=header_row + 1):
                    raw = {headers[idx]: clean_space(values[idx]) if idx < len(values) else "" for idx in range(len(headers))}
                    if not any(raw.values()):
                        continue
                    yield {
                        "row": raw,
                        "header_map": header_map,
                        "source_file": path.name,
                        "source_sheet": sheet.title,
                        "source_row": str(row_number),
                    }
        finally:
            workbook.close()


def review_row(base: Dict[str, str], reason: str) -> Dict[str, str]:
    row = {column: clean_space(base.get(column, "")) for column in OUTPUT_COLUMNS}
    row["review_reason"] = reason
    return row


def build_upload_row(
    item: Dict[str, object],
    counts: Counter[str],
    review_rows: List[Dict[str, str]],
) -> Dict[str, str] | None:
    row = item["row"]
    header_map = item["header_map"]
    assert isinstance(row, dict)
    assert isinstance(header_map, dict)

    email = normalize_email(source_value(row, header_map, "email") or source_value(row, header_map, "AuthorEmail"))
    author_email = normalize_email(source_value(row, header_map, "AuthorEmail") or source_value(row, header_map, "email"))
    base = {
        "email": email,
        "AuthorEmail": author_email,
        "AuthorName": source_value(row, header_map, "AuthorName"),
        "BookTitle": source_value(row, header_map, "BookTitle"),
        "source_file": str(item["source_file"]),
        "source_sheet": str(item["source_sheet"]),
        "source_row": str(item["source_row"]),
    }
    if not email or not valid_email(email):
        counts["invalid_email_rows_removed"] += 1
        review_rows.append(review_row(base, "invalid_or_missing_email"))
        return None
    if not author_email:
        author_email = email

    author_name = source_value(row, header_map, "AuthorName")
    first_name = source_value(row, header_map, "first_name")
    last_name = source_value(row, header_map, "last_name")
    if not author_name and first_name:
        author_name = author_from_parts(first_name)
    if not first_name and author_name:
        first_name = first_from_author(author_name)
    if not last_name and author_name:
        last_name = last_from_author(author_name)
    if unsafe_name(author_name) or unsafe_name(first_name):
        counts["unsafe_author_rows_removed"] += 1
        base.update({"first_name": first_name, "last_name": last_name, "AuthorEmail": author_email, "AuthorName": author_name})
        review_rows.append(review_row(base, "unsafe_author_name"))
        return None

    book_title = source_value(row, header_map, "BookTitle")
    if unsafe_book_title(book_title, author_name, first_name):
        counts["book_titles_cleared"] += 1
        base.update({"first_name": first_name, "last_name": last_name, "AuthorEmail": author_email, "AuthorName": author_name})
        review_rows.append(review_row(base, "suspicious_book_title_cleared"))
        book_title = ""
    else:
        normalized_book_title, title_normalized = normalize_placeholder_punctuation(book_title)
        if title_normalized:
            counts["book_titles_normalized"] += 1
            base.update({"first_name": first_name, "last_name": last_name, "AuthorEmail": author_email, "AuthorName": author_name})
            review_rows.append(review_row(base, "book_title_placeholder_punctuation_normalized"))
            book_title = normalized_book_title
    if book_title:
        counts["safe_book_titles_preserved"] += 1

    confidence = confidence_score(source_value(row, header_map, "ConfidenceScore"))
    if not confidence:
        counts["confidence_rows_removed"] += 1
        base.update({"first_name": first_name, "last_name": last_name, "AuthorEmail": author_email, "AuthorName": author_name, "BookTitle": book_title})
        review_rows.append(review_row(base, "confidence_score_not_2_or_3"))
        return None
    counts[f"confidence_{confidence}_rows"] += 1

    personalized_opening = opening_line_for_book(book_title)

    return {
        "first_name": first_name,
        "last_name": last_name,
        "email": email,
        "AuthorName": author_name,
        "AuthorEmail": author_email,
        "BookTitle": book_title,
        "PersonalizedOpeningLine": personalized_opening,
        "SourceURL": source_value(row, header_map, "SourceURL"),
        "ConfidenceScore": confidence,
        "Website": source_value(row, header_map, "Website"),
        "BookURL": source_value(row, header_map, "BookURL"),
        "WhyAstraFit": source_value(row, header_map, "WhyAstraFit"),
        "source_file": str(item["source_file"]),
        "source_sheet": str(item["source_sheet"]),
        "source_row": str(item["source_row"]),
    }


def winner_score(row: Dict[str, str]) -> tuple[int, int, int]:
    return (
        int(row.get("ConfidenceScore") == "3"),
        int(bool(row.get("BookTitle"))),
        len(row.get("AuthorName", "")),
        sum(1 for column in OUTPUT_COLUMNS if clean_space(row.get(column, ""))),
    )


def validate_upload_rows(rows: Sequence[Dict[str, str]]) -> None:
    seen: set[str] = set()
    for row in rows:
        if list(row.keys()) != OUTPUT_COLUMNS:
            raise AssertionError("Upload CSV columns do not match required schema.")
        email = row["email"]
        if not email or not valid_email(email):
            raise AssertionError(f"Invalid email in upload CSV: {email}")
        if email in seen:
            raise AssertionError(f"Duplicate email in upload CSV: {email}")
        seen.add(email)
        if not row["AuthorEmail"]:
            raise AssertionError("AuthorEmail missing in upload CSV.")
        if unsafe_name(row["AuthorName"]) or unsafe_name(row["first_name"]):
            raise AssertionError("Unsafe author name in upload CSV.")
        if unsafe_book_title(row["BookTitle"], row["AuthorName"], row["first_name"]):
            raise AssertionError("Unsafe BookTitle in upload CSV.")
        if row["ConfidenceScore"] not in {"2", "3"}:
            raise AssertionError("ConfidenceScore must be 2 or 3.")
        if PLACEHOLDER_RE.search(" ".join(row.get(column, "") for column in OUTPUT_COLUMNS)):
            raise AssertionError("Unresolved placeholder token in upload CSV.")
        if row["PersonalizedOpeningLine"] and row["BookTitle"] not in row["PersonalizedOpeningLine"]:
            raise AssertionError("PersonalizedOpeningLine does not use same-row BookTitle.")


def write_upload_csv(path: Path, rows: Sequence[Dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_review_csv(path: Path, rows: Sequence[Dict[str, str]]) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=REVIEW_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def clean_workbooks(paths: Sequence[Path], output_dir: Path) -> Dict[str, object]:
    counts: Counter[str] = Counter()
    winners: Dict[str, Dict[str, str]] = {}
    review_rows: List[Dict[str, str]] = []
    for item in iter_xlsx_rows(paths):
        counts["total_input_rows"] += 1
        row = build_upload_row(item, counts, review_rows)
        if row is None:
            continue
        email = row["email"]
        current = winners.get(email)
        if current is None:
            winners[email] = row
        else:
            counts["duplicate_emails_removed"] += 1
            if winner_score(row) > winner_score(current):
                review_rows.append(review_row(current, "duplicate_email_lost_to_stronger_row"))
                winners[email] = row
            else:
                review_rows.append(review_row(row, "duplicate_email_lost_to_stronger_row"))

    upload_rows = list(winners.values())
    upload_rows.sort(key=lambda item: (item["source_file"], item["source_sheet"], int(item["source_row"] or 0)))
    validate_upload_rows(upload_rows)
    upload_path = output_dir / OUTPUT_FILENAME
    review_path = output_dir / REVIEW_FILENAME
    write_upload_csv(upload_path, upload_rows)
    write_review_csv(review_path, review_rows)

    counts["upload_rows"] = len(upload_rows)
    counts["review_rows"] = len(review_rows)
    counts["rows_with_book_title"] = sum(1 for row in upload_rows if row["BookTitle"])
    return {"upload_path": upload_path, "review_path": review_path if review_rows else None, "counts": counts}


def print_summary(counts: Counter[str]) -> None:
    print(f"total input rows: {counts['total_input_rows']}")
    print(f"upload rows: {counts['upload_rows']}")
    print(f"review rows: {counts['review_rows']}")
    print(f"invalid email rows removed: {counts['invalid_email_rows_removed']}")
    print(f"unsafe author rows removed: {counts['unsafe_author_rows_removed']}")
    print(f"book titles cleared: {counts['book_titles_cleared']}")
    print(f"book titles normalized: {counts['book_titles_normalized']}")
    print(f"safe BookTitle values preserved: {counts['safe_book_titles_preserved']}")
    print(f"duplicate emails removed: {counts['duplicate_emails_removed']}")
    print(f"rows with BookTitle: {counts['rows_with_book_title']}")
    print(f"confidence 2 rows: {counts['confidence_2_rows']}")
    print(f"confidence 3 rows: {counts['confidence_3_rows']}")
    print(f"confidence rows removed: {counts['confidence_rows_removed']}")


def resolve_inputs(values: Sequence[str], input_dir: Path) -> List[Path]:
    paths = [Path(value).expanduser() for value in values]
    if not paths:
        paths = sorted(input_dir.glob("*.xlsx"))
    resolved = [(path if path.is_absolute() else (Path.cwd() / path)).resolve() for path in paths]
    missing = [path for path in resolved if not path.exists()]
    if missing:
        raise SystemExit(f"Missing input file(s): {', '.join(str(path) for path in missing)}")
    return resolved


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Clean AUTHOR-PERSONALIZED Astra Lead Op XLSX files.")
    parser.add_argument("xlsx", nargs="*", help="Input .xlsx files. If omitted, scans --input-dir.")
    parser.add_argument("--input-dir", default=str(settings.UPLOADS_DIR), help="Folder to scan for .xlsx files when none are passed.")
    parser.add_argument("--output-dir", default=".", help="Folder for lead_op_author_personalized_upload.csv.")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    input_paths = resolve_inputs(args.xlsx, Path(args.input_dir).expanduser())
    if not input_paths:
        raise SystemExit("No .xlsx files found to clean.")
    output_dir = Path(args.output_dir).expanduser()
    if not output_dir.is_absolute():
        output_dir = Path.cwd() / output_dir
    result = clean_workbooks(input_paths, output_dir)
    print(f"wrote upload CSV: {result['upload_path']}")
    print_summary(result["counts"])  # type: ignore[arg-type]
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
