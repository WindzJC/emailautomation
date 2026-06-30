# pyright: reportMissingImports=false, reportMissingModuleSource=false
from __future__ import annotations

import asyncio
import base64
import csv
import io
import json
import os
import re
import secrets
import shutil
import subprocess
import threading
import time
import uuid
from datetime import datetime, timedelta, timezone
from functools import lru_cache
from pathlib import Path
from typing import Callable, List

try:
    from cryptography.exceptions import InvalidSignature
except Exception:  # pragma: no cover - dependency fallback
    class InvalidSignature(Exception):
        pass

from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.primitives.asymmetric import ec
from fastapi import (
    FastAPI,
    File,
    Form,
    Query,
    Request,
    UploadFile,
    WebSocket,
    WebSocketDisconnect,
)
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.middleware.sessions import SessionMiddleware

try:
    import importlib
    defuse_stdlib = importlib.import_module("defusedxml").defuse_stdlib
except Exception:  # pragma: no cover - dependency fallback
    defuse_stdlib = None

import runtime_control
import runtime_audit
import settings
from dashboard_core import (
    SENDGRID_PROFILES,
    append_campaign_run_history,
    build_dashboard_queue_safety_report,
    build_dashboard_snapshot,
    build_profile_message_readiness,
    campaign_history_record,
    active_or_locked_sender_profiles,
    detect_running_preview_profiles,
    detect_running_sender_profiles,
    load_dashboard_run_settings,
    message_preview_output_paths,
    message_preview_path_for_profile,
    profile_expected_pitch_mode,
    save_dashboard_send_cap_per_profile,
)
from important_leads_verify import (
    TRIAGE_MODE_FAST,
    TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH,
    TRIAGE_MODE_STRICT,
    TRIAGE_PATHS_STATE_KEY,
    TRIAGE_STATE_KEY,
    _triage_path_state_labels,
    fast_triage_master_leads,
    important_leads_triage_path_state,
    important_leads_verify_path_state,
    important_leads_verify_status,
    verify_master_leads,
)
from important_leads_workflow import (
    DISPATCH_CAP_ALL,
    DISPATCH_SOURCE_CLEANED,
    DISPATCH_SOURCE_STRICT_VERIFIED,
    DISPATCH_SOURCE_TRIAGED_KEEP,
    MASTER_DISPATCH_STATE_KEY,
    STRICT_VERIFIED_PATH,
    TRIAGED_KEEP_PATH,
    ImportantLeadsCheckError,
    check_master_leads,
    check_warm_research_leads,
    confirm_dispatch_preview,
    confirm_warm_private_jc_preview,
    create_safer_recontact_pool_from_preview,
    generate_warm_email_preview,
    important_leads_path_state,
    important_leads_status,
    load_dispatch_preview,
    preview_dispatch_master_leads,
    validate_dispatch_preview,
    warm_private_jc_lane_status,
)
from lead_ledger import (
    apply_quarantine_review_action,
    connect_lead_ledger,
    ingest_send_outcome_events,
    list_quarantine_review_lead_ids,
    list_quarantine_review_leads,
    load_quarantine_review_lead,
    load_recent_quarantine_review_actions,
)
from leads_workflow import (
    clean_uploaded_leads,
    iso_utc,
    load_state,
    preview_shard_cleaned_leads,
    save_state,
    save_uploaded_csv,
    shard_cleaned_leads,
    shard_status,
    timestamp_slug,
)
from tools.package_campaign_handoff import pack_archive
from tools.rebuild_recipient_queues import active_campaign_manifest_path, build_queue_safety_report
from private_bounce_hygiene import (
    PRIVATE_BOUNCE_MONITOR_ENABLED,
    PRIVATE_BOUNCE_SYNC_INTERVAL_SECONDS,
    run_private_bounce_monitor_cycle,
)
from provider_pacing import mark_recovery_started, provider_pacing_status
from send_shard import CAMPAIGN_TYPE_COLD, PROFILES, is_recontact_cold_campaign, normalize_campaign_type
from sendgrid_hygiene import (
    WEBHOOK_EVENTS_JSONL,
    append_events_jsonl,
    dedupe_webhook_events,
    normalize_webhook_events,
    update_suppressions_from_events,
)

if defuse_stdlib is not None:
    try:
        defuse_stdlib()
    except Exception:
        pass


def _int_env(name: str, default: int, minimum: int = 1) -> int:
    raw = str(os.environ.get(name, "") or "").strip()
    if not raw:
        return max(minimum, int(default))
    try:
        return max(minimum, int(raw))
    except Exception:
        return max(minimum, int(default))


def _lead_ledger_db_path() -> Path:
    return Path(getattr(settings, "LEAD_LEDGER_DB_PATH", settings.STATE_DIR / "lead_ledger.sqlite3"))

STATIC_DIR = settings.STATIC_DIR
SUPPRESSION_CSV = settings.SENDGRID_SUPPRESSIONS_PATH
WEBHOOK_EVENTS_PATH = settings.WEBHOOK_EVENTS_PATH
WEBHOOK_DEDUPE_PATH = settings.WEBHOOK_DEDUPE_PATH
IMPORTANT_LEADS_INPUT = settings.APP_ROOT / "_important" / "leadschecker.csv"
IMPORTANT_LEADS_OUTPUT = settings.APP_ROOT / "_important" / "leads.csv"
IMPORTANT_LEADS_REJECTED = settings.APP_ROOT / "_important" / "leads_rejected.csv"
IMPORTANT_LEADS_RUNS = settings.APP_ROOT / "_important" / "runs"
IMPORTANT_LEADS_CHECK_RUNS = settings.APP_ROOT / "_important" / "check_runs"
IMPORTANT_LEADS_CHECK_JOBS = IMPORTANT_LEADS_CHECK_RUNS / "jobs"
IMPORTANT_LEADS_VERIFY_JOBS = settings.APP_ROOT / "_important" / "verify_jobs"
IMPORTANT_LEADS_DISPATCH_JOBS = settings.APP_ROOT / "_important" / "dispatch_jobs"
IMPORTANT_LEADS_DISPATCH_PREVIEWS = IMPORTANT_LEADS_DISPATCH_JOBS / "previews"
IMPORTANT_LEADS_PASTE_WARNING_ROWS = _int_env("IMPORTANT_LEADS_PASTE_WARNING_ROWS", 250)
IMPORTANT_LEADS_PASTE_MAX_ROWS = max(IMPORTANT_LEADS_PASTE_WARNING_ROWS, _int_env("IMPORTANT_LEADS_PASTE_MAX_ROWS", 1000))
IMPORTANT_LEADS_CHECK_UPLOAD_MAX_BYTES = _int_env("DASHBOARD_CHECK_UPLOAD_MAX_BYTES", 150 * 1024 * 1024)
app = FastAPI(title="Email Automation Live Dashboard")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

_AUTH_PUBLIC_PATHS = {
    "/",
    "/api/health",
    "/api/auth/login",
    "/api/auth/logout",
    "/api/auth/status",
    "/webhooks/sendgrid/events",
}
_AUTH_PUBLIC_PREFIXES = ("/static/",)
_AUTH_PROTECTED_DOC_PATHS = {"/docs", "/redoc", "/openapi.json"}
_AUTH_SESSION_KEY = "dashboard_authenticated"


class DashboardAuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):  # type: ignore[override]
        path = request.url.path
        if request.method == "OPTIONS" or path in _AUTH_PUBLIC_PATHS or any(path.startswith(prefix) for prefix in _AUTH_PUBLIC_PREFIXES):
            return await call_next(request)
        if path in _AUTH_PROTECTED_DOC_PATHS or path.startswith("/api/") or path == "/ws":
            if not settings.DASHBOARD_AUTH_PASSWORD:
                return await call_next(request)
            if not bool(request.session.get(_AUTH_SESSION_KEY)):
                return JSONResponse(
                    {"ok": False, "message": "Authentication required."},
                    status_code=401,
                )
        return await call_next(request)


app.add_middleware(DashboardAuthMiddleware)
app.add_middleware(
    SessionMiddleware,
    secret_key=settings.DASHBOARD_SESSION_SECRET,
    session_cookie=settings.DASHBOARD_AUTH_COOKIE_NAME,
    same_site="lax",
    https_only=False,
)

SENDGRID_EVENT_PUBLIC_KEY = os.environ.get("SENDGRID_EVENT_PUBLIC_KEY", "").strip()
SENDGRID_SIG_HEADER = "X-Twilio-Email-Event-Webhook-Signature"
SENDGRID_TS_HEADER = "X-Twilio-Email-Event-Webhook-Timestamp"
PRIVATE_BOUNCE_PROFILE = "private_jc"
AUTOMATION_LOOP_SECONDS = max(15, min(60, max(30, int(PRIVATE_BOUNCE_SYNC_INTERVAL_SECONDS or 120)) // 2))
DASHBOARD_AUTO_START_STATE_PATH = settings.STATE_DIR / "dashboard_auto_start_state.json"
DASHBOARD_TIMER_STATE_PATH = settings.STATE_DIR / "dashboard_timer_state.json"
AUTO_START_RETRY_MINUTES = 10
_PARSER_EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)


class SendCapPayload(BaseModel):
    send_cap_per_profile: int = Field(..., ge=1, le=100000)


class ColumnMappingPayload(BaseModel):
    email: str = ""
    first_name: str = ""
    book_title: str = ""


class CleanLeadsPayload(BaseModel):
    upload_filename: str
    mapping: ColumnMappingPayload | None = None
    remove_invalid_emails: bool = True
    dedupe_by_email: bool = True
    remove_suppressed: bool = True
    drop_role_emails: bool = False
    exclude_domains: List[str] = Field(default_factory=list)


class ShardLeadsPayload(BaseModel):
    cleaned_filename: str
    shard_count: int = Field(default=5, ge=1, le=5)
    strategy: str = Field(default="domain_balanced")


class ImportantLeadPathsPayload(BaseModel):
    input_path: str = ""
    output_path: str = ""
    rejected_path: str = ""
    dispatch_source_mode: str = DISPATCH_SOURCE_TRIAGED_KEEP
    input_text: str = ""
    intake_mode: str = "standard"


class ImportantLeadVerifyPayload(BaseModel):
    input_path: str = ""
    verified_path: str = ""
    rejected_path: str = ""
    quarantine_path: str = ""
    mode: str = TRIAGE_MODE_FAST


def _normalize_intake_mode(value: object) -> str:
    text = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
    if text in {"manual_author_research", "author_research", "manual"}:
        return TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH
    return "STANDARD"


def _triage_mode_label(mode: object) -> str:
    normalized = str(mode or "").strip().upper()
    if normalized == TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH:
        return "Manual Author Research"
    if normalized == TRIAGE_MODE_STRICT:
        return "Strict Public Proof"
    return "Fast Triage"


class ImportantLeadDispatchPayload(BaseModel):
    input_path: str = ""
    output_path: str = ""
    rejected_path: str = ""
    dispatch_source_mode: str = DISPATCH_SOURCE_TRIAGED_KEEP
    dispatch_cap: str = DISPATCH_CAP_ALL
    campaign_type: str = CAMPAIGN_TYPE_COLD
    preview_id: str = ""
    recontact_recency_override: bool = False


class QuarantineReviewActionPayload(BaseModel):
    lead_ids: list[str] = Field(default_factory=list)
    excluded_lead_ids: list[str] = Field(default_factory=list)
    action: str = ""
    operator_note: str = ""
    select_all_filtered: bool = False
    reason_code: str = ""
    stage: str = ""
    status: str = "QUARANTINE"
    sort: str = "score_desc"


class DashboardAuthPayload(BaseModel):
    username: str = ""
    password: str = ""


def _profile_runtime_active(profile_name: str) -> bool:
    try:
        snapshots = runtime_control.list_sender_snapshots(tail_lines=8)
    except Exception:
        return False
    active_states = {"starting", "running", "cooldown", "sleeping"}
    for snapshot in snapshots:
        if getattr(snapshot, "name", "") != profile_name:
            continue
        if getattr(snapshot, "tmux_dead", False):
            return False
        return str(getattr(snapshot, "runtime_state", "") or "") in active_states
    return False


def _resolve_dashboard_csv_path(raw_value: str, default_path: Path) -> Path:
    candidate = str(raw_value or "").strip()
    if not candidate:
        return default_path
    if os.name != "nt" and re.match(r"^[A-Za-z]:[\\/]", candidate):
        raise ValueError("Leads paths must stay inside this workspace.")
    path = Path(candidate)
    if not path.is_absolute():
        path = settings.APP_ROOT / path
    resolved_root = settings.APP_ROOT.resolve()
    resolved_path = path.resolve(strict=False)
    if resolved_path.suffix.lower() != ".csv":
        raise ValueError("Leads paths must point to .csv files.")
    if resolved_path != resolved_root and resolved_root not in resolved_path.parents:
        raise ValueError("Leads paths must stay inside this workspace.")
    return resolved_path


def _resolve_dashboard_csv_path_or_default(raw_value: str, default_path: Path) -> Path:
    try:
        return _resolve_dashboard_csv_path(raw_value, default_path)
    except ValueError:
        return default_path.resolve(strict=False)


def _important_path_labels_for_state(input_path: Path, output_path: Path, rejected_path: Path) -> dict[str, str]:
    root = settings.APP_ROOT.resolve()

    def label(path: Path) -> str:
        try:
            return str(path.resolve().relative_to(root))
        except Exception:
            return str(path)

    return {
        "input_path": label(input_path),
        "output_path": label(output_path),
        "rejected_path": label(rejected_path),
    }


def _important_verify_path_labels_for_state(input_path: Path, verified_path: Path, rejected_path: Path, quarantine_path: Path) -> dict[str, str]:
    root = settings.APP_ROOT.resolve()

    def label(path: Path) -> str:
        try:
            return str(path.resolve().relative_to(root))
        except Exception:
            return str(path)

    return {
        "input_path": label(input_path),
        "verified_path": label(verified_path),
        "rejected_path": label(rejected_path),
        "quarantine_path": label(quarantine_path),
    }


def _important_triage_path_labels_for_state(input_path: Path, keep_path: Path, rejected_path: Path, quarantine_path: Path) -> dict[str, str]:
    root = settings.APP_ROOT.resolve()

    def label(path: Path) -> str:
        try:
            return str(path.resolve().relative_to(root))
        except Exception:
            return str(path)

    return {
        "input_path": label(input_path),
        "keep_path": label(keep_path),
        "rejected_path": label(rejected_path),
        "quarantine_path": label(quarantine_path),
    }


def _important_dispatch_source_labels_for_state(mode: str) -> dict[str, str]:
    normalized = str(mode or "").strip().lower() or DISPATCH_SOURCE_TRIAGED_KEEP
    aliases = {
        "verified": DISPATCH_SOURCE_TRIAGED_KEEP,
        "fast_triage": DISPATCH_SOURCE_TRIAGED_KEEP,
        "fast_triage_keep": DISPATCH_SOURCE_TRIAGED_KEEP,
        "strict": DISPATCH_SOURCE_STRICT_VERIFIED,
        "strict_public_proof": DISPATCH_SOURCE_STRICT_VERIFIED,
    }
    normalized = aliases.get(normalized, normalized)
    if normalized not in {DISPATCH_SOURCE_TRIAGED_KEEP, DISPATCH_SOURCE_STRICT_VERIFIED, DISPATCH_SOURCE_CLEANED}:
        normalized = DISPATCH_SOURCE_TRIAGED_KEEP
    return {"dispatch_source_mode": normalized}


def _dashboard_path_label(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(settings.APP_ROOT.resolve()))
    except Exception:
        return str(path)


def _important_check_batch_policy() -> dict[str, object]:
    return {
        "paste_mode": "small_manual_only",
        "paste_warning_rows": IMPORTANT_LEADS_PASTE_WARNING_ROWS,
        "paste_max_rows": IMPORTANT_LEADS_PASTE_MAX_ROWS,
        "upload_required_rows": IMPORTANT_LEADS_PASTE_MAX_ROWS,
        "upload_recommended_rows": IMPORTANT_LEADS_PASTE_WARNING_ROWS,
    }


def _important_check_job_path(job_id: str) -> Path:
    return IMPORTANT_LEADS_CHECK_JOBS / f"{job_id}.json"


def _load_important_check_job(job_id: str) -> dict[str, object]:
    path = _important_check_job_path(job_id)
    if not path.exists():
        raise FileNotFoundError(f"Check job not found: {job_id}")
    raw = json.loads(path.read_text(encoding="utf-8"))
    return raw if isinstance(raw, dict) else {}


def _save_important_check_job(job: dict[str, object]) -> None:
    job_id = str(job.get("job_id") or "").strip()
    if not job_id:
        raise ValueError("Missing check job id.")
    IMPORTANT_LEADS_CHECK_JOBS.mkdir(parents=True, exist_ok=True)
    payload = dict(job)
    payload["updated_at_utc"] = iso_utc()
    write_path = _important_check_job_path(job_id)
    tmp_path = write_path.with_suffix(f".{os.getpid()}.tmp")
    tmp_path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
    tmp_path.replace(write_path)
    settings.secure_private_file(write_path)


def _important_check_job_with_progress(job: dict[str, object]) -> dict[str, object]:
    payload = dict(job)
    total_rows = max(0, int(payload.get("total_input_rows") or 0))
    processed_rows = max(0, int(payload.get("processed_rows") or 0))
    remaining_rows = max(0, int(payload.get("remaining_rows") or max(0, total_rows - processed_rows)))
    if "progress_percent" not in payload or payload.get("progress_percent") in {"", None}:
        if total_rows > 0:
            payload["progress_percent"] = round(min(100.0, max(0.0, (processed_rows / total_rows) * 100)), 1)
        elif str(payload.get("status") or "") == "completed":
            payload["progress_percent"] = 100
        else:
            payload["progress_percent"] = 0
    payload["processed_rows"] = processed_rows
    payload["remaining_rows"] = remaining_rows
    if payload.get("source_sheet") and not payload.get("current_sheet"):
        payload["current_sheet"] = payload.get("source_sheet")
    return payload


def _find_active_important_check_job() -> dict[str, object] | None:
    if not IMPORTANT_LEADS_CHECK_JOBS.exists():
        return None
    active_statuses = {"queued", "running", "checking", "auto_triage_running"}
    candidates: list[tuple[float, dict[str, object]]] = []
    for path in IMPORTANT_LEADS_CHECK_JOBS.glob("*.json"):
        try:
            job = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(job, dict):
            continue
        status = str(job.get("status") or job.get("stage") or "").strip().lower()
        if status not in active_statuses:
            continue
        try:
            sort_key = path.stat().st_mtime
        except Exception:
            sort_key = 0.0
        candidates.append((sort_key, _important_check_job_with_progress(job)))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def _parse_iso_timestamp(value: object) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).astimezone(timezone.utc)
    except Exception:
        return None


def _check_job_created_sort_key(job: dict[str, object], path: Path | None = None) -> float:
    created = _parse_iso_timestamp(job.get("created_at_utc"))
    if created:
        return created.timestamp()
    if path is not None:
        try:
            return path.stat().st_mtime
        except Exception:
            return 0.0
    return 0.0


def _has_newer_important_check_job(job: dict[str, object]) -> bool:
    if not IMPORTANT_LEADS_CHECK_JOBS.exists():
        return False
    job_id = str(job.get("job_id") or "").strip()
    current_key = _check_job_created_sort_key(job, _important_check_job_path(job_id) if job_id else None)
    for path in IMPORTANT_LEADS_CHECK_JOBS.glob("*.json"):
        try:
            other = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(other, dict):
            continue
        other_id = str(other.get("job_id") or path.stem).strip()
        if other_id == job_id:
            continue
        if _check_job_created_sort_key(other, path) > current_key:
            return True
    return False


def _auto_triage_already_running(exclude_check_job_id: str = "") -> bool:
    active_verify = _find_active_dashboard_job(IMPORTANT_LEADS_VERIFY_JOBS)
    if active_verify and str(active_verify.get("mode") or "").upper() in {TRIAGE_MODE_FAST, TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH}:
        return True
    if not IMPORTANT_LEADS_CHECK_JOBS.exists():
        return False
    for path in IMPORTANT_LEADS_CHECK_JOBS.glob("*.json"):
        try:
            job = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(job, dict):
            continue
        if str(job.get("job_id") or path.stem).strip() == exclude_check_job_id:
            continue
        if str(job.get("auto_triage_status") or "").strip().lower() == "running":
            return True
    return False


def _file_signature(path: Path) -> tuple[int, int]:
    stat = path.stat()
    return stat.st_mtime_ns, stat.st_size


def _copy_csv_atomic(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = destination.with_suffix(f"{destination.suffix}.{os.getpid()}.tmp")
    shutil.copyfile(source, tmp_path)
    tmp_path.replace(destination)
    settings.secure_private_file(destination)


def _promote_auto_triage_outputs(
    *,
    staged_keep_path: Path,
    staged_rejected_path: Path,
    staged_quarantine_path: Path,
    keep_path: Path,
    rejected_path: Path,
    quarantine_path: Path,
) -> None:
    _copy_csv_atomic(staged_keep_path, keep_path)
    _copy_csv_atomic(staged_rejected_path, rejected_path)
    _copy_csv_atomic(staged_quarantine_path, quarantine_path)


def _count_book_title_rows(path: Path) -> dict[str, int]:
    if not path.exists():
        return {"rows_with_booktitle": 0, "rows_without_booktitle": 0}
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        book_title_header = next((field for field in fieldnames if str(field or "").strip().lower() == "booktitle"), "")
        with_title = 0
        without_title = 0
        for row in reader:
            if book_title_header and str(row.get(book_title_header) or "").strip():
                with_title += 1
            else:
                without_title += 1
    return {"rows_with_booktitle": with_title, "rows_without_booktitle": without_title}


def _book_title_fallback_readiness() -> dict[str, object]:
    from send_shard import PROFILES, PITCHES, book_title_fallback_supported, template_requires_book_title

    profile_names = [
        "private_jc",
        "sendgrid_annette",
        "sendgrid_jordan",
        "sendgrid_jodi",
        "sendgrid_alison",
        "sendgrid_fiorela",
    ]
    profiles: dict[str, dict[str, object]] = {}
    all_ok = True
    for profile_name in profile_names:
        cfg = PROFILES.get(profile_name) or {}
        pitch_key = str(cfg.get("pitch") or "")
        pitch = PITCHES.get(pitch_key) or {}
        subject = str(pitch.get("subject") or "")
        body = str(pitch.get("body") or "")
        subject_fallback = str(pitch.get("subject_fallback") or "")
        body_fallback = str(pitch.get("body_fallback") or "")
        requires_booktitle = template_requires_book_title(subject, body)
        fallback_error = ""
        try:
            fallback_supported = book_title_fallback_supported(subject, body, subject_fallback, body_fallback=body_fallback) if requires_booktitle else True
        except ValueError as exc:
            fallback_supported = False
            fallback_error = str(exc)
        strict_required = bool(pitch.get("require_book_title"))
        ok = (not requires_booktitle) or (fallback_supported and not strict_required)
        profiles[profile_name] = {
            "pitch": pitch_key,
            "requires_booktitle": requires_booktitle,
            "fallback_supported": fallback_supported,
            "strict_booktitle_required": strict_required,
            "ready_for_missing_booktitle": ok,
        }
        if fallback_error:
            profiles[profile_name]["fallback_error"] = fallback_error
        all_ok = all_ok and ok
    return {"fallback_capable": all_ok, "profiles": profiles}


def _active_sender_state_summary() -> dict[str, object]:
    active_profiles = runtime_control.list_active_sender_snapshots(tail_lines=12)
    states = {str(item.name): str(item.runtime_state) for item in active_profiles}
    return {
        "active_sender_count": len(states),
        "active_profiles": list(states.keys()),
        "active_sender_states": states,
        "any_sender_running": bool(states),
    }


def _auto_dispatch_preview_summary(
    *,
    preview: dict[str, object],
    triage_report: dict[str, object],
    checked_path: Path,
    keep_path: Path,
    rejected_path: Path,
) -> dict[str, object]:
    booktitle_counts = _count_book_title_rows(keep_path)
    fallback = _book_title_fallback_readiness()
    sender_state = _active_sender_state_summary()
    try:
        queue_safety = build_queue_safety_report(
            intended_source_path=keep_path,
            checked_path=checked_path,
            triaged_keep_path=keep_path,
            triaged_reject_path=rejected_path,
        )
    except Exception as exc:
        queue_safety = {
            "safe": False,
            "unsafe_reasons": ["QUEUE_SAFETY_CHECK_FAILED"],
            "message": str(exc),
        }
    planned_counts = dict(preview.get("rows_written_per_queue") or {})
    return {
        "preview_id": str(preview.get("preview_id") or ""),
        "preview_path": str(preview.get("preview_path") or ""),
        "status": "previewed",
        "generated_at_utc": iso_utc(),
        "campaign_type": str(preview.get("campaign_type") or CAMPAIGN_TYPE_COLD),
        "dispatch_source_path": str(preview.get("dispatch_source_path") or keep_path),
        "dispatch_source_mode": str(preview.get("dispatch_source_mode") or DISPATCH_SOURCE_TRIAGED_KEEP),
        "dispatch_source_row_count": int(preview.get("dispatch_source_row_count") or triage_report.get("keep_count") or 0),
        "dispatch_eligible_row_count": int(preview.get("dispatch_eligible_row_count") or triage_report.get("keep_count") or 0),
        "total_keep_rows": int(triage_report.get("keep_count") or preview.get("dispatch_source_row_count") or 0),
        "rejected_rows": int(triage_report.get("reject_count") or 0),
        "quarantine_rows": int(triage_report.get("quarantine_count") or 0),
        **booktitle_counts,
        "fallback_capable": bool(fallback.get("fallback_capable")),
        "fallback_readiness": fallback,
        "suppression_unsubscribe_skip_count": int(preview.get("suppressed_skipped") or preview.get("skipped_suppressed") or 0),
        "suppression_summary": dict(preview.get("suppression_summary") or {}),
        "per_profile_planned_counts": planned_counts,
        "queue_safety": queue_safety,
        **sender_state,
        "manual_rebuild_allowed": not bool(sender_state.get("any_sender_running")),
        "manual_rebuild_required": True,
        "manual_start_required": True,
        "auto_rebuild_performed": False,
        "auto_dispatch_performed": False,
        "auto_start_performed": False,
        "message": (
            "Dispatch preview ready. Manual queue rebuild/confirm is required."
            if not bool(sender_state.get("any_sender_running"))
            else "Dispatch preview ready, but active senders must be stopped before queue rebuild/confirm."
        ),
    }


def _run_auto_dispatch_preview_after_triage(
    *,
    job: dict[str, object],
    triage_report: dict[str, object],
    master_path: Path,
    keep_path: Path,
    rejected_path: Path,
    quarantine_path: Path,
    preview_dir: Path | None = None,
) -> dict[str, object]:
    job["auto_dispatch_preview_status"] = "running"
    job["auto_dispatch_preview_started_at_utc"] = iso_utc()
    job["message"] = "Auto triage complete. Building dispatch preview."
    _save_important_check_job(job)
    try:
        preview = preview_dispatch_master_leads(
            master_path=master_path,
            rejected_path=Path(str(job.get("rejected_path") or rejected_path)),
            verified_path=STRICT_VERIFIED_PATH,
            triaged_keep_path=keep_path,
            dispatch_source_mode=DISPATCH_SOURCE_TRIAGED_KEEP,
            dispatch_cap=DISPATCH_CAP_ALL,
            jc_queue_path=settings.SHARDS_DIR / "recipients_private_jc.csv",
            sendgrid_queue_paths=[
                settings.SHARDS_DIR / "recipients_sendgrid_1.csv",
                settings.SHARDS_DIR / "recipients_sendgrid_2.csv",
                settings.SHARDS_DIR / "recipients_sendgrid_3.csv",
                settings.SHARDS_DIR / "recipients_sendgrid_4.csv",
                settings.SHARDS_DIR / "recipients_sendgrid_5.csv",
            ],
            jc_log_path=settings.LOGS_DIR / "private_jc_log.csv",
            sendgrid_log_paths=[
                settings.LOGS_DIR / "sendgrid_annette_log.csv",
                settings.LOGS_DIR / "sendgrid_jordan_log.csv",
                settings.LOGS_DIR / "sendgrid_jodi_log.csv",
                settings.LOGS_DIR / "sendgrid_alison_log.csv",
                settings.LOGS_DIR / "sendgrid_fiorela_log.csv",
                settings.LOGS_DIR / "sendgrid_domain_log.csv",
            ],
            preview_dir=preview_dir or IMPORTANT_LEADS_DISPATCH_PREVIEWS,
        )
        summary = _auto_dispatch_preview_summary(
            preview=preview,
            triage_report=triage_report,
            checked_path=master_path,
            keep_path=keep_path,
            rejected_path=rejected_path,
        )
        job["auto_dispatch_preview_status"] = "completed"
        job["auto_dispatch_preview_completed_at_utc"] = iso_utc()
        job["auto_dispatch_preview_id"] = summary["preview_id"]
        job["auto_dispatch_preview"] = summary
        save_state(latest_auto_dispatch_preview=summary)
    except Exception as exc:
        job["auto_dispatch_preview_status"] = "failed"
        job["auto_dispatch_preview_error"] = str(exc)
        job["auto_dispatch_preview_completed_at_utc"] = iso_utc()
        job["message"] = f"Auto triage complete. Dispatch preview failed: {exc}"
    return job


def _run_auto_fast_triage_after_check(job: dict[str, object]) -> dict[str, object]:
    job_id = str(job.get("job_id") or "").strip()
    if not job_id:
        return job
    if str(job.get("source_mode") or "").strip() != "uploaded_file":
        job["auto_triage_status"] = "skipped"
        job["auto_triage_skip_reason"] = "not_upload_check"
        return job
    if str(job.get("status") or "").strip().lower() in {"failed", "canceled", "cancelled"} or job.get("cancel_requested"):
        job["auto_triage_status"] = "skipped"
        job["auto_triage_skip_reason"] = "check_not_successful"
        return job
    existing_status = str(job.get("auto_triage_status") or "").strip().lower()
    if existing_status in {"running", "completed", "failed", "skipped"}:
        return job
    check_status = str(job.get("status") or "").strip().lower()
    if check_status not in {"completed", "done"}:
        job["auto_triage_status"] = "skipped"
        job["auto_triage_skip_reason"] = "check_still_running"
        job["message"] = "Check still running. Auto triage will wait for a finalized leads.csv."
        return job

    output_path = Path(str(job.get("output_path") or ""))
    rejected_path = Path(str(job.get("rejected_path") or ""))
    if not output_path.exists() or not rejected_path.exists():
        job["auto_triage_status"] = "skipped"
        job["auto_triage_skip_reason"] = "fresh_check_outputs_missing"
        job["message"] = "Check output is not ready. Auto triage requires finalized leads.csv and leads_rejected.csv."
        return job
    try:
        checked_output_rows = _count_csv_rows(output_path)
    except Exception as exc:
        job["auto_triage_status"] = "skipped"
        job["auto_triage_skip_reason"] = "fresh_check_output_unreadable"
        job["message"] = f"Check output is not readable yet. Auto triage skipped: {exc}"
        return job
    if checked_output_rows <= 0:
        job["auto_triage_status"] = "skipped"
        job["auto_triage_skip_reason"] = "fresh_check_output_empty"
        job["message"] = "Check output has no accepted rows. Auto triage was not started."
        return job
    if _has_newer_important_check_job(job):
        job["auto_triage_status"] = "skipped"
        job["auto_triage_skip_reason"] = "newer_check_job_started"
        return job
    if _auto_triage_already_running(exclude_check_job_id=job_id):
        job["auto_triage_status"] = "skipped"
        job["auto_triage_skip_reason"] = "triage_already_running"
        return job

    checked_signature = _file_signature(output_path)
    staged_dir = Path(str(job.get("staged_run_dir") or "")) if job.get("staged_run_dir") else IMPORTANT_LEADS_RUNS / job_id
    if not staged_dir.is_absolute():
        staged_dir = settings.APP_ROOT / staged_dir
    staged_dir.mkdir(parents=True, exist_ok=True)
    staged_keep_path = staged_dir / "leads_triaged_keep.csv"
    staged_rejected_path = staged_dir / "leads_triaged_reject.csv"
    staged_quarantine_path = staged_dir / "leads_triaged_quarantine.csv"
    staged_preview_dir = staged_dir / "dispatch_previews"
    keep_path = staged_keep_path
    triage_rejected_path = staged_rejected_path
    quarantine_path = staged_quarantine_path

    job["status"] = "auto_triage_running"
    job["stage"] = "auto_triage"
    job["phase"] = "auto_triage"
    job["auto_triage_status"] = "running"
    job["auto_triage_source_check_job_id"] = job_id
    job["auto_triage_started_at_utc"] = iso_utc()
    job["auto_triage_keep_path"] = str(keep_path)
    job["auto_triage_rejected_path"] = str(triage_rejected_path)
    job["auto_triage_quarantine_path"] = str(quarantine_path)
    job["staged_run_dir"] = str(staged_dir)
    job["message"] = "Check complete. Auto triage running."
    _save_important_check_job(job)

    started_at = time.monotonic()
    last_progress_save_at = 0.0

    def save_auto_triage_progress(processed_rows: int, total_rows: int) -> None:
        nonlocal job, last_progress_save_at
        now = time.monotonic()
        if processed_rows < total_rows and now - last_progress_save_at < 0.75:
            return
        last_progress_save_at = now
        total = max(0, int(total_rows or 0))
        processed = min(total, max(0, int(processed_rows or 0)))
        remaining = max(0, total - processed)
        elapsed = max(0.001, now - started_at)
        rate = processed / elapsed if processed > 0 else 0.0
        job["auto_triage_total_rows"] = total
        job["auto_triage_processed_rows"] = processed
        job["auto_triage_remaining_rows"] = remaining
        job["auto_triage_progress_percent"] = round(min(100.0, max(0.0, (processed / total) * 100)), 1) if total else 0
        job["auto_triage_eta_seconds"] = int(remaining / rate) if rate > 0 and remaining > 0 else 0
        _save_important_check_job(job)

    def should_cancel_auto_triage() -> bool:
        try:
            latest_job = _load_important_check_job(job_id)
            if latest_job.get("cancel_requested"):
                return True
        except Exception:
            pass
        return _has_newer_important_check_job(job)

    try:
        triage_mode = TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH if _normalize_intake_mode(job.get("intake_mode")) == TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH else TRIAGE_MODE_FAST
        book_title_fallback_supported = (
            bool(_book_title_fallback_readiness().get("fallback_capable"))
            if triage_mode == TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH
            else False
        )
        report = fast_triage_master_leads(
            input_path=output_path,
            keep_path=staged_keep_path,
            rejected_path=staged_rejected_path,
            quarantine_path=staged_quarantine_path,
            persist_state=False,
            progress_callback=save_auto_triage_progress,
            should_cancel=should_cancel_auto_triage,
            mode=triage_mode,
            book_title_fallback_supported=book_title_fallback_supported,
        )
        if bool(report.get("canceled")) or should_cancel_auto_triage():
            job["status"] = "completed"
            job["stage"] = "done"
            job["phase"] = "done"
            job["auto_triage_status"] = "skipped"
            job["auto_triage_skip_reason"] = "newer_check_job_started" if _has_newer_important_check_job(job) else "canceled"
            job["message"] = "Check complete. Auto triage skipped before publishing outputs."
            return job
        if _file_signature(output_path) != checked_signature:
            job["status"] = "completed"
            job["stage"] = "done"
            job["phase"] = "done"
            job["auto_triage_status"] = "skipped"
            job["auto_triage_skip_reason"] = "checked_output_changed"
            job["message"] = "Check complete. Auto triage skipped because checked output changed."
            return job

        final_report = dict(report)
        final_report["input_label"] = str(output_path)
        final_report["verified_label"] = str(keep_path)
        final_report["rejected_label"] = str(triage_rejected_path)
        final_report["quarantine_label"] = str(quarantine_path)
        final_report["auto_triage_source_check_job_id"] = job_id
        final_report["intake_mode"] = triage_mode
        final_report["intake_mode_label"] = "Manual Author Research" if triage_mode == TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH else "Standard"
        if triage_mode == TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH:
            final_report["book_title_fallback_supported"] = book_title_fallback_supported
        save_state(
            **{
                TRIAGE_PATHS_STATE_KEY: _triage_path_state_labels(
                    output_path,
                    keep_path,
                    triage_rejected_path,
                    quarantine_path,
                ),
                TRIAGE_STATE_KEY: final_report,
            }
        )
        job["status"] = "completed"
        job["stage"] = "done"
        job["phase"] = "done"
        job["auto_triage_status"] = "completed"
        job["auto_triage_completed_at_utc"] = iso_utc()
        job["auto_triage_report"] = final_report
        job["auto_triage_total_rows"] = int(final_report.get("total_input_rows") or final_report.get("input_rows") or 0)
        job["auto_triage_processed_rows"] = int(final_report.get("processed_rows") or 0)
        job["auto_triage_remaining_rows"] = 0
        job["auto_triage_progress_percent"] = 100
        job = _run_auto_dispatch_preview_after_triage(
            job=job,
            triage_report=final_report,
            master_path=output_path,
            keep_path=keep_path,
            rejected_path=triage_rejected_path,
            quarantine_path=quarantine_path,
            preview_dir=staged_preview_dir,
        )
        job["message"] = (
            f"Check complete. Auto triage complete: KEEP {int(final_report.get('keep_count') or 0)}, "
            f"REJECT {int(final_report.get('reject_count') or 0)}, "
            f"QUARANTINE {int(final_report.get('quarantine_count') or 0)}. "
            + (
                f"Dispatch preview failed: {job.get('auto_dispatch_preview_error')}"
                if str(job.get("auto_dispatch_preview_status") or "").lower() == "failed"
                else f"Dispatch preview {str(job.get('auto_dispatch_preview_status') or 'not_started')}."
            )
        )
    except Exception as exc:
        job["status"] = "triage_failed"
        job["stage"] = "triage_failed"
        job["phase"] = "triage_failed"
        job["auto_triage_status"] = "failed"
        job["auto_triage_error"] = str(exc)
        job["auto_triage_completed_at_utc"] = iso_utc()
        job["message"] = f"Check complete. Auto triage failed: {exc}"
    return job


def _job_path(directory: Path, job_id: str) -> Path:
    return directory / f"{job_id}.json"


def _load_dashboard_job(directory: Path, job_id: str, label: str) -> dict[str, object]:
    path = _job_path(directory, job_id)
    if not path.exists():
        raise FileNotFoundError(f"{label} job not found: {job_id}")
    raw = json.loads(path.read_text(encoding="utf-8"))
    return raw if isinstance(raw, dict) else {}


def _save_dashboard_job(directory: Path, job: dict[str, object]) -> None:
    job_id = str(job.get("job_id") or "").strip()
    if not job_id:
        raise ValueError("Missing job id.")
    directory.mkdir(parents=True, exist_ok=True)
    payload = dict(job)
    payload["updated_at_utc"] = iso_utc()
    write_path = _job_path(directory, job_id)
    tmp_path = write_path.with_suffix(f".{os.getpid()}.tmp")
    tmp_path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
    tmp_path.replace(write_path)
    settings.secure_private_file(write_path)


def _job_progress_payload(job: dict[str, object]) -> dict[str, object]:
    payload = dict(job)
    total_rows = max(0, int(payload.get("total_rows") or payload.get("total_input_rows") or 0))
    processed_rows = max(0, int(payload.get("processed_rows") or 0))
    remaining_rows = max(0, int(payload.get("remaining_rows") or max(0, total_rows - processed_rows)))
    if "progress_percent" not in payload or payload.get("progress_percent") in {"", None}:
        if total_rows > 0:
            payload["progress_percent"] = round(min(100.0, max(0.0, (processed_rows / total_rows) * 100)), 1)
        elif str(payload.get("status") or "") == "completed":
            payload["progress_percent"] = 100
        else:
            payload["progress_percent"] = 0
    payload["total_rows"] = total_rows
    payload["processed_rows"] = processed_rows
    payload["remaining_rows"] = remaining_rows
    return payload


def _create_pre_dispatch_archive(job_id: str) -> dict[str, object]:
    safe_job_id = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(job_id or "dispatch")).strip("_") or "dispatch"
    archive_dir = settings.BACKUPS_DIR / "pre_dispatch_handoffs"
    archive_path = archive_dir / f"{safe_job_id}_{timestamp_slug()}.tar.gz"
    manifest = pack_archive(archive_path, include_check_history=False)
    return {
        "archive_path": str(archive_path),
        "archive_name": archive_path.name,
        "file_count": int(manifest.get("file_count") or 0),
        "created_at_utc": str(manifest.get("created_at_utc") or ""),
        "queue_counts": dict(manifest.get("queue_counts") or {}),
        "state_summaries": dict(manifest.get("state_summaries") or {}),
    }


def _find_active_dashboard_job(directory: Path) -> dict[str, object] | None:
    if not directory.exists():
        return None
    active_statuses = {"queued", "running", "checking", "verifying", "dispatching", "auto_triage_running"}
    candidates: list[tuple[float, dict[str, object]]] = []
    for path in directory.glob("*.json"):
        try:
            job = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(job, dict):
            continue
        status = str(job.get("status") or job.get("stage") or "").strip().lower()
        if status not in active_statuses:
            continue
        try:
            sort_key = path.stat().st_mtime
        except Exception:
            sort_key = 0.0
        candidates.append((sort_key, _job_progress_payload(job)))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def _count_csv_rows(path: Path) -> int:
    if not path.exists():
        return 0
    count = 0
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if any(str(value or "").strip() for value in row.values()):
                count += 1
    return count


def _dispatch_source_readiness_block(
    dispatch_source_mode: str,
    source_path: Path,
    *,
    source_resolution: str = "",
) -> dict[str, str] | None:
    mode = str(dispatch_source_mode or "").strip().lower() or DISPATCH_SOURCE_TRIAGED_KEEP
    label = _dashboard_path_label(source_path)
    if mode == DISPATCH_SOURCE_TRIAGED_KEEP:
        if source_resolution == "latest_completed_staged_run" and (not source_path.exists() or _count_csv_rows(source_path) <= 0):
            return {
                "error": "triage_not_ready",
                "message": "Current staged Fast Triage Keep is empty. Run Check Leads / Fast Triage first.",
                "retry_action": "Run Check Leads / Fast Triage, then retry Preview Dispatch.",
            }
        if not source_path.exists():
            return {
                "error": "triage_not_ready",
                "message": f"Triage not ready: leads_triaged_keep.csv is missing at {label}. Run Fast Triage after Check Leads completes.",
                "retry_action": "Run Fast Triage, then retry Preview Dispatch.",
            }
        if _count_csv_rows(source_path) <= 0:
            return {
                "error": "triage_not_ready",
                "message": "Triage not ready: leads_triaged_keep.csv has no Keep rows. Review/Quarantine rows are not dispatched automatically.",
                "retry_action": "Review the triage results or rerun Fast Triage after fixing the source, then retry Preview Dispatch.",
            }
        return None
    if mode == DISPATCH_SOURCE_STRICT_VERIFIED:
        if not source_path.exists():
            return {
                "error": "strict_verified_not_ready",
                "message": f"Strict verified source is missing at {label}. Run Verify Leads before Preview Dispatch.",
                "retry_action": "Run Verify Leads, then retry Preview Dispatch.",
            }
        if _count_csv_rows(source_path) <= 0:
            return {
                "error": "strict_verified_not_ready",
                "message": "Strict verified source has no eligible rows. Preview Dispatch cannot run.",
                "retry_action": "Review Verify Leads output, then retry Preview Dispatch.",
            }
        return None
    if not source_path.exists():
        return {
            "error": "checked_output_not_ready",
            "message": f"Check still running or output missing: leads.csv is not available at {label}.",
            "retry_action": "Wait for Check Leads to complete, then retry Preview Dispatch.",
        }
    if _count_csv_rows(source_path) <= 0:
        return {
            "error": "checked_output_not_ready",
            "message": "Check output has no accepted rows. Preview Dispatch cannot run.",
            "retry_action": "Rerun Check Leads with a valid source, then retry Preview Dispatch.",
        }
    return None


def _normalize_uploaded_check_file(filename: str, content: bytes) -> tuple[str, str, str]:
    extension = Path(str(filename or "").strip()).suffix.lower()
    if extension == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        return ".csv", text if text.endswith("\n") else f"{text}\n", ""
    if extension == ".xlsx":
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Failed to read XLSX upload: {exc}") from exc
        try:
            source_sheet = str(workbook.sheetnames[0]) if workbook.sheetnames else ""
            worksheet = workbook[source_sheet] if source_sheet else None
            if worksheet is None:
                raise ValueError("Uploaded XLSX file has no worksheets.")
            rows = []
            for row in worksheet.iter_rows(values_only=True):
                values = ["" if cell is None else str(cell) for cell in row]
                if any(str(value or "").strip() for value in values):
                    rows.append(values)
            if not rows:
                raise ValueError("Uploaded XLSX file has no usable rows.")
            buffer = io.StringIO()
            writer = csv.writer(buffer, lineterminator="\n")
            for row in rows:
                writer.writerow(row)
            text = buffer.getvalue()
            if text and not text.endswith("\n"):
                text += "\n"
            return ".xlsx", text, source_sheet
        finally:
            workbook.close()
    raise ValueError(f"Unsupported upload file type: {extension or '[none]'}")


def _execute_important_check(
    *,
    input_path: Path,
    output_path: Path,
    rejected_path: Path,
    effective_input_path: Path,
    intake_mode: str = "STANDARD",
    progress_callback=None,
) -> dict[str, object]:
    save_state(important_leads_paths=_important_path_labels_for_state(input_path, output_path, rejected_path))
    report = check_master_leads(
        input_path=effective_input_path,
        output_path=output_path,
        rejected_path=rejected_path,
        progress_callback=progress_callback,
    )
    report["intake_mode"] = _normalize_intake_mode(intake_mode)
    report["intake_mode_label"] = "Manual Author Research" if report["intake_mode"] == TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH else "Standard"
    save_state(latest_master_check=report)
    return report


def _check_important_leads_response(
    *,
    input_path: Path,
    output_path: Path,
    rejected_path: Path,
    effective_input_path: Path,
    source_label: str | None = None,
    intake_mode: str = "STANDARD",
) -> JSONResponse:
    run_id = f"check_{timestamp_slug()}_{uuid.uuid4().hex[:8]}"
    staged_run_dir = IMPORTANT_LEADS_RUNS / run_id
    staged_run_dir.mkdir(parents=True, exist_ok=True)
    live_output_path = output_path
    live_rejected_path = rejected_path
    output_path = staged_run_dir / "leads.csv"
    rejected_path = staged_run_dir / "leads_rejected.csv"
    try:
        report = _execute_important_check(
            input_path=input_path,
            output_path=output_path,
            rejected_path=rejected_path,
            effective_input_path=effective_input_path,
            intake_mode=intake_mode,
        )
    except FileNotFoundError as exc:
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=404)
    except ImportantLeadsCheckError as exc:
        return JSONResponse(
            {
                "ok": False,
                "message": exc.message,
                "error": exc.code,
                "details": exc.details,
            },
            status_code=400,
        )
    except ValueError as exc:
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Lead check failed: {exc}"}, status_code=500)

    prefix = f"Uploaded {source_label} and " if source_label else ""
    return JSONResponse(
        {
            "ok": True,
            "run_id": run_id,
            "staged_run_dir": str(staged_run_dir),
            "live_output_path": str(live_output_path),
            "live_rejected_path": str(live_rejected_path),
            "message": (
                f"{prefix}checked {report['input_label']} into {report['output_label']}. "
                f"Kept {int(report['cleaned_rows'] or 0)} row(s), rejected "
                f"{sum(int(report['reason_counts'].get(reason, 0)) for reason in report.get('reason_counts', {}))}."
            ),
            "check": report,
            "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
        }
    )


def _run_important_check_job(job_id: str) -> None:
    try:
        job = _load_important_check_job(job_id)
    except Exception:
        return
    if str(job.get("status") or "") in {"completed", "failed"}:
        return
    try:
        job["status"] = "running"
        job["stage"] = "checking"
        job["processed_rows"] = 0
        job["remaining_rows"] = int(job.get("total_input_rows") or 0)
        job["eta_seconds"] = ""
        job["progress_percent"] = 0
        _save_important_check_job(job)
        started_at = time.monotonic()
        last_progress_save_at = 0.0

        def save_progress(processed_rows: int, total_rows: int) -> None:
            nonlocal job, last_progress_save_at
            try:
                latest_job = _load_important_check_job(job_id)
                if latest_job.get("cancel_requested"):
                    job["cancel_requested"] = True
            except Exception:
                pass
            now = time.monotonic()
            if processed_rows < total_rows and now - last_progress_save_at < 0.75:
                return
            last_progress_save_at = now
            total = max(0, int(total_rows or 0))
            processed = min(total, max(0, int(processed_rows or 0)))
            elapsed = max(0.001, now - started_at)
            rate = processed / elapsed if processed > 0 else 0.0
            remaining = max(0, total - processed)
            job["processed_rows"] = processed
            job["remaining_rows"] = remaining
            job["progress_percent"] = round(min(100.0, max(0.0, (processed / total) * 100)), 1) if total else 0
            job["eta_seconds"] = int(remaining / rate) if rate > 0 and remaining > 0 else 0
            _save_important_check_job(job)

        is_warm_research = str(job.get("upload_type") or "").strip().lower() == "warm_research"
        if is_warm_research:
            report = check_warm_research_leads(
                input_path=Path(str(job.get("effective_input_path") or "")),
                email_ready_path=Path(str(job.get("output_path") or "")),
                contact_form_review_path=Path(str(job.get("contact_form_review_path") or "")),
                rejected_path=Path(str(job.get("rejected_path") or "")),
                progress_callback=save_progress,
            )
        else:
            report = _execute_important_check(
                input_path=Path(str(job.get("input_path") or "")),
                output_path=Path(str(job.get("output_path") or "")),
                rejected_path=Path(str(job.get("rejected_path") or "")),
                effective_input_path=Path(str(job.get("effective_input_path") or "")),
                intake_mode=str(job.get("intake_mode") or "STANDARD"),
                progress_callback=save_progress,
            )
        if job.get("cancel_requested"):
            job["status"] = "canceled"
            job["stage"] = "canceled"
            job["phase"] = "canceled"
            job["completed_at_utc"] = iso_utc()
            job["check"] = report
            job["processed_rows"] = int(report.get("total_input_rows") or report.get("input_rows") or job.get("total_input_rows") or 0)
            job["remaining_rows"] = 0
            job["eta_seconds"] = 0
            job["progress_percent"] = 100
            job["auto_triage_status"] = "skipped"
            job["auto_triage_skip_reason"] = "check_canceled"
            job["message"] = "Check completed after cancellation request. Auto triage skipped."
            _save_important_check_job(job)
            return
        job["status"] = "completed"
        job["stage"] = "done"
        job["phase"] = "check_complete"
        job["completed_at_utc"] = iso_utc()
        job["check"] = report
        job["processed_rows"] = int(report.get("total_input_rows") or report.get("input_rows") or job.get("total_input_rows") or 0)
        job["remaining_rows"] = 0
        job["eta_seconds"] = 0
        job["progress_percent"] = 100
        if is_warm_research:
            job["message"] = "Warm upload checked. Generate Warm Draft Preview before explicit Warm Private JC confirmation."
            job["auto_triage_status"] = "skipped"
            job["auto_triage_skip_reason"] = "warm_research_upload"
        else:
            job["message"] = (
                f"Checked {report['input_label']} into {report['output_label']}. "
                f"Kept {int(report['cleaned_rows'] or 0)} row(s), rejected "
                f"{sum(int(report['reason_counts'].get(reason, 0)) for reason in report.get('reason_counts', {}))}."
            )
            job = _run_auto_fast_triage_after_check(job)
        _save_important_check_job(job)
    except Exception as exc:
        job["status"] = "failed"
        job["stage"] = "failed"
        job["completed_at_utc"] = iso_utc()
        job["processed_rows"] = 0
        job["remaining_rows"] = int(job.get("total_input_rows") or 0)
        job["progress_percent"] = 0
        job["error"] = str(exc)
        _save_important_check_job(job)


def _start_important_check_job(
    *,
    input_path: Path,
    output_path: Path,
    rejected_path: Path,
    effective_input_path: Path,
    source_label: str,
    source_mode: str,
    original_uploaded_filename: str = "",
    server_received_filename: str = "",
    selected_filename: str = "",
    selected_size_bytes: int = 0,
    selected_extension: str = "",
    source_sheet: str = "",
    intake_mode: str = "STANDARD",
    total_input_rows: int = 0,
    upload_type: str = "cold",
) -> dict[str, object]:
    job_id = f"check_{timestamp_slug()}_{uuid.uuid4().hex[:8]}"
    staged_run_dir = IMPORTANT_LEADS_RUNS / job_id
    normalized_upload_type = "warm_research" if str(upload_type or "").strip().lower() == "warm_research" else "cold"
    contact_form_review_path = Path("")
    if str(source_mode or "").strip() == "uploaded_file":
        staged_run_dir.mkdir(parents=True, exist_ok=True)
        if normalized_upload_type == "warm_research":
            output_path = staged_run_dir / "warm_email_ready.csv"
            contact_form_review_path = staged_run_dir / "warm_contact_form_review.csv"
            rejected_path = staged_run_dir / "warm_rejected.csv"
        else:
            output_path = staged_run_dir / "leads.csv"
            rejected_path = staged_run_dir / "leads_rejected.csv"
    job = {
        "job_id": job_id,
        "status": "queued",
        "stage": "queued",
        "created_at_utc": iso_utc(),
        "updated_at_utc": iso_utc(),
        "source_label": source_label,
        "source_mode": str(source_mode or "").strip() or "uploaded_file",
        "upload_type": normalized_upload_type,
        "upload_type_label": "Warm Research" if normalized_upload_type == "warm_research" else "Cold Leads",
        "original_uploaded_filename": str(original_uploaded_filename or "").strip() or source_label,
        "server_received_filename": str(server_received_filename or "").strip() or source_label,
        "selected_filename": str(selected_filename or "").strip() or source_label,
        "selected_size_bytes": int(selected_size_bytes or 0),
        "selected_extension": str(selected_extension or "").strip(),
        "source_sheet": str(source_sheet or "").strip(),
        "intake_mode": _normalize_intake_mode(intake_mode),
        "intake_mode_label": "Manual Author Research" if _normalize_intake_mode(intake_mode) == TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH else "Standard",
        "current_sheet": str(source_sheet or "").strip(),
        "input_path": str(input_path),
        "saved_input_path": str(effective_input_path),
        "output_path": str(output_path),
        "rejected_path": str(rejected_path),
        "contact_form_review_path": str(contact_form_review_path) if normalized_upload_type == "warm_research" else "",
        "effective_input_path": str(effective_input_path),
        "staged_run_dir": str(staged_run_dir) if str(source_mode or "").strip() == "uploaded_file" else "",
        "live_output_path": str(IMPORTANT_LEADS_OUTPUT),
        "live_rejected_path": str(IMPORTANT_LEADS_REJECTED),
        "total_input_rows": int(total_input_rows or 0),
        "processed_rows": 0,
        "remaining_rows": int(total_input_rows or 0),
        "eta_seconds": "",
        "progress_percent": 0,
    }
    _save_important_check_job(job)
    thread = threading.Thread(target=_run_important_check_job, args=(job_id,), daemon=True)
    thread.start()
    return job


def _resume_pending_important_check_jobs() -> None:
    if not IMPORTANT_LEADS_CHECK_JOBS.exists():
        return
    for path in sorted(IMPORTANT_LEADS_CHECK_JOBS.glob("*.json")):
        try:
            job = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(job, dict):
            continue
        if str(job.get("status") or "") not in {"queued", "running"}:
            continue
        job_id = str(job.get("job_id") or path.stem).strip()
        if not job_id:
            continue
        thread = threading.Thread(target=_run_important_check_job, args=(job_id,), daemon=True)
        thread.start()


def _load_important_verify_job(job_id: str) -> dict[str, object]:
    return _load_dashboard_job(IMPORTANT_LEADS_VERIFY_JOBS, job_id, "Verify")


def _save_important_verify_job(job: dict[str, object]) -> None:
    _save_dashboard_job(IMPORTANT_LEADS_VERIFY_JOBS, job)


def _load_important_dispatch_job(job_id: str) -> dict[str, object]:
    return _load_dashboard_job(IMPORTANT_LEADS_DISPATCH_JOBS, job_id, "Dispatch")


def _save_important_dispatch_job(job: dict[str, object]) -> None:
    _save_dashboard_job(IMPORTANT_LEADS_DISPATCH_JOBS, job)


def _run_important_verify_job(job_id: str) -> None:
    try:
        job = _load_important_verify_job(job_id)
    except Exception:
        return
    if str(job.get("status") or "") in {"completed", "failed", "canceled", "cancelled"}:
        return
    try:
        mode = str(job.get("mode") or TRIAGE_MODE_FAST).strip().upper()
        if mode not in {TRIAGE_MODE_FAST, TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH, TRIAGE_MODE_STRICT}:
            mode = TRIAGE_MODE_FAST
        is_fast_triage = mode != TRIAGE_MODE_STRICT
        job["status"] = "running"
        job["mode"] = mode
        job["stage"] = "manual_author_research" if mode == TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH else "fast_triage" if is_fast_triage else "strict_public_proof"
        job["phase"] = job["stage"]
        job["eta_seconds"] = ""
        job["progress_percent"] = float(job.get("progress_percent") or 0)
        _save_important_verify_job(job)
        started_at = time.monotonic()
        last_progress_save_at = 0.0

        def save_progress(processed_rows: int, total_rows: int) -> None:
            nonlocal job, last_progress_save_at
            now = time.monotonic()
            if processed_rows < total_rows and now - last_progress_save_at < 0.75:
                return
            last_progress_save_at = now
            total = max(0, int(total_rows or 0))
            processed = min(total, max(0, int(processed_rows or 0)))
            remaining = max(0, total - processed)
            elapsed = max(0.001, now - started_at)
            rate = processed / elapsed if processed > 0 else 0.0
            job["total_rows"] = total
            job["total_input_rows"] = total
            job["processed_rows"] = processed
            job["remaining_rows"] = remaining
            job["progress_percent"] = round(min(100.0, max(0.0, (processed / total) * 100)), 1) if total else 0
            job["eta_seconds"] = int(remaining / rate) if rate > 0 and remaining > 0 else 0
            _save_important_verify_job(job)

        def should_cancel() -> bool:
            try:
                latest_job = _load_important_verify_job(job_id)
            except Exception:
                return bool(job.get("cancel_requested"))
            if latest_job.get("cancel_requested"):
                job["cancel_requested"] = True
                return True
            return bool(job.get("cancel_requested"))

        if is_fast_triage:
            report = fast_triage_master_leads(
                input_path=Path(str(job.get("input_path") or "")),
                keep_path=Path(str(job.get("verified_path") or "")),
                rejected_path=Path(str(job.get("rejected_path") or "")),
                quarantine_path=Path(str(job.get("quarantine_path") or "")),
                progress_callback=save_progress,
                should_cancel=should_cancel,
                mode=mode,
            )
        else:
            report = verify_master_leads(
                input_path=Path(str(job.get("input_path") or "")),
                verified_path=Path(str(job.get("verified_path") or "")),
                rejected_path=Path(str(job.get("rejected_path") or "")),
                quarantine_path=Path(str(job.get("quarantine_path") or "")),
                progress_callback=save_progress,
                should_cancel=should_cancel,
            )
        total = int(report.get("total_input_rows") or report.get("input_rows") or job.get("total_rows") or 0)
        processed = int(report.get("processed_rows") or total)
        canceled = should_cancel() and processed < total
        job["status"] = "canceled" if canceled else "completed"
        job["stage"] = "canceled" if canceled else "done"
        job["phase"] = "canceled" if canceled else "done"
        job["completed_at_utc"] = iso_utc()
        job["verify"] = report
        job["total_rows"] = total
        job["total_input_rows"] = total
        job["processed_rows"] = processed
        job["remaining_rows"] = max(0, total - processed) if canceled else 0
        job["eta_seconds"] = 0
        job["progress_percent"] = round(min(100.0, max(0.0, (processed / total) * 100)), 1) if total else 0
        if canceled:
            job["message"] = f"Verify stopped safely at row {processed} of {total}. Checkpoint/output files were preserved."
        else:
            job["progress_percent"] = 100
            mode_label = _triage_mode_label(mode)
            job["message"] = (
                f"{mode_label} {report['input_label']} into {report['verified_label']}. "
                f"KEEP {int(report['keep_count'] or 0)}, REJECT {int(report['reject_count'] or 0)}, "
                f"QUARANTINE {int(report['quarantine_count'] or 0)}."
            )
        _save_important_verify_job(job)
    except Exception as exc:
        job["status"] = "failed"
        job["stage"] = "failed"
        job["phase"] = "failed"
        job["completed_at_utc"] = iso_utc()
        job["error"] = str(exc)
        _save_important_verify_job(job)


def _start_important_verify_job(
    *,
    input_path: Path,
    verified_path: Path,
    rejected_path: Path,
    quarantine_path: Path,
    mode: str = TRIAGE_MODE_FAST,
) -> dict[str, object]:
    job_id = f"verify_{timestamp_slug()}_{uuid.uuid4().hex[:8]}"
    total_rows = _count_csv_rows(input_path)
    raw_mode = str(mode or "").strip().upper()
    if raw_mode == TRIAGE_MODE_STRICT:
        mode = TRIAGE_MODE_STRICT
    elif raw_mode == TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH:
        mode = TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH
    else:
        mode = TRIAGE_MODE_FAST
    job = {
        "job_id": job_id,
        "mode": mode,
        "status": "queued",
        "stage": "queued",
        "phase": "queued",
        "created_at_utc": iso_utc(),
        "updated_at_utc": iso_utc(),
        "input_path": str(input_path),
        "verified_path": str(verified_path),
        "rejected_path": str(rejected_path),
        "quarantine_path": str(quarantine_path),
        "total_rows": total_rows,
        "total_input_rows": total_rows,
        "processed_rows": 0,
        "remaining_rows": total_rows,
        "eta_seconds": "",
        "progress_percent": 0,
    }
    _save_important_verify_job(job)
    thread = threading.Thread(target=_run_important_verify_job, args=(job_id,), daemon=True)
    thread.start()
    return _job_progress_payload(job)


def _run_important_dispatch_job(job_id: str) -> None:
    try:
        job = _load_important_dispatch_job(job_id)
    except Exception:
        return
    if str(job.get("status") or "") in {"completed", "failed", "canceled", "cancelled"}:
        return
    try:
        job["status"] = "running"
        job["stage"] = "dispatching"
        job["phase"] = "dispatching"
        job["eta_seconds"] = ""
        job["progress_percent"] = float(job.get("progress_percent") or 0)
        job["message"] = "Creating pre-dispatch archive before queue writes."
        _save_important_dispatch_job(job)
        archive = _create_pre_dispatch_archive(job_id)
        job["pre_dispatch_archive"] = archive
        job["pre_dispatch_archive_path"] = archive["archive_path"]
        job["message"] = "Pre-dispatch archive created. Confirming queue write."
        _save_important_dispatch_job(job)
        report = confirm_dispatch_preview(
            str(job.get("preview_id") or ""),
            require_stopped=True,
            allow_high_risk_recontact=bool(job.get("recontact_recency_override")),
            backup_root=settings.BACKUPS_DIR,
            report_dir=settings.STATE_DIR,
            persist_state=True,
            preview_dir=IMPORTANT_LEADS_DISPATCH_PREVIEWS,
        )
        report["pre_dispatch_archive"] = archive
        report["pre_dispatch_archive_path"] = archive["archive_path"]
        report_path_text = str(report.get("report_path") or "").strip()
        if report_path_text:
            report_path = Path(report_path_text)
            report_path.write_text(json.dumps(report, indent=2, sort_keys=True) + "\n", encoding="utf-8")
            settings.secure_private_file(report_path)
        save_state(**{MASTER_DISPATCH_STATE_KEY: report})
        assigned_rows = int(report.get("added_astra") or 0) + int(report.get("added_sendgrid") or 0)
        skipped_rows = (
            int(report.get("suppressed_skipped") or 0)
            + int(report.get("duplicate_master_skipped") or 0)
            + int(report.get("skipped_both") or 0)
            + int(report.get("invalid_malformed_skipped") or 0)
        )
        total_rows = int(report.get("dispatch_selected_row_count") or job.get("total_rows") or 0)
        job["status"] = "completed"
        job["stage"] = "done"
        job["phase"] = "done"
        job["completed_at_utc"] = iso_utc()
        job["run_id"] = report.get("run_id")
        job["dispatch"] = report
        job["total_rows"] = total_rows
        job["processed_rows"] = total_rows
        job["assigned_rows"] = assigned_rows
        job["skipped_rows"] = skipped_rows
        job["remaining_rows"] = 0
        job["eta_seconds"] = 0
        job["progress_percent"] = 100
        job["message"] = str(
            report.get("message")
            or (
                f"Dispatch complete. Astra added {report['added_astra']} row(s), SendGrid added "
                f"{report['added_sendgrid']} row(s), skipped both {report['skipped_both']}."
            )
        )
        _save_important_dispatch_job(job)
    except Exception as exc:
        job["status"] = "failed"
        job["stage"] = "failed"
        job["phase"] = "failed"
        job["completed_at_utc"] = iso_utc()
        job["error"] = str(exc)
        _save_important_dispatch_job(job)


def _start_important_dispatch_job(
    *,
    preview_id: str,
    campaign_type: str,
    dispatch_source_mode: str,
    dispatch_source_name: str,
    dispatch_source_path: str,
    dispatch_cap: str,
    total_source_rows: int,
    eligible_rows: int,
    selected_rows: int,
    total_rows_would_write: int,
    recontact_recency_override: bool = False,
) -> dict[str, object]:
    job_id = f"dispatch_{timestamp_slug()}_{uuid.uuid4().hex[:8]}"
    job = {
        "job_id": job_id,
        "preview_id": preview_id,
        "status": "queued",
        "stage": "queued",
        "phase": "queued",
        "created_at_utc": iso_utc(),
        "updated_at_utc": iso_utc(),
        "campaign_type": campaign_type,
        "dispatch_source_mode": dispatch_source_mode,
        "dispatch_source_name": dispatch_source_name,
        "dispatch_source_path": dispatch_source_path,
        "dispatch_cap": dispatch_cap,
        "recontact_recency_override": bool(recontact_recency_override),
        "total_source_rows": max(0, int(total_source_rows or 0)),
        "eligible_rows": max(0, int(eligible_rows or 0)),
        "selected_rows": max(0, int(selected_rows or 0)),
        "total_rows_would_write": max(0, int(total_rows_would_write or 0)),
        "total_rows": max(0, int(selected_rows or 0)),
        "processed_rows": 0,
        "assigned_rows": 0,
        "skipped_rows": 0,
        "remaining_rows": max(0, int(selected_rows or 0)),
        "eta_seconds": "",
        "progress_percent": 0,
    }
    _save_important_dispatch_job(job)
    thread = threading.Thread(target=_run_important_dispatch_job, args=(job_id,), daemon=True)
    thread.start()
    return _job_progress_payload(job)


def _count_pasted_lead_rows(input_text: str) -> int:
    normalized_text = _normalize_pasted_leads_csv(input_text)
    if not normalized_text.strip():
        return 0
    reader = csv.DictReader(io.StringIO(normalized_text))
    count = 0
    for row in reader:
        if any(str(value or "").strip() for value in row.values()):
            count += 1
    return count


def _csv_count_from_status_label(status: dict[str, object], key: str, default_path: Path) -> int:
    raw = str(status.get(key) or "").strip()
    if raw:
        path = Path(raw)
        if not path.is_absolute():
            path = settings.APP_ROOT / path
        if path.suffix.lower() == ".csv" and path.exists():
            return _count_csv_rows(path)
        return 0
    path = default_path
    return _count_csv_rows(path)


def _csv_funnel_stage(path: Path, *, missing_status: str = "not_available") -> dict[str, object]:
    if not path.is_absolute():
        path = settings.APP_ROOT / path
    payload: dict[str, object] = {
        "path": str(path),
        "row_count": None,
        "status": missing_status,
    }
    if path.exists():
        payload["row_count"] = _count_csv_rows(path)
        payload["status"] = "ready"
    return payload


def _number_funnel_stage(row_count: object, *, path: object = "", missing_status: str = "pending") -> dict[str, object]:
    try:
        value = int(row_count) if row_count not in ("", None) else None
    except Exception:
        value = None
    return {
        "path": str(path or ""),
        "row_count": value,
        "status": "ready" if value is not None else missing_status,
    }


def _latest_important_check_job() -> dict[str, object] | None:
    if not IMPORTANT_LEADS_CHECK_JOBS.exists():
        return None
    candidates: list[tuple[float, dict[str, object]]] = []
    for path in IMPORTANT_LEADS_CHECK_JOBS.glob("*.json"):
        try:
            job = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(job, dict):
            continue
        if str(job.get("upload_type") or "cold").strip().lower() == "warm_research":
            continue
        candidates.append((_check_job_created_sort_key(job, path), _job_progress_payload(job)))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def _latest_completed_important_check_job() -> dict[str, object] | None:
    if not IMPORTANT_LEADS_CHECK_JOBS.exists():
        return None
    candidates: list[tuple[float, dict[str, object]]] = []
    for path in IMPORTANT_LEADS_CHECK_JOBS.glob("*.json"):
        try:
            job = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(job, dict):
            continue
        if str(job.get("status") or "").strip().lower() not in {"completed", "done"}:
            continue
        if str(job.get("upload_type") or "cold").strip().lower() == "warm_research":
            continue
        candidates.append((_check_job_created_sort_key(job, path), job))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def _latest_completed_warm_check_job() -> dict[str, object] | None:
    if not IMPORTANT_LEADS_CHECK_JOBS.exists():
        return None
    candidates: list[tuple[float, dict[str, object]]] = []
    for path in IMPORTANT_LEADS_CHECK_JOBS.glob("*.json"):
        try:
            job = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(job, dict):
            continue
        if str(job.get("status") or "").strip().lower() not in {"completed", "done"}:
            continue
        if str(job.get("upload_type") or "").strip().lower() != "warm_research":
            continue
        candidates.append((_check_job_created_sort_key(job, path), job))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def _staged_run_dir_for_job(job: dict[str, object]) -> Path:
    staged_dir = Path(str(job.get("staged_run_dir") or "")) if job.get("staged_run_dir") else IMPORTANT_LEADS_RUNS / str(job.get("job_id") or "")
    if not staged_dir.is_absolute():
        staged_dir = settings.APP_ROOT / staged_dir
    return staged_dir


def _staged_triage_paths_for_job(job: dict[str, object]) -> dict[str, Path]:
    staged_dir = _staged_run_dir_for_job(job)

    def path_from_job(key: str, filename: str) -> Path:
        path = Path(str(job.get(key) or staged_dir / filename))
        if not path.is_absolute():
            path = settings.APP_ROOT / path
        return path

    return {
        "input": path_from_job("output_path", "leads.csv"),
        "rejected": path_from_job("rejected_path", "leads_rejected.csv"),
        "keep": path_from_job("auto_triage_keep_path", "leads_triaged_keep.csv"),
        "triage_reject": path_from_job("auto_triage_rejected_path", "leads_triaged_reject.csv"),
        "triage_quarantine": path_from_job("auto_triage_quarantine_path", "leads_triaged_quarantine.csv"),
    }


def _latest_fast_triage_keep_source() -> dict[str, object]:
    job = _latest_completed_important_check_job()
    if job:
        paths = _staged_triage_paths_for_job(job)
        keep_path = paths["keep"]
        return {
            "source_resolution": "latest_completed_staged_run",
            "job": job,
            "run_id": str(job.get("job_id") or ""),
            "path": keep_path,
            "paths": paths,
            "exists": keep_path.exists(),
            "row_count": _count_csv_rows(keep_path) if keep_path.exists() else 0,
        }
    legacy_path = TRIAGED_KEEP_PATH
    return {
        "source_resolution": "legacy_important_triaged_keep",
        "job": None,
        "run_id": "",
        "path": legacy_path,
        "paths": {},
        "exists": legacy_path.exists(),
        "row_count": _count_csv_rows(legacy_path) if legacy_path.exists() else 0,
    }


def _normalize_dashboard_path_text(value: object) -> str:
    text = str(value or "").strip().replace("\\", "/")
    while "//" in text:
        text = text.replace("//", "/")
    return text.rstrip("/")


def _dashboard_paths_match(left: object, right: object) -> bool:
    left_text = _normalize_dashboard_path_text(left)
    right_text = _normalize_dashboard_path_text(right)
    if not left_text or not right_text:
        return False
    try:
        return Path(left_text).resolve(strict=False) == Path(right_text).resolve(strict=False)
    except Exception:
        return left_text == right_text


def _dispatch_summary_timestamp(summary: dict[str, object]) -> datetime | None:
    for key in ("generated_at_utc", "confirmed_at", "confirmed_at_utc", "completed_at_utc", "completed_at"):
        parsed = _parse_iso_timestamp(summary.get(key))
        if parsed is not None:
            return parsed
    return None


def _preview_summary_timestamp(summary: dict[str, object]) -> datetime | None:
    for key in ("generated_at_utc", "completed_at_utc", "completed_at", "created_at_utc", "created_at"):
        parsed = _parse_iso_timestamp(summary.get(key))
        if parsed is not None:
            return parsed
    return None


def _summary_current_for_staged_source(
    summary: dict[str, object],
    *,
    source_path: Path,
    source_generated_at: object,
    timestamp_reader: Callable[[dict[str, object]], datetime | None],
) -> bool:
    if not summary:
        return False
    summary_source = (
        summary.get("dispatch_source_path")
        or summary.get("source_path")
        or summary.get("triaged_keep_path")
        or ""
    )
    if summary_source and not _dashboard_paths_match(summary_source, source_path):
        return False
    source_time = _parse_iso_timestamp(source_generated_at)
    summary_time = timestamp_reader(summary)
    if source_time is not None and summary_time is not None and summary_time < source_time:
        return False
    if source_time is not None and summary_time is None:
        return False
    return True


def _read_dispatch_source_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not path.exists():
        return [], []
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        reader = csv.DictReader(handle)
        headers = [str(field or "").strip().lstrip("\ufeff") for field in (reader.fieldnames or [])]
        rows = [
            {header: str(row.get(header, "") or "") for header in headers}
            for row in reader
            if any(str(value or "").strip() for value in row.values())
        ]
    return headers, rows


def _dispatch_source_status_for_path(
    *,
    path: Path,
    mode: str,
    source_resolution: str = "",
    run_id: str = "",
) -> dict[str, object]:
    headers, rows = _read_dispatch_source_rows(path)
    exists = path.exists()
    status_keep_filter = mode in {DISPATCH_SOURCE_TRIAGED_KEEP, DISPATCH_SOURCE_STRICT_VERIFIED}
    eligible_rows = rows
    if status_keep_filter and "Status" in headers:
        eligible_rows = [row for row in rows if str(row.get("Status") or "").strip().upper() == "KEEP"]
    source_name = "Fast Triage Keep" if mode == DISPATCH_SOURCE_TRIAGED_KEEP else "Strict Public Proof Verified" if mode == DISPATCH_SOURCE_STRICT_VERIFIED else "Cleaned Leads"
    block_reason = ""
    if mode == DISPATCH_SOURCE_TRIAGED_KEEP and source_resolution == "latest_completed_staged_run" and (not exists or not rows):
        block_reason = "Current staged Fast Triage Keep is empty. Run Check Leads / Fast Triage first."
    elif not exists:
        block_reason = f"{source_name} dispatch source missing: {_dashboard_path_label(path)}"
    elif not rows:
        block_reason = f"{source_name} dispatch source is empty: {_dashboard_path_label(path)}"
    elif status_keep_filter and "Status" in headers and not eligible_rows:
        block_reason = f"{source_name} dispatch source has no KEEP rows: {_dashboard_path_label(path)}"
    verification_file_mtime = ""
    if exists:
        try:
            verification_file_mtime = datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()
        except Exception:
            verification_file_mtime = ""
    return {
        "dispatch_source_mode": mode,
        "dispatch_source_name": source_name,
        "dispatch_source_path": str(path),
        "dispatch_source_label": _dashboard_path_label(path),
        "dispatch_source_exists": exists,
        "dispatch_source_row_count": len(rows),
        "dispatch_eligible_row_count": len(eligible_rows),
        "dispatch_block_reason": block_reason,
        "verification_required": mode == DISPATCH_SOURCE_STRICT_VERIFIED,
        "status_keep_filter": status_keep_filter,
        "verification_file_mtime": verification_file_mtime,
        "dispatch_source_preview_rows": eligible_rows[:5],
        "dispatch_source_headers": headers,
        "source_resolution": source_resolution,
        "run_id": run_id,
    }


def _apply_latest_staged_run_status(status: dict[str, object]) -> dict[str, object]:
    fast_triage_source = _latest_fast_triage_keep_source()
    if fast_triage_source.get("source_resolution") != "latest_completed_staged_run":
        return status
    job = fast_triage_source.get("job") if isinstance(fast_triage_source.get("job"), dict) else None
    paths = fast_triage_source.get("paths") if isinstance(fast_triage_source.get("paths"), dict) else {}
    if not job or not paths:
        return status

    input_path = Path(paths["input"])
    rejected_path = Path(paths["rejected"])
    keep_path = Path(paths["keep"])
    triage_reject_path = Path(paths["triage_reject"])
    triage_quarantine_path = Path(paths["triage_quarantine"])
    intake_mode = _normalize_intake_mode(job.get("intake_mode"))
    intake_mode_label = "Manual Author Research" if intake_mode == TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH else "Standard"

    check_rows = _count_csv_rows(input_path) if input_path.exists() else 0
    rejected_rows = _count_csv_rows(rejected_path) if rejected_path.exists() else 0
    keep_rows = _count_csv_rows(keep_path) if keep_path.exists() else 0
    triage_reject_rows = _count_csv_rows(triage_reject_path) if triage_reject_path.exists() else 0
    quarantine_rows = _count_csv_rows(triage_quarantine_path) if triage_quarantine_path.exists() else 0

    latest_master_check = dict(job.get("check") or status.get("latest_master_check") or {})
    latest_master_check.update(
        {
            "intake_mode": intake_mode,
            "intake_mode_label": intake_mode_label,
            "input_rows": int(job.get("total_input_rows") or latest_master_check.get("input_rows") or check_rows),
            "cleaned_rows": check_rows,
            "output_rows": check_rows,
            "rejected_rows": rejected_rows,
            "output_label": _dashboard_path_label(input_path),
            "rejected_label": _dashboard_path_label(rejected_path),
            "generated_at_utc": str(
                latest_master_check.get("generated_at_utc")
                or job.get("completed_at_utc")
                or job.get("updated_at_utc")
                or ""
            ),
        }
    )

    latest_triage = dict(job.get("auto_triage_report") or status.get("latest_lead_triage") or {})
    latest_triage.update(
        {
            "mode": TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH if intake_mode == TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH else TRIAGE_MODE_FAST,
            "intake_mode": intake_mode,
            "intake_mode_label": intake_mode_label,
            "input_count": check_rows,
            "keep_count": keep_rows,
            "reject_count": triage_reject_rows,
            "rejected_count": triage_reject_rows,
            "quarantine_count": quarantine_rows,
            "review_count": quarantine_rows,
            "keep_path": _dashboard_path_label(keep_path),
            "rejected_path": _dashboard_path_label(triage_reject_path),
            "quarantine_path": _dashboard_path_label(triage_quarantine_path),
            "generated_at_utc": str(
                latest_triage.get("generated_at_utc")
                or job.get("auto_triage_completed_at_utc")
                or job.get("updated_at_utc")
                or ""
            ),
        }
    )
    triage_generated_at = latest_triage.get("generated_at_utc")

    source_status = _dispatch_source_status_for_path(
        path=keep_path,
        mode=DISPATCH_SOURCE_TRIAGED_KEEP,
        source_resolution=str(fast_triage_source.get("source_resolution") or ""),
        run_id=str(fast_triage_source.get("run_id") or ""),
    )
    source_options = dict(status.get("dispatch_source_options") or {})
    source_options[DISPATCH_SOURCE_TRIAGED_KEEP] = source_status
    latest_preview = dict(job.get("auto_dispatch_preview") or status.get("latest_auto_dispatch_preview") or {})
    latest_preview_current = _summary_current_for_staged_source(
        latest_preview,
        source_path=keep_path,
        source_generated_at=triage_generated_at,
        timestamp_reader=_preview_summary_timestamp,
    )
    latest_dispatch = status.get("latest_dispatch") if isinstance(status.get("latest_dispatch"), dict) else {}
    latest_dispatch_current = _summary_current_for_staged_source(
        latest_dispatch,
        source_path=keep_path,
        source_generated_at=triage_generated_at,
        timestamp_reader=_dispatch_summary_timestamp,
    )

    status.update(
        {
            "important_output_label": _dashboard_path_label(input_path),
            "important_rejected_label": _dashboard_path_label(rejected_path),
            "important_triage_input_label": _dashboard_path_label(input_path),
            "important_triage_keep_label": _dashboard_path_label(keep_path),
            "important_triage_rejected_label": _dashboard_path_label(triage_reject_path),
            "important_triage_quarantine_label": _dashboard_path_label(triage_quarantine_path),
            "latest_master_check": latest_master_check,
            "latest_lead_triage": latest_triage,
            "latest_auto_dispatch_preview": latest_preview if latest_preview_current else {},
            "latest_dispatch": latest_dispatch if latest_dispatch_current else {},
            "stale_latest_dispatch": latest_dispatch if latest_dispatch and not latest_dispatch_current else {},
            "latest_confirmed_dispatch_current": latest_dispatch_current,
            "latest_auto_dispatch_preview_current": latest_preview_current,
            "dispatch_source_options": source_options,
        }
    )
    if str(status.get("dispatch_source_mode") or DISPATCH_SOURCE_TRIAGED_KEEP).strip().lower() == DISPATCH_SOURCE_TRIAGED_KEEP:
        status.update(
            {
                "dispatch_source_path": source_status["dispatch_source_label"],
                "dispatch_source_exists": source_status["dispatch_source_exists"],
                "dispatch_source_row_count": source_status["dispatch_source_row_count"],
                "dispatch_eligible_row_count": source_status["dispatch_eligible_row_count"],
                "dispatch_block_reason": source_status["dispatch_block_reason"],
                "verification_required": source_status["verification_required"],
                "verification_file_mtime": source_status["verification_file_mtime"],
                "dispatch_source_preview_rows": source_status["dispatch_source_preview_rows"],
                "dispatch_source": source_status,
            }
        )
    return status


def _funnel_stage_count(stage: dict[str, object]) -> int | None:
    if str(stage.get("status") or "") != "ready":
        return None
    try:
        return int(stage.get("row_count")) if stage.get("row_count") is not None else None
    except Exception:
        return None


def _finish_funnel_summary(summary: dict[str, object]) -> dict[str, object]:
    raw_count = _funnel_stage_count(summary["raw_input"]) if isinstance(summary.get("raw_input"), dict) else None
    final_count = _funnel_stage_count(summary["final_eligible"]) if isinstance(summary.get("final_eligible"), dict) else None
    component_removed = 0
    component_ready = False
    for key in ("check_rejected", "triage_reject", "triage_quarantine"):
        stage = summary.get(key)
        if isinstance(stage, dict):
            count = _funnel_stage_count(stage)
            if count is not None:
                component_removed += count
                component_ready = True

    removed_count: int | None = None
    if raw_count is not None and final_count is not None:
        removed_count = max(0, raw_count - final_count)
    elif component_ready:
        removed_count = component_removed

    summary["total_removed_excluded"] = _number_funnel_stage(removed_count)
    pass_through: float | None = None
    if raw_count and final_count is not None:
        pass_through = round((final_count / raw_count) * 100, 1)
    summary["pass_through_rate"] = {
        "value": pass_through,
        "status": "ready" if pass_through is not None else "pending",
    }
    return summary


def _build_current_live_funnel_summary() -> dict[str, object]:
    checked = _csv_funnel_stage(IMPORTANT_LEADS_OUTPUT)
    triage_keep = _csv_funnel_stage(TRIAGED_KEEP_PATH)
    summary: dict[str, object] = {
        "label": "Current Live Funnel",
        "source": "live_current_files",
        "raw_input": _csv_funnel_stage(IMPORTANT_LEADS_INPUT),
        "cleaned_after_check": checked,
        "check_rejected": _csv_funnel_stage(IMPORTANT_LEADS_REJECTED),
        "triage_keep": triage_keep,
        "triage_reject": _csv_funnel_stage(TRIAGED_KEEP_PATH.with_name("leads_triaged_reject.csv")),
        "triage_quarantine": _csv_funnel_stage(IMPORTANT_LEADS_OUTPUT.with_name("leads_triaged_quarantine.csv")),
        "final_eligible": dict(triage_keep) if triage_keep.get("status") == "ready" else _number_funnel_stage(None),
    }
    return _finish_funnel_summary(summary)


def _build_next_batch_funnel_summary(status: dict[str, object]) -> dict[str, object]:
    active_job = status.get("active_important_check_job") if isinstance(status.get("active_important_check_job"), dict) else None
    if active_job and str(active_job.get("upload_type") or "cold").strip().lower() == "warm_research":
        active_job = None
    latest_job = active_job or _latest_important_check_job()
    if not latest_job:
        pending = _number_funnel_stage(None)
        summary: dict[str, object] = {
            "label": "Next Batch Funnel",
            "source": "latest_staged_run",
            "run_id": "",
            "staged_run_dir": "",
            "raw_input": pending,
            "cleaned_after_check": pending,
            "check_rejected": pending,
            "triage_keep": pending,
            "triage_reject": pending,
            "triage_quarantine": pending,
            "final_eligible": pending,
        }
        return _finish_funnel_summary(summary)

    staged_dir = Path(str(latest_job.get("staged_run_dir") or "")) if latest_job.get("staged_run_dir") else IMPORTANT_LEADS_RUNS / str(latest_job.get("job_id") or "")
    if not staged_dir.is_absolute():
        staged_dir = settings.APP_ROOT / staged_dir
    preview = latest_job.get("auto_dispatch_preview") if isinstance(latest_job.get("auto_dispatch_preview"), dict) else {}
    keep_path = Path(str(latest_job.get("auto_triage_keep_path") or staged_dir / "leads_triaged_keep.csv"))
    if not keep_path.is_absolute():
        keep_path = settings.APP_ROOT / keep_path
    final_rows = preview.get("dispatch_eligible_row_count") if preview else None
    if final_rows in {"", None}:
        keep_stage = _csv_funnel_stage(keep_path, missing_status="pending")
        final_stage = dict(keep_stage) if keep_stage.get("status") == "ready" else _number_funnel_stage(None)
    else:
        final_stage = _number_funnel_stage(final_rows, path=preview.get("source_path") or keep_path)

    summary = {
        "label": "Next Batch Funnel",
        "source": "latest_staged_run",
        "run_id": str(latest_job.get("job_id") or ""),
        "staged_run_dir": str(staged_dir),
        "raw_input": _number_funnel_stage(latest_job.get("total_input_rows"), path=latest_job.get("effective_input_path") or latest_job.get("input_path")),
        "cleaned_after_check": _csv_funnel_stage(Path(str(latest_job.get("output_path") or staged_dir / "leads.csv")), missing_status="pending"),
        "check_rejected": _csv_funnel_stage(Path(str(latest_job.get("rejected_path") or staged_dir / "leads_rejected.csv")), missing_status="pending"),
        "triage_keep": _csv_funnel_stage(keep_path, missing_status="pending"),
        "triage_reject": _csv_funnel_stage(Path(str(latest_job.get("auto_triage_rejected_path") or staged_dir / "leads_triaged_reject.csv")), missing_status="pending"),
        "triage_quarantine": _csv_funnel_stage(Path(str(latest_job.get("auto_triage_quarantine_path") or staged_dir / "leads_triaged_quarantine.csv")), missing_status="pending"),
        "final_eligible": final_stage,
    }
    return _finish_funnel_summary(summary)


def _build_lead_funnel_summary(status: dict[str, object]) -> dict[str, object]:
    return {
        "current_live": _build_current_live_funnel_summary(),
        "next_batch": _build_next_batch_funnel_summary(status),
    }


def _build_leads_pipeline_status(status: dict[str, object]) -> dict[str, object]:
    active_check = status.get("active_important_check_job") if isinstance(status.get("active_important_check_job"), dict) else None
    active_verify = status.get("active_important_verify_job") if isinstance(status.get("active_important_verify_job"), dict) else None
    active_dispatch = status.get("active_important_dispatch_job") if isinstance(status.get("active_important_dispatch_job"), dict) else None
    latest_check = status.get("latest_master_check") if isinstance(status.get("latest_master_check"), dict) else {}
    latest_triage = status.get("latest_lead_triage") if isinstance(status.get("latest_lead_triage"), dict) else {}
    latest_verify = status.get("latest_lead_verify") if isinstance(status.get("latest_lead_verify"), dict) else {}
    latest_dispatch = status.get("latest_dispatch") if isinstance(status.get("latest_dispatch"), dict) else {}
    dispatch_source = status.get("dispatch_source") if isinstance(status.get("dispatch_source"), dict) else {}

    checked_rows = _csv_count_from_status_label(status, "important_output_label", IMPORTANT_LEADS_OUTPUT)
    triaged_rows = _csv_count_from_status_label(status, "important_triage_keep_label", TRIAGED_KEEP_PATH)
    strict_rows = _csv_count_from_status_label(status, "important_verify_keep_label", STRICT_VERIFIED_PATH)
    quarantine_rows = _csv_count_from_status_label(status, "important_triage_quarantine_label", IMPORTANT_LEADS_OUTPUT.with_name("leads_triaged_quarantine.csv"))
    eligible_rows = int(dispatch_source.get("dispatch_eligible_row_count") or status.get("dispatch_eligible_row_count") or 0)
    auto_triage_running = bool(active_check and str(active_check.get("auto_triage_status") or "").lower() == "running")
    check_running = bool(active_check and not auto_triage_running)

    steps = [
        {
            "key": "check",
            "label": "Check",
            "state": "active" if check_running else ("done" if checked_rows or latest_check.get("generated_at_utc") or auto_triage_running else "waiting"),
            "count": checked_rows or int(latest_check.get("cleaned_rows") or 0),
            "note": "Cleaning/upload job running." if check_running else "Check complete." if auto_triage_running else "Cleaned leads ready." if checked_rows else "Run Check Leads.",
        },
        {
            "key": "triage",
            "label": "Triage",
            "state": "active" if (auto_triage_running or (active_verify and str(active_verify.get("mode") or "").upper() == TRIAGE_MODE_FAST)) else ("done" if triaged_rows or latest_triage.get("generated_at_utc") else "waiting"),
            "count": triaged_rows or int(active_check.get("auto_triage_processed_rows") or 0) if auto_triage_running else triaged_rows or int(latest_triage.get("keep_count") or latest_triage.get("verified_count") or 0),
            "note": "Auto triage running." if auto_triage_running else "Fast triage running." if active_verify else "Fast triage keep rows ready." if triaged_rows else "Run fast triage.",
        },
        {
            "key": "quarantine",
            "label": "Review",
            "state": "warn" if quarantine_rows else ("done" if triaged_rows else "waiting"),
            "count": quarantine_rows,
            "note": "Review quarantined rows before dispatch." if quarantine_rows else "No triage quarantine rows detected." if triaged_rows else "Waiting for triage.",
        },
        {
            "key": "preview",
            "label": "Preview",
            "state": "done" if eligible_rows else "waiting",
            "count": eligible_rows,
            "note": "Dispatch source has eligible rows." if eligible_rows else "Run Preview Dispatch after source is ready.",
        },
        {
            "key": "dispatch",
            "label": "Dispatch",
            "state": "active" if active_dispatch else ("done" if latest_dispatch.get("generated_at_utc") else "waiting"),
            "count": int(latest_dispatch.get("dispatch_selected_row_count") or latest_dispatch.get("total_rows_would_write") or 0),
            "note": "Dispatch job running." if active_dispatch else "Last dispatch complete." if latest_dispatch.get("generated_at_utc") else "Confirm Dispatch after preview.",
        },
    ]
    active_step = next((step["key"] for step in steps if step["state"] == "active"), "")
    next_step = next((step["key"] for step in steps if step["state"] in {"waiting", "warn"}), "")
    return {
        "steps": steps,
        "active_step": active_step,
        "next_step": active_step or next_step,
        "checked_rows": checked_rows,
        "triaged_keep_rows": triaged_rows,
        "strict_verified_rows": strict_rows,
        "quarantine_rows": quarantine_rows,
        "dispatch_eligible_rows": eligible_rows,
        "latest_pre_dispatch_archive_path": str((latest_dispatch.get("pre_dispatch_archive") or {}).get("archive_path") or latest_dispatch.get("pre_dispatch_archive_path") or ""),
    }


def _queue_status_missing_booktitle(status: dict[str, object]) -> list[str]:
    missing: list[str] = []
    jc_queue = status.get("jc_queue") if isinstance(status.get("jc_queue"), dict) else {}
    if jc_queue:
        fields = [str(field).strip().lower() for field in jc_queue.get("fieldnames", []) if str(field).strip()]
        if "booktitle" not in fields:
            missing.append(str(jc_queue.get("profile") or jc_queue.get("name") or "private_jc"))
    sendgrid_queues = status.get("sendgrid_queues")
    if isinstance(sendgrid_queues, list):
        for queue in sendgrid_queues:
            if not isinstance(queue, dict):
                continue
            fields = [str(field).strip().lower() for field in queue.get("fieldnames", []) if str(field).strip()]
            if "booktitle" not in fields:
                missing.append(str(queue.get("profile") or queue.get("name") or queue.get("path") or "sendgrid"))
    return missing


def _build_current_send_safety_status(status: dict[str, object]) -> dict[str, object]:
    dispatch_source = status.get("dispatch_source") if isinstance(status.get("dispatch_source"), dict) else {}
    source_resolution = str(dispatch_source.get("source_resolution") or "").strip()
    if source_resolution == "latest_completed_staged_run":
        checked_path = _state_label_path(status.get("important_output_label")) or IMPORTANT_LEADS_OUTPUT
        keep_path = _state_label_path(status.get("important_triage_keep_label")) or TRIAGED_KEEP_PATH
        reject_path = _state_label_path(status.get("important_triage_rejected_label")) or IMPORTANT_LEADS_OUTPUT.with_name("leads_triaged_reject.csv")

        def scoped_queue_safety(provider: str, shard_paths: list[Path] | None) -> dict[str, object]:
            try:
                report = build_queue_safety_report(
                    shard_paths=shard_paths,
                    intended_source_path=keep_path,
                    checked_path=checked_path,
                    triaged_keep_path=keep_path,
                    triaged_reject_path=reject_path,
                )
            except Exception as exc:
                return {
                    "safe": False,
                    "unsafe_reasons": ["QUEUE_SAFETY_CHECK_FAILED"],
                    "message": f"Queue safety check failed: {exc}",
                    "provider": provider,
                    "affected_provider": provider,
                    "validated_shard_paths": [str(path) for path in shard_paths] if shard_paths is not None else [],
                }
            report["provider"] = provider
            report["affected_provider"] = provider
            report["validated_shard_paths"] = [str(item.get("path") or "") for item in report.get("shards", []) if isinstance(item, dict)]
            return report

        sendgrid_paths = [settings.SHARDS_DIR / f"recipients_sendgrid_{index}.csv" for index in range(1, 6)]
        private_paths = [settings.SHARDS_DIR / "recipients_private_jc.csv"]
        queue_safety = scoped_queue_safety("all", None)
        sendgrid_queue_safety = scoped_queue_safety("sendgrid", sendgrid_paths)
        private_queue_safety = scoped_queue_safety("private_jc", private_paths)
    else:
        queue_safety = build_dashboard_queue_safety_report("all")
        sendgrid_queue_safety = build_dashboard_queue_safety_report("sendgrid")
        private_queue_safety = build_dashboard_queue_safety_report("private_jc")
    reasons: list[str] = []
    if not bool(queue_safety.get("safe")):
        raw_reasons = queue_safety.get("unsafe_reasons")
        if isinstance(raw_reasons, list) and raw_reasons:
            reasons.extend(str(reason) for reason in raw_reasons)
        else:
            reasons.append(str(queue_safety.get("message") or "Live recipient queue is unsafe."))
    missing_booktitle = _queue_status_missing_booktitle(status)
    if missing_booktitle:
        reasons.append(f"Live recipient queues are missing BookTitle: {', '.join(missing_booktitle)}.")
    return {
        "status": "BLOCKED" if reasons else "READY",
        "sendgrid_status": "READY" if bool(sendgrid_queue_safety.get("safe")) else "BLOCKED",
        "private_status": "READY" if bool(private_queue_safety.get("safe")) else "BLOCKED",
        "blocked": bool(reasons),
        "reasons": reasons,
        "queue_safety": queue_safety,
        "combined_queue_safety": queue_safety,
        "sendgrid_queue_safety": sendgrid_queue_safety,
        "private_queue_safety": private_queue_safety,
        "missing_booktitle_queues": missing_booktitle,
    }


def _build_next_batch_prep_status(status: dict[str, object]) -> dict[str, object]:
    active_check = status.get("active_important_check_job") if isinstance(status.get("active_important_check_job"), dict) else None
    pipeline = status.get("pipeline") if isinstance(status.get("pipeline"), dict) else {}
    latest_check = status.get("latest_master_check") if isinstance(status.get("latest_master_check"), dict) else {}
    latest_triage = status.get("latest_lead_triage") if isinstance(status.get("latest_lead_triage"), dict) else {}
    latest_preview = status.get("latest_auto_dispatch_preview") if isinstance(status.get("latest_auto_dispatch_preview"), dict) else {}
    reasons: list[str] = []
    if active_check:
        reasons.append("Check Leads is running for the next batch.")
    if not latest_check.get("generated_at_utc") and not active_check:
        reasons.append("Staged leads.csv is not ready.")
    if latest_check.get("generated_at_utc") and not latest_triage.get("generated_at_utc"):
        reasons.append("Staged triage is not ready.")
    if latest_triage.get("generated_at_utc") and not latest_preview:
        reasons.append("Staged dispatch preview is not ready.")
    return {
        "status": "WAIT" if active_check else ("SAFE TO PROMOTE" if not reasons else "NOT READY"),
        "blocks_current_send": False,
        "reasons": reasons,
        "active_check_job": active_check,
        "pipeline": pipeline,
    }


def _load_latest_confirmed_dispatch_summary() -> dict[str, object]:
    confirmed_dir = settings.STATE_DIR / "dispatch_confirmed"
    if not confirmed_dir.exists():
        return {}
    candidates = sorted(
        confirmed_dir.glob("dispatch_confirmed_*.json"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    for path in candidates:
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(payload, dict):
            continue
        report = payload.get("report") if isinstance(payload.get("report"), dict) else {}
        summary = dict(report or {})
        summary.update(
            {
                "confirmed_summary_path": str(path),
                "confirmed_at": str(payload.get("confirmed_at") or payload.get("confirmed_at_utc") or ""),
                "private_jc_added": int(payload.get("private_jc_added") or summary.get("added_astra") or 0),
                "sendgrid_added": int(payload.get("sendgrid_added") or summary.get("added_sendgrid") or 0),
                "sg1_added": int(payload.get("sg1_added") or summary.get("assigned_sg1") or 0),
                "sg2_added": int(payload.get("sg2_added") or summary.get("assigned_sg2") or 0),
                "sg3_added": int(payload.get("sg3_added") or summary.get("assigned_sg3") or 0),
                "sg4_added": int(payload.get("sg4_added") or summary.get("assigned_sg4") or 0),
                "sg5_added": int(payload.get("sg5_added") or summary.get("assigned_sg5") or 0),
                "backup_path": str(payload.get("backup_path") or summary.get("backup_dir") or ""),
                "assigned_preview_archive_path": str(payload.get("assigned_preview_archive_path") or summary.get("assigned_preview_archive_path") or ""),
            }
        )
        if not summary.get("generated_at_utc"):
            summary["generated_at_utc"] = summary.get("confirmed_at") or payload.get("confirmed_at_utc") or ""
        if not summary.get("dispatch_source_path"):
            summary["dispatch_source_path"] = payload.get("source_path") or ""
        if not summary.get("dispatch_source_row_count"):
            summary["dispatch_source_row_count"] = int(payload.get("source_rows") or 0)
        if not summary.get("dispatch_eligible_row_count"):
            summary["dispatch_eligible_row_count"] = int(payload.get("eligible_rows") or 0)
        if not summary.get("added_astra"):
            summary["added_astra"] = summary["private_jc_added"]
        if not summary.get("added_sendgrid"):
            summary["added_sendgrid"] = summary["sendgrid_added"]
        for index in range(1, 6):
            summary.setdefault(f"assigned_sg{index}", summary.get(f"sg{index}_added") or 0)
        return summary
    return {}


def _load_active_campaign_snapshot_summary() -> dict[str, object]:
    path = active_campaign_manifest_path(settings.STATE_DIR)
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    if not isinstance(payload, dict):
        return {}
    files = payload.get("files") if isinstance(payload.get("files"), dict) else {}
    intended = files.get("intended_source") if isinstance(files.get("intended_source"), dict) else {}
    checked = files.get("checked") if isinstance(files.get("checked"), dict) else {}
    return {
        "manifest_path": str(path),
        "created_at_utc": str(payload.get("created_at_utc") or payload.get("created_at") or ""),
        "campaign_type": str(payload.get("campaign_type") or ""),
        "source": str(payload.get("source") or ""),
        "intended_source_path": str(payload.get("intended_source_path") or intended.get("path") or ""),
        "intended_source_row_count": int(intended.get("row_count") or 0),
        "checked_path": str(payload.get("checked_path") or checked.get("path") or ""),
        "checked_row_count": int(checked.get("row_count") or 0),
    }


def _combined_leads_status() -> dict[str, object]:
    state = load_state()
    latest_warm_job = _latest_completed_warm_check_job()
    warm_status = build_warm_private_jc_live_status()
    latest_confirmed_dispatch = _load_latest_confirmed_dispatch_summary()
    active_campaign_snapshot = _load_active_campaign_snapshot_summary()
    status = {
        **shard_status(),
        **important_leads_status(),
        **important_leads_verify_status(),
        "active_important_check_job": _find_active_important_check_job(),
        "active_important_verify_job": _find_active_dashboard_job(IMPORTANT_LEADS_VERIFY_JOBS),
        "active_important_dispatch_job": _find_active_dashboard_job(IMPORTANT_LEADS_DISPATCH_JOBS),
        "latest_auto_dispatch_preview": state.get("latest_auto_dispatch_preview", {}),
        "latest_confirmed_dispatch": latest_confirmed_dispatch,
        "active_campaign_snapshot": active_campaign_snapshot,
        "safer_recontact_source_summary": _load_safer_recontact_source_summary(),
        "latest_warm_check": latest_warm_job.get("check", {}) if latest_warm_job else {},
        "warm_private_jc_status": warm_status,
        "warm_private_jc_lane": warm_status,
    }
    if latest_confirmed_dispatch:
        status["latest_dispatch"] = latest_confirmed_dispatch
    status = _apply_latest_staged_run_status(status)
    status["pipeline"] = _build_leads_pipeline_status(status)
    status["lead_funnel"] = _build_lead_funnel_summary(status)
    status["current_send_safety"] = _build_current_send_safety_status(status)
    status["next_batch_prep"] = _build_next_batch_prep_status(status)
    return status


def _load_safer_recontact_source_summary() -> dict[str, object]:
    path = settings.STATE_DIR / "safer_recontact_source_summary.json"
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if not isinstance(raw, dict):
        return {}
    safe_keys = {
        "preview_id",
        "campaign_type",
        "dispatch_source_mode",
        "source_path",
        "original_source_rows",
        "planned_unique",
        "found_in_active_history",
        "found_in_active_history_pct",
        "seen_this_month",
        "seen_this_month_pct",
        "not_found_in_active_history",
        "not_found_in_active_history_pct",
        "risk_level",
        "safer_found_in_active_history",
        "safer_rows_written",
        "output_path",
        "created_at",
    }
    return {key: value for key, value in raw.items() if key in safe_keys}


def _normalize_pasted_leads_csv(input_text: str) -> str:
    normalized_text = str(input_text or "").lstrip("\ufeff").replace("\r\n", "\n").replace("\r", "\n")
    if not normalized_text.endswith("\n"):
        normalized_text += "\n"

    non_empty_lines = [line for line in normalized_text.splitlines() if line.strip()]
    if not non_empty_lines:
        return normalized_text

    sample = "\n".join(non_empty_lines[:5])
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        delimiter = "," if "," in non_empty_lines[0] else "\t"

    first_row = next(csv.reader([non_empty_lines[0]], delimiter=delimiter), [])
    normalized_headers = {"".join(ch for ch in str(cell or "").strip().lower() if ch.isalnum()) for cell in first_row}
    known_email_headers = {
        "email",
        "emailaddress",
        "email_address",
        "e_mail",
        "mail",
        "authoremail",
        "contactemail",
    }
    if normalized_headers & known_email_headers:
        return normalized_text

    def _safe_email_cell(value: str) -> str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        raw = re.sub(r"^\s*mailto:\s*", "", raw, flags=re.IGNORECASE)
        matches = _PARSER_EMAIL_RE.findall(raw)
        if len(matches) > 1:
            return ""
        candidate = matches[0] if matches else raw
        candidate = candidate.strip().strip("<>()[]{}\"'")
        candidate = candidate.rstrip(".,;:!?")
        if candidate.count("@") != 1:
            return ""
        local, domain = candidate.split("@", 1)
        if not local or not domain:
            return ""
        return f"{local}@{domain.lower()}"

    swapped = io.StringIO()
    writer = csv.writer(swapped, delimiter=",", lineterminator="\n")
    writer.writerow(["Email", "FirstName"])

    def _emit(email: str, first_name: str = "") -> None:
        writer.writerow([email, first_name])

    reader = csv.reader(non_empty_lines, delimiter=delimiter)
    for row in reader:
        cells = [str(cell or "").strip() for cell in row if str(cell or "").strip()]
        if not cells:
            continue
        if len(cells) == 1:
            email = _safe_email_cell(cells[0])
            if email:
                _emit(email, "")
            else:
                _emit("", cells[0])
            continue
        if len(cells) == 2:
            first_email = _safe_email_cell(cells[0])
            second_email = _safe_email_cell(cells[1])
            if first_email and not second_email:
                _emit(first_email, cells[1])
                continue
            if second_email and not first_email:
                _emit(second_email, cells[0])
                continue
            if first_email and second_email:
                _emit(first_email, "")
                if second_email != first_email:
                    _emit(second_email, "")
                continue
            _emit("", cells[0])
            continue
        extracted = [_safe_email_cell(cell) for cell in cells]
        emails = [email for email in extracted if email]
        if emails and len(emails) == len(cells):
            for email in emails:
                _emit(email, "")
            continue
        if emails:
            _emit(emails[0], "")
            continue
        _emit("", cells[0])

    normalized_rows = swapped.getvalue()
    if not normalized_rows.endswith("\n"):
        normalized_rows += "\n"
    return normalized_rows


def _load_dashboard_auto_start_state() -> dict[str, str]:
    state = {
        "sendgrid_last_started_local_date": "",
        "sendgrid_last_attempt_utc": "",
        "private_jc_last_started_local_date": "",
        "private_jc_last_attempt_utc": "",
        "private_jc_recovery_last_attempt_utc": "",
        "updated_at_utc": "",
    }
    if not DASHBOARD_AUTO_START_STATE_PATH.exists():
        return state
    try:
        raw = json.loads(DASHBOARD_AUTO_START_STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return state
    if not isinstance(raw, dict):
        return state
    for key in state:
        state[key] = str(raw.get(key) or "")
    return state


def _save_dashboard_auto_start_state(state: dict[str, str]) -> None:
    payload = {
        **state,
        "updated_at_utc": datetime.now(timezone.utc).isoformat(),
    }
    DASHBOARD_AUTO_START_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = DASHBOARD_AUTO_START_STATE_PATH.with_suffix(f".{os.getpid()}.tmp")
    tmp_path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
    tmp_path.replace(DASHBOARD_AUTO_START_STATE_PATH)


def _load_dashboard_timer_state() -> dict[str, str]:
    state = {
        "private_jc_recovery_start_at_utc": "",
        "private_jc_recovery_note": "",
        "updated_at_utc": "",
    }
    if not DASHBOARD_TIMER_STATE_PATH.exists():
        return state
    try:
        raw = json.loads(DASHBOARD_TIMER_STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return state
    if not isinstance(raw, dict):
        return state
    for key in state:
        state[key] = str(raw.get(key) or "")
    return state


def _save_dashboard_timer_state(state: dict[str, str]) -> None:
    payload = {
        **state,
        "updated_at_utc": datetime.now(timezone.utc).isoformat(),
    }
    DASHBOARD_TIMER_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = DASHBOARD_TIMER_STATE_PATH.with_suffix(f".{os.getpid()}.tmp")
    tmp_path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
    tmp_path.replace(DASHBOARD_TIMER_STATE_PATH)


def _clear_dashboard_recovery_timer() -> None:
    _save_dashboard_timer_state(
        {
            "private_jc_recovery_start_at_utc": "",
            "private_jc_recovery_note": "",
        }
    )


def _parse_local_trigger_time(raw: object) -> tuple[int, int]:
    text = str(raw or "").strip()
    try:
        hour_text, minute_text = text.split(":", 1)
        hour = int(hour_text)
        minute = int(minute_text)
    except Exception:
        return 18, 0
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return 18, 0
    return hour, minute


def _auto_start_due(now_local: datetime, trigger_text: object) -> bool:
    hour, minute = _parse_local_trigger_time(trigger_text)
    return (now_local.hour, now_local.minute) >= (hour, minute)


def _retry_due(last_attempt_utc: str) -> bool:
    text = str(last_attempt_utc or "").strip()
    if not text:
        return True
    try:
        previous = datetime.fromisoformat(text)
    except Exception:
        return True
    if previous.tzinfo is None:
        previous = previous.replace(tzinfo=timezone.utc)
    return datetime.now(timezone.utc) - previous >= timedelta(minutes=AUTO_START_RETRY_MINUTES)


def _format_local_offset(dt: datetime) -> str:
    offset = dt.utcoffset() or timedelta()
    total_minutes = int(offset.total_seconds() // 60)
    sign = "+" if total_minutes >= 0 else "-"
    total_minutes = abs(total_minutes)
    hours, minutes = divmod(total_minutes, 60)
    return f"{sign}{hours:02d}:{minutes:02d}"


def _format_local_label(dt: datetime) -> str:
    local_dt = dt.astimezone()
    return f"{local_dt.strftime('%Y-%m-%d %H:%M')} {_format_local_offset(local_dt)}"


def _format_local_clock(dt: datetime) -> str:
    local_dt = dt.astimezone()
    return f"{local_dt.strftime('%H:%M')} {_format_local_offset(local_dt)}"


def _next_local_run(now_local: datetime, trigger_text: object) -> datetime:
    hour, minute = _parse_local_trigger_time(trigger_text)
    target = now_local.replace(hour=hour, minute=minute, second=0, microsecond=0)
    if target <= now_local:
        target += timedelta(days=1)
    return target


def _build_automation_status() -> dict[str, object]:
    run_settings = load_dashboard_run_settings()
    auto_state = _load_dashboard_auto_start_state()
    timer_state = _load_dashboard_timer_state()
    now_local = datetime.now().astimezone()
    now_utc = datetime.now(timezone.utc)

    sendgrid_next = _next_local_run(now_local, run_settings.get("auto_start_sendgrid_local_time"))
    jc_daily_next = _next_local_run(now_local, run_settings.get("auto_start_private_jc_local_time"))

    jc_configured_cooldown = max(0, int(PROFILES.get(PRIVATE_BOUNCE_PROFILE, {}).get("cooldown_seconds") or 0))
    jc_pacing = provider_pacing_status(PRIVATE_BOUNCE_PROFILE, "private", jc_configured_cooldown, now=now_utc)

    recovery_target_utc = None
    if jc_pacing.get("recovery_pending"):
        recovery_raw = str(jc_pacing.get("cooldown_until_utc") or "").strip()
    else:
        recovery_raw = str(timer_state.get("private_jc_recovery_start_at_utc") or "").strip()
    if recovery_raw:
        try:
            recovery_target_utc = datetime.fromisoformat(recovery_raw)
            if recovery_target_utc.tzinfo is None:
                recovery_target_utc = recovery_target_utc.replace(tzinfo=timezone.utc)
            else:
                recovery_target_utc = recovery_target_utc.astimezone(timezone.utc)
        except Exception:
            recovery_target_utc = None

    recovery_active = False
    recovery_remaining_seconds = 0
    if recovery_target_utc and not _profile_runtime_active(PRIVATE_BOUNCE_PROFILE):
        recovery_active = True
        recovery_remaining_seconds = max(0, int((recovery_target_utc - now_utc).total_seconds()))

    return {
        "local_timezone_offset": _format_local_offset(now_local),
        "sendgrid_daily": {
            "enabled": bool(run_settings.get("auto_start_sendgrid_enabled")),
            "local_time": str(run_settings.get("auto_start_sendgrid_local_time") or ""),
            "next_run_utc": sendgrid_next.astimezone(timezone.utc).isoformat(),
            "next_run_local_label": _format_local_label(sendgrid_next),
            "next_run_local_clock": _format_local_clock(sendgrid_next),
            "remaining_seconds": max(0, int((sendgrid_next - now_local).total_seconds())),
            "last_started_local_date": str(auto_state.get("sendgrid_last_started_local_date") or ""),
        },
        "private_jc_daily": {
            "enabled": bool(run_settings.get("auto_start_private_jc_enabled")),
            "local_time": str(run_settings.get("auto_start_private_jc_local_time") or ""),
            "next_run_utc": jc_daily_next.astimezone(timezone.utc).isoformat(),
            "next_run_local_label": _format_local_label(jc_daily_next),
            "next_run_local_clock": _format_local_clock(jc_daily_next),
            "remaining_seconds": max(0, int((jc_daily_next - now_local).total_seconds())),
            "last_started_local_date": str(auto_state.get("private_jc_last_started_local_date") or ""),
        },
        "private_jc_recovery": {
            "active": recovery_active,
            "target_utc": recovery_target_utc.isoformat() if recovery_target_utc else "",
            "target_local_label": _format_local_label(recovery_target_utc.astimezone()) if recovery_target_utc else "",
            "target_local_clock": _format_local_clock(recovery_target_utc.astimezone()) if recovery_target_utc else "",
            "remaining_seconds": recovery_remaining_seconds,
            "note": str(
                jc_pacing.get("recovery_reason")
                or jc_pacing.get("last_throttle_reason")
                or timer_state.get("private_jc_recovery_note")
                or ""
            ),
        },
    }


def _build_live_snapshot(activity_hours: int = 24, tail_lines: int = 12) -> dict[str, object]:
    snapshot = build_dashboard_snapshot(activity_hours=activity_hours, tail_lines=tail_lines)
    snapshot["automation"] = _build_automation_status()
    warm_status = build_warm_private_jc_live_status()
    snapshot["warm_private_jc_status"] = warm_status
    snapshot["warm_private_jc_lane"] = warm_status
    for profile in snapshot.get("profiles", []):
        if str(profile.get("name") or "") != "private_jc_warm":
            continue
        profile["pending_count"] = int(warm_status["queued_remaining_count"])
        profile["pending"] = int(warm_status["queued_remaining_count"])
        profile["run_sent_display"] = int(warm_status["sent_count"])
        profile["sent_today"] = int(warm_status["sent_count"])
        profile["last_email"] = str(warm_status["last_sent_email"])
        profile["last_timestamp"] = str(warm_status["last_sent_timestamp"])
        profile["runtime_state"] = "running" if warm_status["running"] else "stopped"
    return snapshot


def _read_csv_dict_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    try:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    except (OSError, csv.Error):
        return []


def _read_worker_events(path: Path) -> list[dict[str, object]]:
    if not path.exists():
        return []
    events: list[dict[str, object]] = []
    try:
        with path.open(encoding="utf-8") as handle:
            for line in handle:
                try:
                    payload = json.loads(line)
                except (TypeError, json.JSONDecodeError):
                    continue
                if isinstance(payload, dict):
                    events.append(payload)
    except OSError:
        return []
    return events


def build_warm_private_jc_live_status() -> dict[str, object]:
    profile = PROFILES.get("private_jc_warm", {})
    queue_path = settings.SHARDS_DIR / str(profile.get("csv") or "recipients_private_jc_warm.csv")
    log_path = settings.LOGS_DIR / str(profile.get("log") or "private_jc_warm_log.csv")
    worker_path = settings.LOGS_DIR / "private_jc_warm_log_worker.jsonl"
    lane = warm_private_jc_lane_status()
    queue_rows = _read_csv_dict_rows(queue_path)
    sent_rows = [
        row for row in _read_csv_dict_rows(log_path)
        if str(row.get("Status") or "").strip().upper() == "SENT"
    ]
    worker_events = _read_worker_events(worker_path)
    last_sent = sent_rows[-1] if sent_rows else {}
    last_worker = worker_events[-1] if worker_events else {}
    last_worker_event = str(last_worker.get("event_type") or last_worker.get("event") or "").strip()
    last_worker_reason = str(last_worker.get("reason") or last_worker.get("message") or "").strip()
    running = "private_jc_warm" in _active_dashboard_profiles()
    remaining = len(queue_rows)
    sent_count = len(sent_rows)
    confirmed = bool(lane.get("confirmed"))
    blocked = bool(
        remaining > 0
        and not running
        and last_worker_reason == "queue_exhausted_no_eligible_rows"
    )
    if running:
        state = "Running"
    elif blocked:
        state = "Blocked"
    elif sent_count > 0 and remaining == 0:
        state = "Complete"
    elif sent_count > 0 and remaining > 0:
        state = "Partial"
    elif confirmed and remaining > 0:
        state = "Ready"
    elif remaining > 0:
        state = "Not confirmed"
    else:
        state = "No queue"

    timeline: list[dict[str, str]] = []
    for row in sent_rows:
        timeline.append({
            "type": "SENT",
            "timestamp": str(row.get("TimestampUTC") or row.get("Timestamp") or row.get("timestamp") or ""),
            "email": str(row.get("Email") or row.get("email") or ""),
            "reason": "",
        })
    for event in worker_events:
        event_type = str(event.get("event_type") or event.get("event") or "").strip().upper()
        reason = str(event.get("reason") or event.get("message") or "").strip()
        if event_type not in {"START", "DONE"} and reason != "queue_exhausted_no_eligible_rows":
            continue
        timeline.append({
            "type": event_type or "DONE",
            "timestamp": str(event.get("ts_utc") or event.get("timestamp") or event.get("timestamp_utc") or event.get("created_at_utc") or ""),
            "email": str(event.get("email") or ""),
            "reason": reason,
        })
    timeline.sort(key=lambda item: item.get("timestamp") or "")

    original_count = int(lane.get("confirmed_rows") or 0)
    return {
        **lane,
        "profile": "private_jc_warm",
        "confirmed": confirmed,
        "running": running,
        "blocked": blocked,
        "state": state,
        "queue_path": str(queue_path),
        "log_path": str(log_path),
        "worker_log_path": str(worker_path),
        "queued_remaining_count": remaining,
        "remaining": remaining,
        "sent_count": sent_count,
        "cap": int(profile.get("max_total") or 10),
        "original_count": original_count,
        "ready_original_count": original_count or sent_count + remaining,
        "last_sent_email": str(last_sent.get("Email") or last_sent.get("email") or ""),
        "last_sent_timestamp": str(last_sent.get("TimestampUTC") or last_sent.get("Timestamp") or last_sent.get("timestamp") or ""),
        "next_queued_email": str(
            (queue_rows[0].get("Email") or queue_rows[0].get("AuthorEmail") or "") if queue_rows else ""
        ),
        "last_worker_event": last_worker_event,
        "last_worker_reason": last_worker_reason,
        "timeline": timeline[-10:],
    }


def _active_dashboard_profiles() -> set[str]:
    active_states = {"starting", "running", "cooldown", "sleeping"}
    active: set[str] = set()
    try:
        snapshots = runtime_control.list_sender_snapshots(tail_lines=6)
    except Exception:
        return active
    for snapshot in snapshots:
        name = str(getattr(snapshot, "name", "") or "")
        state = str(getattr(snapshot, "runtime_state", "") or "")
        dead = bool(getattr(snapshot, "tmux_dead", False))
        if name and not dead and state in active_states:
            active.add(name)
    return active


def _run_dashboard_daily_auto_start_once() -> None:
    settings_payload = load_dashboard_run_settings()
    now_local = datetime.now().astimezone()
    today = now_local.date().isoformat()
    now_utc_iso = datetime.now(timezone.utc).isoformat()
    state = _load_dashboard_auto_start_state()
    active_profiles = _active_dashboard_profiles()
    dirty = False

    if bool(settings_payload.get("auto_start_sendgrid_enabled")) and _auto_start_due(now_local, settings_payload.get("auto_start_sendgrid_local_time")):
        if state.get("sendgrid_last_started_local_date") != today:
            if all(name in active_profiles for name in SENDGRID_PROFILES):
                state["sendgrid_last_started_local_date"] = today
                dirty = True
            elif _retry_due(state.get("sendgrid_last_attempt_utc", "")):
                state["sendgrid_last_attempt_utc"] = now_utc_iso
                dirty = True
                all_ready = True
                for profile_name in SENDGRID_PROFILES:
                    if profile_name in active_profiles:
                        continue
                    ok, _message = runtime_control.start_sender(profile_name)
                    if ok:
                        active_profiles.add(profile_name)
                        continue
                    all_ready = False
                if all_ready and all(name in active_profiles for name in SENDGRID_PROFILES):
                    state["sendgrid_last_started_local_date"] = today
                    dirty = True

    if bool(settings_payload.get("auto_start_private_jc_enabled")) and _auto_start_due(now_local, settings_payload.get("auto_start_private_jc_local_time")):
        if state.get("private_jc_last_started_local_date") != today:
            if PRIVATE_BOUNCE_PROFILE in active_profiles:
                state["private_jc_last_started_local_date"] = today
                dirty = True
            elif _retry_due(state.get("private_jc_last_attempt_utc", "")):
                state["private_jc_last_attempt_utc"] = now_utc_iso
                dirty = True
                ok, _message = runtime_control.start_sender(PRIVATE_BOUNCE_PROFILE)
                if ok or _profile_runtime_active(PRIVATE_BOUNCE_PROFILE):
                    state["private_jc_last_started_local_date"] = today
                    dirty = True

    if dirty:
        _save_dashboard_auto_start_state(state)


def _run_private_jc_recovery_auto_start_once() -> None:
    now_utc = datetime.now(timezone.utc)
    state = _load_dashboard_auto_start_state()
    pacing = provider_pacing_status(
        PRIVATE_BOUNCE_PROFILE,
        "private",
        max(0, int(PROFILES.get(PRIVATE_BOUNCE_PROFILE, {}).get("cooldown_seconds") or 0)),
        now=now_utc,
    )

    if bool(pacing.get("recovery_pending")):
        if _profile_runtime_active(PRIVATE_BOUNCE_PROFILE):
            mark_recovery_started(PRIVATE_BOUNCE_PROFILE, now=now_utc)
            return
        if bool(pacing.get("cooldown_active")):
            return
        if not _retry_due(state.get("private_jc_recovery_last_attempt_utc", "")):
            return
        state["private_jc_recovery_last_attempt_utc"] = now_utc.isoformat()
        _save_dashboard_auto_start_state(state)
        ok, _message = runtime_control.start_sender(PRIVATE_BOUNCE_PROFILE)
        if ok or _profile_runtime_active(PRIVATE_BOUNCE_PROFILE):
            mark_recovery_started(PRIVATE_BOUNCE_PROFILE, now=now_utc)
        return

    timer_state = _load_dashboard_timer_state()
    recovery_raw = str(timer_state.get("private_jc_recovery_start_at_utc") or "").strip()
    if not recovery_raw:
        return
    try:
        recovery_target_utc = datetime.fromisoformat(recovery_raw)
    except Exception:
        _clear_dashboard_recovery_timer()
        return
    if recovery_target_utc.tzinfo is None:
        recovery_target_utc = recovery_target_utc.replace(tzinfo=timezone.utc)
    else:
        recovery_target_utc = recovery_target_utc.astimezone(timezone.utc)
    if recovery_target_utc > now_utc or _profile_runtime_active(PRIVATE_BOUNCE_PROFILE):
        return
    if not _retry_due(state.get("private_jc_recovery_last_attempt_utc", "")):
        return
    state["private_jc_recovery_last_attempt_utc"] = now_utc.isoformat()
    _save_dashboard_auto_start_state(state)
    ok, _message = runtime_control.start_sender(PRIVATE_BOUNCE_PROFILE)
    if ok or _profile_runtime_active(PRIVATE_BOUNCE_PROFILE):
        _clear_dashboard_recovery_timer()


def _run_background_automation_once() -> None:
    try:
        _run_dashboard_daily_auto_start_once()
    except Exception:
        pass
    try:
        _run_private_jc_recovery_auto_start_once()
    except Exception:
        pass
    try:
        runtime_control.apply_delivery_guards()
    except Exception:
        pass
    if not PRIVATE_BOUNCE_MONITOR_ENABLED:
        return
    profile_active = _profile_runtime_active(PRIVATE_BOUNCE_PROFILE)
    try:
        run_private_bounce_monitor_cycle(
            profile_name=PRIVATE_BOUNCE_PROFILE,
            profile_active=profile_active,
            stop_profile=runtime_control.stop_sender,
            start_profile=runtime_control.start_sender,
        )
    except Exception:
        return


async def _background_automation_loop() -> None:
    while True:
        await asyncio.to_thread(_run_background_automation_once)
        await asyncio.sleep(AUTOMATION_LOOP_SECONDS)


@app.on_event("startup")
async def _startup_background_automation() -> None:
    runtime_audit.write_app_start()
    if getattr(app.state, "automation_task", None) is None:
        app.state.automation_task = asyncio.create_task(_background_automation_loop())
    _resume_pending_important_check_jobs()


@app.on_event("shutdown")
async def _shutdown_background_automation() -> None:
    runtime_audit.write_app_shutdown()
    task = getattr(app.state, "automation_task", None)
    if task is None:
        return
    task.cancel()
    try:
        await task
    except asyncio.CancelledError:
        pass
    app.state.automation_task = None


def _csv_preview(path: Path, limit: int = 8) -> dict[str, object]:
    if not path.exists():
        return {"fieldnames": [], "preview_rows": [], "row_count": 0}
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        reader = csv.DictReader(handle)
        fieldnames = [str(name or "").lstrip("\ufeff") for name in (reader.fieldnames or [])]
        rows: list[dict[str, str]] = []
        total = 0
        for row in reader:
            cleaned = {field: str(row.get(field, "") or "") for field in fieldnames}
            if not any(value.strip() for value in cleaned.values()):
                continue
            total += 1
            if len(rows) < limit:
                rows.append(cleaned)
    return {"fieldnames": fieldnames, "preview_rows": rows, "row_count": total}


@lru_cache(maxsize=1)
def _load_sendgrid_public_key():
    if not SENDGRID_EVENT_PUBLIC_KEY:
        return None
    key_bytes = base64.b64decode(SENDGRID_EVENT_PUBLIC_KEY)
    return serialization.load_der_public_key(key_bytes)


def _verify_sendgrid_signature(raw_body: bytes, signature_b64: str, timestamp: str) -> bool:
    public_key = _load_sendgrid_public_key()
    if public_key is None:
        return True
    signature = base64.b64decode((signature_b64 or "").strip())
    signed_payload = (timestamp or "").encode("utf-8") + raw_body
    try:
        public_key.verify(signature, signed_payload, ec.ECDSA(hashes.SHA256()))
        return True
    except InvalidSignature:
        return False


def _dashboard_auth_enabled() -> bool:
    return bool(settings.DASHBOARD_AUTH_PASSWORD)


def _dashboard_is_authenticated(scope: Request | WebSocket) -> bool:
    session = getattr(scope, "session", None)
    if not isinstance(session, dict):
        session = {}
    return bool(session.get(_AUTH_SESSION_KEY))


def _dashboard_auth_response() -> dict[str, object]:
    return {
        "auth_enabled": _dashboard_auth_enabled(),
        "authenticated": True if not _dashboard_auth_enabled() else False,
        "username": str(settings.DASHBOARD_AUTH_USERNAME or "admin"),
    }


async def _read_upload_bytes_with_limit(file: UploadFile, *, limit: int) -> bytes:
    content = await file.read(limit + 1)
    if len(content) > limit:
        raise ValueError(
            f"Upload too large. Limit is {limit} bytes."
        )
    return content


def _human_upload_limit(limit: int) -> str:
    megabytes = max(1, int(limit or 0) // (1024 * 1024))
    return f"{megabytes} MB"


@app.get("/")
def index() -> FileResponse:
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/auth/status")
def auth_status(request: Request) -> JSONResponse:
    authenticated = bool(request.session.get(_AUTH_SESSION_KEY))
    return JSONResponse(
        {
            "ok": True,
            **_dashboard_auth_response(),
            "authenticated": True if not _dashboard_auth_enabled() else authenticated,
            "username": str(request.session.get("dashboard_username") or settings.DASHBOARD_AUTH_USERNAME or "admin"),
        }
    )


@app.post("/api/auth/login")
def auth_login(payload: DashboardAuthPayload, request: Request) -> JSONResponse:
    if not _dashboard_auth_enabled():
        return JSONResponse(
            {"ok": False, "message": "Dashboard auth is not configured."},
            status_code=503,
        )
    expected_user = str(settings.DASHBOARD_AUTH_USERNAME or "admin")
    expected_pass = str(settings.DASHBOARD_AUTH_PASSWORD or "")
    if not secrets.compare_digest(str(payload.username or ""), expected_user) or not secrets.compare_digest(str(payload.password or ""), expected_pass):
        return JSONResponse({"ok": False, "message": "Invalid dashboard credentials."}, status_code=401)

    request.session[_AUTH_SESSION_KEY] = True
    request.session["dashboard_username"] = expected_user
    request.session["dashboard_authenticated_at_utc"] = iso_utc()
    return JSONResponse({"ok": True, "message": "Signed in.", **_dashboard_auth_response(), "authenticated": True, "username": expected_user})


@app.post("/api/auth/logout")
def auth_logout(request: Request) -> JSONResponse:
    request.session.clear()
    return JSONResponse({"ok": True, "message": "Signed out.", **_dashboard_auth_response()})


@app.get("/api/snapshot")
def snapshot(
    hours: int = Query(default=24, ge=1, le=168),
    tail_lines: int = Query(default=12, ge=4, le=50),
) -> dict[str, object]:
    return _build_live_snapshot(activity_hours=hours, tail_lines=tail_lines)


def _queue_safety_start_block_response(profile_name: str = "") -> JSONResponse | None:
    report = build_dashboard_queue_safety_report()
    if bool(report.get("safe")):
        return None
    reasons = report.get("unsafe_reasons")
    if not isinstance(reasons, list):
        reasons = []
    message = str(report.get("message") or "").strip() or (
        "Recipient queue is not safe to start. Rerun Upload & Check, confirm dispatch, "
        "or rebuild queues from the current campaign source before starting senders."
    )
    snapshot_for_record = _build_live_snapshot()
    _append_campaign_history(
        "start_profile_blocked" if profile_name else "start_all_blocked",
        profile=str(profile_name or "all"),
        snapshot=snapshot_for_record,
        queue_safety=report,
        blocked_reasons=[str(reason) for reason in reasons],
    )
    payload = {
        "ok": False,
        "blocked": True,
        "error": "queue_safety_unsafe",
        "profile": str(profile_name or ""),
        "safety_status": "unsafe",
        "safe": False,
        "reasons": [str(reason) for reason in reasons],
        "queue_safety": report,
        "suggested_fix": "Rerun Upload & Check, confirm dispatch, or rebuild queues from the current campaign source before starting senders.",
        "message": message,
        "snapshot": _build_live_snapshot(),
    }
    return JSONResponse(payload, status_code=409)


def _start_precondition_profiles(profile_name: str = "") -> list[str]:
    requested = str(profile_name or "").strip()
    if requested:
        return [requested]
    return [str(profile) for profile in SENDGRID_PROFILES]


def _queue_safety_provider_for_start(profile_name: str = "") -> str:
    requested = str(profile_name or "").strip()
    if requested:
        cfg = PROFILES.get(requested) or {}
        return "sendgrid" if str(cfg.get("provider") or "").strip().lower() == "sendgrid" else "private_jc"
    return "all"


def _profile_readiness_from_snapshot(snapshot: dict[str, object], profile_name: str) -> dict[str, object]:
    profiles = snapshot.get("profiles")
    if isinstance(profiles, list):
        for profile in profiles:
            if isinstance(profile, dict) and str(profile.get("name") or "") == profile_name:
                readiness = profile.get("message_readiness")
                if isinstance(readiness, dict):
                    return dict(readiness)
    return build_profile_message_readiness(profile_name)


def _profile_provider_block_reason_from_snapshot(snapshot: dict[str, object], profile_name: str) -> str:
    profiles = snapshot.get("profiles")
    profile: dict[str, object] = {}
    if isinstance(profiles, list):
        for candidate in profiles:
            if isinstance(candidate, dict) and str(candidate.get("name") or "") == profile_name:
                profile = candidate
                break
    if not profile:
        return ""
    readiness_label = str(profile.get("readiness_label") or "").strip()
    readiness_tone = str(profile.get("readiness_tone") or "").strip().lower()
    reason_code = str(profile.get("reason_code") or "").strip()
    reason_note = str(profile.get("reason_note") or profile.get("readiness_note") or profile.get("health_note") or "").strip()
    if readiness_tone == "bad" or readiness_label.lower() in {"blocked", "not ready"}:
        note = reason_note or readiness_label or "Provider is not ready."
        return f"Provider readiness for {profile_name} is {readiness_label or 'Blocked'} ({reason_code or 'PROVIDER_BLOCKED'}). {note}"
    return ""


def _state_label_path(value: object) -> Path | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    path = Path(raw)
    if not path.is_absolute():
        path = settings.APP_ROOT / path
    return path


def _path_is_temp_artifact(path: Path) -> bool:
    for part in path.parts:
        lowered = part.lower()
        if lowered.startswith("tmp") or lowered.startswith("pytest-"):
            return True
    return False


def _state_file_mtime(path: Path | None) -> float:
    if path is None:
        return 0.0
    try:
        return path.stat().st_mtime
    except Exception:
        return 0.0


def _queue_safety_report_path(value: object) -> Path | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    path = Path(raw)
    if not path.is_absolute():
        path = settings.APP_ROOT / path
    return path


def _path_exists_nonempty(path: Path | None) -> bool:
    if path is None:
        return False
    try:
        return path.exists() and path.stat().st_size > 0
    except Exception:
        return False


def _queue_safety_has_valid_archived_source(queue_safety: dict[str, object]) -> bool:
    if not bool(queue_safety.get("safe")):
        return False
    source_resolution = str(queue_safety.get("source_resolution") or "").strip()
    archived_resolutions = {
        "latest_queue_rebuild_archived_dispatch_state",
        "latest_queue_rebuild_manifest",
        "latest_dispatch_staged_batch_archive",
    }
    if source_resolution not in archived_resolutions:
        return False
    intended = _queue_safety_report_path(queue_safety.get("intended_source_path"))
    checked = _queue_safety_report_path(queue_safety.get("checked_path"))
    return _path_exists_nonempty(intended) and _path_exists_nonempty(checked)


def _lead_state_start_block_reasons(queue_safety: dict[str, object] | None = None) -> list[str]:
    reasons: list[str] = []
    archived_queue_source_ready = _queue_safety_has_valid_archived_source(queue_safety or {})
    try:
        status = _combined_leads_status()
    except Exception as exc:
        return [f"Lead Op state could not be checked: {exc}"]

    active_check = status.get("active_important_check_job")
    if isinstance(active_check, dict):
        job_id = str(active_check.get("job_id") or "").strip() or "unknown"
        reasons.append(f"Check Leads is still running or stale: {job_id}.")
    if archived_queue_source_ready:
        return reasons

    checks: list[tuple[str, Path | None]] = []
    latest_check = status.get("latest_master_check") if isinstance(status.get("latest_master_check"), dict) else {}
    latest_triage = status.get("latest_lead_triage") or status.get("latest_lead_verify")
    if not isinstance(latest_triage, dict):
        latest_triage = {}
    latest_preview = status.get("latest_auto_dispatch_preview")
    if not isinstance(latest_preview, dict):
        latest_preview = {}

    check_output = _state_label_path(latest_check.get("output_label") if isinstance(latest_check, dict) else "")
    check_rejected = _state_label_path(latest_check.get("rejected_label") if isinstance(latest_check, dict) else "")
    triage_keep = _state_label_path(latest_triage.get("verified_label") or latest_triage.get("keep_path"))
    triage_reject = _state_label_path(latest_triage.get("rejected_label") or latest_triage.get("rejected_path"))
    checks.extend(
        [
            ("checked leads output", check_output),
            ("rejected leads output", check_rejected),
            ("triage keep output", triage_keep),
            ("triage reject output", triage_reject),
        ]
    )

    for label, path in checks:
        if path is None:
            continue
        if _path_is_temp_artifact(path):
            reasons.append(f"{label} points to a temp artifact: {path.name}.")
        if not archived_queue_source_ready and not path.exists():
            reasons.append(f"{label} is missing: {path.name}.")

    check_mtime = _state_file_mtime(check_output)
    triage_mtime = _state_file_mtime(triage_keep)
    if not archived_queue_source_ready and check_mtime and triage_mtime and triage_mtime < check_mtime:
        reasons.append("Triage output is older than the checked leads output.")

    if not archived_queue_source_ready and latest_preview:
        preview_time = _parse_iso_timestamp(
            latest_preview.get("generated_at_utc")
            or latest_preview.get("completed_at_utc")
            or latest_preview.get("created_at_utc")
        )
        if triage_mtime and preview_time is not None and preview_time.timestamp() < triage_mtime:
            reasons.append("Dispatch preview is older than the triage output.")
        elif triage_mtime and preview_time is None:
            reasons.append("Dispatch preview is missing a current timestamp.")

    return reasons


def _build_start_preconditions_report(profile_name: str = "") -> dict[str, object]:
    requested_profile = str(profile_name or "").strip()
    profiles = _start_precondition_profiles(requested_profile)
    snapshot = _build_live_snapshot()
    queue_safety_provider = _queue_safety_provider_for_start(requested_profile)
    queue_safety = build_dashboard_queue_safety_report(queue_safety_provider)
    active_profiles = _active_sender_names()
    active_preview_profiles = _active_preview_names(profiles if requested_profile else None)
    blocked_reasons: list[str] = []
    warning_reasons: list[str] = []

    queue_safe = bool(queue_safety.get("safe"))
    if not queue_safe:
        raw_reasons = queue_safety.get("unsafe_reasons")
        if isinstance(raw_reasons, list) and raw_reasons:
            blocked_reasons.extend(str(reason) for reason in raw_reasons)
        else:
            blocked_reasons.append(str(queue_safety.get("message") or "Recipient queue is unsafe."))

    if requested_profile:
        if requested_profile in active_profiles:
            blocked_reasons.append(f"{requested_profile} is already active.")
        if requested_profile in active_preview_profiles:
            blocked_reasons.append(f"{requested_profile} preview validation is already running.")
    elif active_profiles:
        blocked_reasons.append(f"Active senders are already running: {', '.join(sorted(active_profiles))}.")
    elif active_preview_profiles:
        blocked_reasons.append(
            f"Preview validation is already running: {', '.join(sorted(active_preview_profiles))}."
        )

    readiness_by_profile: dict[str, dict[str, object]] = {}
    readiness_status_by_profile: dict[str, str] = {}
    for profile in profiles:
        readiness = _profile_readiness_from_snapshot(snapshot, profile)
        readiness_by_profile[profile] = readiness
        provider_block = _profile_provider_block_reason_from_snapshot(snapshot, profile)
        if provider_block:
            blocked_reasons.append(provider_block)
        status = str(readiness.get("status") or "NOT RUN").strip().upper() or "NOT RUN"
        readiness_status_by_profile[profile] = status
        if status == "PASS":
            continue
        reason_text = "; ".join(str(reason) for reason in readiness.get("reasons", []) if str(reason or "").strip())
        suffix = f" {reason_text}" if reason_text else ""
        if status in {"NOT RUN", "STALE"}:
            warning_reasons.append(f"Message Readiness for {profile} is {status}.{suffix}")
        else:
            blocked_reasons.append(f"Message Readiness for {profile} is {status}.{suffix}")

    lead_state_reasons = _lead_state_start_block_reasons(queue_safety)
    if lead_state_reasons:
        warning_reasons.extend(f"Next batch prep: {reason}" for reason in lead_state_reasons)
    blocked_reasons = list(dict.fromkeys(reason for reason in blocked_reasons if str(reason or "").strip()))
    warning_reasons = list(dict.fromkeys(reason for reason in warning_reasons if str(reason or "").strip()))

    ok = not blocked_reasons
    queue_safety_status = "safe" if queue_safe else "unsafe"
    message_readiness_status: object
    if requested_profile and profiles:
        message_readiness_status = readiness_status_by_profile.get(profiles[0], "NOT RUN")
    else:
        message_readiness_status = readiness_status_by_profile
    return {
        "ok": ok,
        "blocked": not ok,
        "profile": requested_profile,
        "profiles": profiles,
        "queue_safety": queue_safety,
        "queue_safety_provider": queue_safety_provider,
        "queue_safety_status": queue_safety_status,
        "safety_status": queue_safety_status,
        "message_readiness": readiness_by_profile,
        "message_readiness_status": message_readiness_status,
        "warning_reasons": warning_reasons,
        "warnings": warning_reasons,
        "blocked_reasons": blocked_reasons,
        "reasons": blocked_reasons,
        "suggested_fix": (
            "Run Preview + Validate for each sender, wait for active jobs/senders to finish, "
            "rerun Upload & Check/Fast Triage/Dispatch Preview if stale, then confirm/rebuild queues if needed."
        ),
        "snapshot": snapshot,
    }


def _start_preconditions_block_response(report: dict[str, object]) -> JSONResponse | None:
    if bool(report.get("ok")):
        return None
    profile_name = str(report.get("profile") or "").strip()
    queue_safety = report.get("queue_safety") if isinstance(report.get("queue_safety"), dict) else {}
    reasons = [str(reason) for reason in (report.get("blocked_reasons") or []) if str(reason or "").strip()]
    _append_campaign_history(
        "start_profile_blocked" if profile_name else "start_all_blocked",
        profile=profile_name or "all",
        snapshot=report.get("snapshot") if isinstance(report.get("snapshot"), dict) else _build_live_snapshot(),
        queue_safety=queue_safety,
        blocked_reasons=reasons,
    )
    error = "queue_safety_unsafe" if str(report.get("queue_safety_status") or "") == "unsafe" else "start_preconditions_failed"
    message = "NOT READY / BLOCKED: " + (" ".join(reasons) if reasons else "Start preconditions failed.")
    payload = {
        "ok": False,
        "blocked": True,
        "error": error,
        "profile": profile_name,
        "profiles": report.get("profiles") or [],
        "queue_safety_status": report.get("queue_safety_status") or "unknown",
        "safety_status": report.get("safety_status") or report.get("queue_safety_status") or "unknown",
        "message_readiness_status": report.get("message_readiness_status") or "",
        "warning_reasons": report.get("warning_reasons") or [],
        "warnings": report.get("warnings") or report.get("warning_reasons") or [],
        "blocked_reasons": reasons,
        "reasons": reasons,
        "queue_safety": queue_safety,
        "suggested_fix": report.get("suggested_fix") or "",
        "message": message,
        "snapshot": _build_live_snapshot(),
    }
    return JSONResponse(payload, status_code=409)


def _active_sender_names() -> set[str]:
    names: set[str] = set()
    try:
        names.update(str(item.name) for item in runtime_control.list_active_sender_snapshots(tail_lines=12))
    except Exception:
        pass
    try:
        names.update(detect_running_sender_profiles())
    except Exception:
        pass
    try:
        names.update(active_or_locked_sender_profiles())
    except Exception:
        pass
    return names


def _active_preview_names(profile_names: list[str] | None = None) -> set[str]:
    try:
        return detect_running_preview_profiles(profile_names)
    except Exception:
        return set()


def _preview_validation_reason_counts(summary_path: Path, limit: int = 5) -> list[str]:
    if not summary_path.exists():
        return []
    try:
        text = summary_path.read_text(encoding="utf-8")
    except Exception:
        return []
    reasons: list[str] = []
    for line in text.splitlines():
        cleaned = line.strip()
        if not cleaned.startswith("- "):
            continue
        reason = cleaned[2:].strip()
        if reason and reason.lower() != "none":
            reasons.append(reason)
        if len(reasons) >= limit:
            break
    return reasons


def _append_campaign_history(
    event_type: str,
    *,
    profile: str = "",
    snapshot: dict[str, object] | None = None,
    queue_safety: dict[str, object] | None = None,
    blocked_reasons: list[object] | None = None,
    preview_file: str = "",
    preview_row_count: int | None = None,
    validation_status: str = "",
) -> None:
    try:
        append_campaign_run_history(
            campaign_history_record(
                event_type,
                profile=profile,
                snapshot=snapshot,
                queue_safety=queue_safety,
                blocked_reasons=blocked_reasons,
                preview_file=preview_file,
                preview_row_count=preview_row_count,
                validation_status=validation_status,
            )
        )
    except Exception:
        pass


@app.post("/api/profiles/{profile_name}/preview-validate")
def preview_validate_profile(profile_name: str) -> JSONResponse:
    profile_name = str(profile_name or "").strip()
    if not runtime_control.is_known_profile(profile_name):
        return JSONResponse({"ok": False, "message": f"Unknown profile: {profile_name}"}, status_code=404)
    if profile_name in _active_sender_names():
        return JSONResponse(
            {
                "ok": False,
                "blocked": True,
                "error": "profile_active",
                "profile": profile_name,
                "message": f"Preview validation blocked: {profile_name} is actively sending.",
                "snapshot": _build_live_snapshot(),
            },
            status_code=409,
        )
    active_dispatch = _find_active_dashboard_job(IMPORTANT_LEADS_DISPATCH_JOBS)
    if active_dispatch:
        return JSONResponse(
            {
                "ok": False,
                "blocked": True,
                "error": "dispatch_active",
                "profile": profile_name,
                "job": active_dispatch,
                "message": "Preview validation blocked while dispatch confirm/rebuild is running.",
                "snapshot": _build_live_snapshot(),
            },
            status_code=409,
        )

    python_bin = settings.APP_ROOT / ".venv" / "bin" / "python"
    if not python_bin.exists():
        python_bin = Path("python")
    preview_cmd = [str(python_bin), "send_shard.py", "--profile", profile_name, "--preview_messages"]
    expected_mode = profile_expected_pitch_mode(profile_name)
    validate_cmd = [
        str(python_bin),
        "tools/validate_message_preview.py",
        "--profile",
        profile_name,
        "--pitch-mode",
        expected_mode,
        "--fail-on-errors",
    ]
    preview_path = message_preview_path_for_profile(profile_name)
    validated_path, failed_path, summary_path = message_preview_output_paths(preview_path)
    _append_campaign_history(
        "preview_validate_started",
        profile=profile_name,
        snapshot=_build_live_snapshot(),
        preview_file=str(preview_path),
        validation_status="RUNNING",
    )

    try:
        preview_proc = subprocess.run(
            preview_cmd,
            cwd=settings.APP_ROOT,
            capture_output=True,
            text=True,
            check=False,
            timeout=900,
        )
    except subprocess.TimeoutExpired:
        _append_campaign_history(
            "preview_validate_completed",
            profile=profile_name,
            snapshot=_build_live_snapshot(),
            preview_file=str(preview_path),
            preview_row_count=_count_csv_rows(preview_path),
            validation_status="TIMEOUT",
            blocked_reasons=["Preview generation timed out."],
        )
        return JSONResponse(
            {
                "ok": False,
                "error": "preview_timeout",
                "profile": profile_name,
                "message": "Preview generation timed out before validation could run.",
                "snapshot": _build_live_snapshot(),
            },
            status_code=504,
        )
    if preview_proc.returncode != 0:
        _append_campaign_history(
            "preview_validate_completed",
            profile=profile_name,
            snapshot=_build_live_snapshot(),
            preview_file=str(preview_path),
            preview_row_count=_count_csv_rows(preview_path),
            validation_status="PREVIEW_FAILED",
            blocked_reasons=["Preview generation failed."],
        )
        return JSONResponse(
            {
                "ok": False,
                "error": "preview_failed",
                "profile": profile_name,
                "returncode": int(preview_proc.returncode),
                "message": "Preview generation failed. No email was sent.",
                "snapshot": _build_live_snapshot(),
            },
            status_code=500,
        )

    try:
        validate_proc = subprocess.run(
            validate_cmd,
            cwd=settings.APP_ROOT,
            capture_output=True,
            text=True,
            check=False,
            timeout=300,
        )
    except subprocess.TimeoutExpired:
        _append_campaign_history(
            "preview_validate_completed",
            profile=profile_name,
            snapshot=_build_live_snapshot(),
            preview_file=str(preview_path),
            preview_row_count=_count_csv_rows(preview_path),
            validation_status="TIMEOUT",
            blocked_reasons=["Preview validation timed out."],
        )
        return JSONResponse(
            {
                "ok": False,
                "error": "validation_timeout",
                "profile": profile_name,
                "preview_path": str(preview_path),
                "preview_row_count": _count_csv_rows(preview_path),
                "message": "Preview validation timed out.",
                "snapshot": _build_live_snapshot(),
            },
            status_code=504,
        )

    validation_passed = validate_proc.returncode == 0
    reason_counts = _preview_validation_reason_counts(summary_path)
    result = {
        "profile": profile_name,
        "pitch_mode": expected_mode,
        "preview_path": str(preview_path),
        "preview_row_count": _count_csv_rows(preview_path),
        "validation_passed": validation_passed,
        "validation_status": "PASS" if validation_passed else "FAIL",
        "validation_returncode": int(validate_proc.returncode),
        "validation_reasons": reason_counts,
        "validated_path": str(validated_path),
        "failed_path": str(failed_path),
        "summary_path": str(summary_path),
        "timestamp_utc": iso_utc(),
    }
    history_snapshot = _build_live_snapshot()
    _append_campaign_history(
        "preview_validate_completed",
        profile=profile_name,
        snapshot=history_snapshot,
        preview_file=str(preview_path),
        preview_row_count=int(result["preview_row_count"]),
        validation_status=str(result["validation_status"]),
        blocked_reasons=reason_counts,
    )
    snapshot = _build_live_snapshot()
    return JSONResponse(
        {
            "ok": True,
            "message": (
                f"Preview + validation passed for {profile_name}."
                if validation_passed
                else f"Preview generated but validation failed for {profile_name}."
            ),
            "result": result,
            "snapshot": snapshot,
        }
    )


@app.post("/api/start")
def start() -> JSONResponse:
    _append_campaign_history("start_all_requested", profile="all", snapshot=_build_live_snapshot())
    preconditions = _build_start_preconditions_report()
    blocked = _start_preconditions_block_response(preconditions)
    if blocked is not None:
        return blocked
    ok, message = runtime_control.start_all_senders()
    time.sleep(0.6)
    snapshot = _build_live_snapshot()
    partially_started = str(message or "").startswith("PARTIALLY_STARTED")
    _append_campaign_history(
        "start_all_started" if ok else "start_all_partially_started" if partially_started else "start_all_blocked",
        profile="all",
        snapshot=snapshot,
        blocked_reasons=[] if ok else [message],
    )
    return JSONResponse({
        "ok": ok,
        "status": "STARTED" if ok else "PARTIALLY_STARTED" if partially_started else "BLOCKED",
        "message": message,
        "warnings": preconditions.get("warning_reasons") or [],
        "preconditions": preconditions,
        "snapshot": _build_live_snapshot(),
    })


@app.post("/api/start/{profile_name}")
def start_profile(profile_name: str) -> JSONResponse:
    if not runtime_control.is_known_profile(profile_name):
        return JSONResponse({"ok": False, "message": f"Unknown profile: {profile_name}"}, status_code=404)
    _append_campaign_history("start_profile_requested", profile=profile_name, snapshot=_build_live_snapshot())
    if profile_name == "private_jc_warm":
        lane = warm_private_jc_lane_status()
        if not bool(lane.get("ready")):
            warm_error = str(lane.get("integrity_reason") or "warm_confirmation_required")
            return JSONResponse(
                {
                    "ok": False,
                    "blocked": True,
                    "error": warm_error,
                    "message": str(lane.get("message") or "Confirm Warm Private JC before starting."),
                    "warm_private_jc_lane": lane,
                    "snapshot": _build_live_snapshot(),
                },
                status_code=409,
            )
        if profile_name in _active_sender_names():
            return JSONResponse(
                {"ok": False, "blocked": True, "error": "profile_active", "message": "Warm Private JC is already running."},
                status_code=409,
            )
        ok, message = runtime_control.start_sender(profile_name)
        time.sleep(0.6)
        return JSONResponse({
            "ok": ok,
            "message": message,
            "warm_private_jc_lane": warm_private_jc_lane_status(),
            "snapshot": _build_live_snapshot(),
        })
    preconditions = _build_start_preconditions_report(profile_name=profile_name)
    blocked = _start_preconditions_block_response(preconditions)
    if blocked is not None:
        return blocked
    ok, message = runtime_control.start_sender(profile_name)
    time.sleep(0.6)
    snapshot = _build_live_snapshot()
    _append_campaign_history(
        "start_profile_started" if ok else "start_profile_blocked",
        profile=profile_name,
        snapshot=snapshot,
        blocked_reasons=[] if ok else [message],
    )
    return JSONResponse({
        "ok": ok,
        "message": message,
        "warnings": preconditions.get("warning_reasons") or [],
        "preconditions": preconditions,
        "snapshot": _build_live_snapshot(),
    })


@app.post("/api/stop")
def stop() -> JSONResponse:
    ok, message = runtime_control.stop_all_senders()
    return JSONResponse({"ok": ok, "message": message, "snapshot": _build_live_snapshot()})


@app.post("/api/stop/{profile_name}")
def stop_profile(profile_name: str) -> JSONResponse:
    if not runtime_control.is_known_profile(profile_name):
        return JSONResponse({"ok": False, "message": f"Unknown profile: {profile_name}"}, status_code=404)
    ok, message = runtime_control.stop_sender(profile_name)
    return JSONResponse({"ok": ok, "message": message, "snapshot": _build_live_snapshot()})


@app.post("/api/archive-reset-logs")
def archive_reset_logs() -> JSONResponse:
    ok, message = runtime_control.archive_reset_logs()
    return JSONResponse({"ok": ok, "message": message, "snapshot": _build_live_snapshot()})


@app.post("/api/settings/send-cap")
def update_send_cap(payload: SendCapPayload) -> JSONResponse:
    settings = save_dashboard_send_cap_per_profile(payload.send_cap_per_profile)
    cap = int(settings.get("send_cap_per_profile") or payload.send_cap_per_profile)
    return JSONResponse(
        {
            "ok": True,
            "message": f"Dashboard SendGrid target saved: {cap} emails from 6 PM to 12 PM.",
            "snapshot": _build_live_snapshot(),
        }
    )


@app.post("/api/leads/upload")
async def upload_leads(file: UploadFile = File(...)) -> JSONResponse:
    filename = (file.filename or "").strip()
    if not filename:
        return JSONResponse({"ok": False, "message": "Missing upload filename."}, status_code=400)
    try:
        content = await _read_upload_bytes_with_limit(file, limit=settings.DASHBOARD_MAX_UPLOAD_BYTES)
    except ValueError as exc:
        return JSONResponse(
            {
                "ok": False,
                "message": str(exc),
                "error": "UPLOAD_TOO_LARGE",
                "details": {"max_upload_bytes": settings.DASHBOARD_MAX_UPLOAD_BYTES},
            },
            status_code=413,
        )
    if not content:
        return JSONResponse({"ok": False, "message": "Uploaded file is empty."}, status_code=400)
    try:
        preview = save_uploaded_csv(filename, content)
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Upload failed: {exc}"}, status_code=400)
    return JSONResponse(
        {
            "ok": True,
            "message": f"Uploaded {preview['original_filename']} ({preview['row_count']} row(s)).",
            "upload": preview,
            "status": shard_status(),
        }
    )


@app.post("/api/leads/check-important/upload")
async def check_important_leads_upload(
    file: UploadFile | None = File(None),
    client_selected_filename: str = Form(""),
    client_selected_size_bytes: str = Form(""),
    client_selected_extension: str = Form(""),
    output_path: str = Form(""),
    rejected_path: str = Form(""),
    intake_mode: str = Form("standard"),
    upload_type: str = Form("cold"),
) -> JSONResponse:
    current_paths = important_leads_path_state()
    resolved_output_path = _resolve_dashboard_csv_path_or_default(
        output_path or current_paths["output_path"],
        IMPORTANT_LEADS_OUTPUT,
    )
    resolved_rejected_path = _resolve_dashboard_csv_path_or_default(
        rejected_path or current_paths["rejected_path"],
        IMPORTANT_LEADS_REJECTED,
    )
    if file is None:
        return JSONResponse(
            {
                "ok": False,
                "message": "Missing uploaded file. Upload CSV files only.",
                "error": "UPLOAD_FILE_REQUIRED",
                "details": {"allowed_extensions": [".csv"]},
                "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
            },
            status_code=400,
        )
    filename = Path(str(file.filename or "").strip()).name
    if not filename:
        return JSONResponse(
            {
                "ok": False,
                "message": "Missing upload filename. Upload CSV files only.",
                "error": "UPLOAD_FILENAME_REQUIRED",
                "details": {"allowed_extensions": [".csv"]},
                "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
            },
            status_code=400,
        )
    extension = Path(filename).suffix.lower()
    if extension not in {".csv", ".xlsx"}:
        return JSONResponse(
            {
                "ok": False,
                "message": f"Unsupported upload file type: {extension or '[none]'}. Only .csv and .xlsx files are supported for upload checks.",
                "error": "UPLOAD_UNSUPPORTED_FILE_TYPE",
                "details": {
                    "allowed_extensions": [".csv", ".xlsx"],
                    "server_received_filename": filename,
                    "server_received_extension": extension or "",
                },
                "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
            },
            status_code=415,
        )
    selected_filename = Path(str(client_selected_filename or "").strip()).name
    if selected_filename and selected_filename != filename:
        return JSONResponse(
            {
                "ok": False,
                "message": (
                    "Selected filename mismatch. "
                    f"Selected {selected_filename}, server received {filename}."
                ),
                "error": "UPLOAD_FILENAME_MISMATCH",
                "details": {
                    "selected_filename": selected_filename,
                    "server_received_filename": filename,
                    "allowed_extensions": [".csv", ".xlsx"],
                },
                "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
            },
            status_code=400,
        )
    try:
        selected_size_bytes = max(0, int(str(client_selected_size_bytes or "0").strip() or 0))
    except Exception:
        selected_size_bytes = 0
    max_upload_bytes = IMPORTANT_LEADS_CHECK_UPLOAD_MAX_BYTES
    if selected_size_bytes > max_upload_bytes:
        return JSONResponse(
            {
                "ok": False,
                "message": f"Upload too large. Upload CSV/XLSX files up to {_human_upload_limit(max_upload_bytes)}.",
                "error": "UPLOAD_TOO_LARGE",
                "details": {"max_upload_bytes": max_upload_bytes, "max_upload_megabytes": max_upload_bytes // (1024 * 1024)},
                "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
            },
            status_code=413,
        )
    try:
        content = await _read_upload_bytes_with_limit(file, limit=max_upload_bytes)
    except ValueError as exc:
        return JSONResponse(
            {
                "ok": False,
                "message": f"{exc} Upload CSV/XLSX files up to {_human_upload_limit(max_upload_bytes)}.",
                "error": "UPLOAD_TOO_LARGE",
                "details": {"max_upload_bytes": max_upload_bytes, "max_upload_megabytes": max_upload_bytes // (1024 * 1024)},
                "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
            },
            status_code=413,
        )
    if not content:
        return JSONResponse(
            {
                "ok": False,
                "message": "Uploaded file is empty.",
                "error": "UPLOAD_FILE_EMPTY",
                "details": {
                    "allowed_extensions": [".csv", ".xlsx"],
                    "server_received_filename": filename,
                    "server_received_extension": extension,
                },
                "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
            },
            status_code=400,
        )
    IMPORTANT_LEADS_CHECK_RUNS.mkdir(parents=True, exist_ok=True)
    effective_input_path = IMPORTANT_LEADS_CHECK_RUNS / f"leadschecker_{timestamp_slug()}.csv"
    try:
        normalized_extension, normalized_text, source_sheet = _normalize_uploaded_check_file(filename, content)
    except ValueError as exc:
        return JSONResponse(
            {
                "ok": False,
                "message": str(exc),
                "error": "UPLOAD_WORKBOOK_INVALID" if extension == ".xlsx" else "UPLOAD_FILE_INVALID",
                "details": {
                    "allowed_extensions": [".csv", ".xlsx"],
                    "server_received_filename": filename,
                    "server_received_extension": extension,
                },
                "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
            },
            status_code=400,
    )
    effective_input_path.write_text(normalized_text, encoding="utf-8")
    settings.secure_private_file(effective_input_path)
    total_input_rows = _count_csv_rows(effective_input_path)
    job = _start_important_check_job(
        input_path=effective_input_path,
        output_path=resolved_output_path,
        rejected_path=resolved_rejected_path,
        effective_input_path=effective_input_path,
        source_label=filename,
        source_mode="uploaded_file",
        original_uploaded_filename=filename,
        server_received_filename=filename,
        selected_filename=selected_filename or filename,
        selected_size_bytes=selected_size_bytes,
        selected_extension=str(client_selected_extension or normalized_extension or extension or "").strip().lower() or normalized_extension or extension,
        source_sheet=source_sheet,
        intake_mode=_normalize_intake_mode(intake_mode),
        total_input_rows=total_input_rows,
        upload_type=upload_type,
    )
    return JSONResponse(
        {
            "ok": True,
            "message": f"Queued upload check for {filename} as {job['job_id']}.",
            "server_received_filename": filename,
            "selected_filename": selected_filename or filename,
            "selected_size_bytes": selected_size_bytes,
            "selected_extension": normalized_extension or extension,
            "job": job,
            "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
        },
        status_code=202,
    )


@app.get("/api/leads/check-important/job/{job_id}")
def get_check_important_leads_job(job_id: str) -> JSONResponse:
    try:
        job = _load_important_check_job(job_id)
    except FileNotFoundError:
        return JSONResponse({"ok": False, "message": f"Check job not found: {job_id}"}, status_code=404)
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Failed to load check job: {exc}"}, status_code=500)
    return JSONResponse({"ok": True, "job": _important_check_job_with_progress(job), "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()}})


@app.get("/api/leads/check-important/active")
def get_active_check_important_leads_job() -> JSONResponse:
    return JSONResponse(
        {
            "ok": True,
            "job": _find_active_important_check_job(),
            "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
        }
    )


@app.post("/api/leads/check-important/warm-preview")
def generate_warm_research_email_preview() -> JSONResponse:
    job = _latest_completed_warm_check_job()
    if not job:
        return JSONResponse(
            {"ok": False, "error": "warm_check_required", "message": "Run a Warm Research upload check first."},
            status_code=404,
        )
    email_ready_path = Path(str(job.get("output_path") or ""))
    if not email_ready_path.exists() or email_ready_path.name != "warm_email_ready.csv":
        return JSONResponse(
            {"ok": False, "error": "warm_email_ready_missing", "message": "The latest warm_email_ready.csv is missing. Re-run the Warm Research upload check."},
            status_code=409,
        )
    preview_path = email_ready_path.with_name("warm_email_preview.csv")
    try:
        preview = generate_warm_email_preview(
            email_ready_path=email_ready_path,
            preview_path=preview_path,
        )
    except ImportantLeadsCheckError as exc:
        return JSONResponse(
            {"ok": False, "error": exc.code, "message": exc.message, "details": exc.details},
            status_code=400,
        )
    except Exception as exc:
        return JSONResponse(
            {"ok": False, "error": "warm_preview_failed", "message": f"Warm draft preview failed: {exc}"},
            status_code=500,
        )

    updated_job = dict(job)
    updated_check = dict(updated_job.get("check") or {})
    updated_check.update({
        "warm_email_preview_rows": int(preview.get("warm_email_preview_rows") or 0),
        "warm_email_preview_label": str(preview.get("output_label") or ""),
        "warm_email_preview_generated_at_utc": str(preview.get("generated_at_utc") or ""),
    })
    updated_job["check"] = updated_check
    updated_job["warm_email_preview"] = preview
    updated_job["warm_email_preview_path"] = str(preview_path)
    _save_important_check_job(updated_job)
    return JSONResponse(
        {
            "ok": True,
            "message": str(preview.get("message") or "Warm draft preview generated."),
            "preview": preview,
            "warm_check": updated_check,
            "status": _combined_leads_status(),
        }
    )


@app.post("/api/leads/check-important/warm-confirm")
def confirm_warm_research_private_jc() -> JSONResponse:
    job = _latest_completed_warm_check_job()
    if not job:
        return JSONResponse(
            {"ok": False, "error": "warm_check_required", "message": "Run a Warm Research upload check first."},
            status_code=404,
        )
    preview_path = Path(str(job.get("warm_email_preview_path") or ""))
    if not preview_path.exists() or preview_path.name != "warm_email_preview.csv":
        return JSONResponse(
            {"ok": False, "error": "warm_preview_required", "message": "Generate the current Warm Draft Preview before confirming Warm Private JC."},
            status_code=409,
        )
    try:
        confirmation = confirm_warm_private_jc_preview(preview_path=preview_path)
    except Exception as exc:
        return JSONResponse(
            {"ok": False, "blocked": True, "error": "warm_confirm_blocked", "message": str(exc)},
            status_code=409,
        )
    updated_job = dict(job)
    updated_check = dict(updated_job.get("check") or {})
    updated_check.update({
        "warm_private_jc_confirmed": True,
        "warm_private_jc_confirmed_rows": int(confirmation.get("row_count") or 0),
        "warm_private_jc_confirmation_id": str(confirmation.get("confirmation_id") or ""),
    })
    updated_job["check"] = updated_check
    updated_job["warm_private_jc_confirmation"] = confirmation
    _save_important_check_job(updated_job)
    return JSONResponse({
        "ok": True,
        "message": str(confirmation.get("message") or "Warm Private JC confirmed."),
        "confirmation": confirmation,
        "warm_check": updated_check,
        "status": _combined_leads_status(),
    })


@app.post("/api/leads/clean")
def clean_leads(payload: CleanLeadsPayload) -> JSONResponse:
    mapping = payload.mapping.model_dump() if payload.mapping else None
    try:
        report = clean_uploaded_leads(
            upload_filename=payload.upload_filename,
            mapping=mapping,
            remove_invalid_emails=payload.remove_invalid_emails,
            dedupe_by_email=payload.dedupe_by_email,
            remove_suppressed=payload.remove_suppressed,
            drop_role_emails=payload.drop_role_emails,
            exclude_domains=payload.exclude_domains,
        )
    except FileNotFoundError as exc:
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=404)
    except ImportantLeadsCheckError as exc:
        return JSONResponse(
            {
                "ok": False,
                "message": exc.message,
                "error": exc.code,
                "details": exc.details,
            },
            status_code=400,
        )
    except ValueError as exc:
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Lead clean failed: {exc}"}, status_code=500)
    return JSONResponse(
        {
            "ok": True,
            "message": f"Cleaned leads written to {report['cleaned_filename']} ({report['output_rows']} row(s) kept).",
            "clean": report,
            "status": shard_status(),
        }
    )


@app.post("/api/leads/check-important")
def check_important_leads(payload: ImportantLeadPathsPayload | None = None) -> JSONResponse:
    current_paths = important_leads_path_state()
    input_path = _resolve_dashboard_csv_path(
        payload.input_path if payload else current_paths["input_path"],
        IMPORTANT_LEADS_INPUT,
    )
    output_path = _resolve_dashboard_csv_path(
        payload.output_path if payload else current_paths["output_path"],
        IMPORTANT_LEADS_OUTPUT,
    )
    rejected_path = _resolve_dashboard_csv_path(
        payload.rejected_path if payload else current_paths["rejected_path"],
        IMPORTANT_LEADS_REJECTED,
    )
    effective_input_path = input_path
    input_text = str(payload.input_text or "") if payload else ""
    if input_text.strip():
        pasted_rows = _count_pasted_lead_rows(input_text)
        if pasted_rows > IMPORTANT_LEADS_PASTE_MAX_ROWS:
            return JSONResponse(
                {
                    "ok": False,
                    "error": "PASTE_TOO_LARGE",
                    "message": (
                        f"Paste intake is limited to {IMPORTANT_LEADS_PASTE_MAX_ROWS} rows. "
                        f"This paste has {pasted_rows} rows; use CSV upload for large batches."
                    ),
                    "details": {
                        "paste_rows": pasted_rows,
                        "paste_max_rows": IMPORTANT_LEADS_PASTE_MAX_ROWS,
                    },
                    "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
                },
                status_code=413,
            )
        IMPORTANT_LEADS_CHECK_RUNS.mkdir(parents=True, exist_ok=True)
        effective_input_path = IMPORTANT_LEADS_CHECK_RUNS / f"leadschecker_{timestamp_slug()}.csv"
        normalized_text = _normalize_pasted_leads_csv(input_text)
        effective_input_path.write_text(normalized_text, encoding="utf-8")
    return _check_important_leads_response(
        input_path=input_path,
        output_path=output_path,
        rejected_path=rejected_path,
        effective_input_path=effective_input_path,
        intake_mode=_normalize_intake_mode(payload.intake_mode if payload else "standard"),
    )


@app.post("/api/leads/verify-important")
def verify_important_leads(payload: ImportantLeadVerifyPayload | None = None) -> JSONResponse:
    try:
        mode = str(payload.mode if payload else TRIAGE_MODE_FAST).strip().upper()
        if mode == TRIAGE_MODE_STRICT:
            mode = TRIAGE_MODE_STRICT
        elif mode == TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH:
            mode = TRIAGE_MODE_MANUAL_AUTHOR_RESEARCH
        else:
            mode = TRIAGE_MODE_FAST
        if mode != TRIAGE_MODE_STRICT:
            current_paths = important_leads_triage_path_state()
            default_keep_path = IMPORTANT_LEADS_OUTPUT.with_name("leads_triaged_keep.csv")
            default_rejected_path = IMPORTANT_LEADS_OUTPUT.with_name("leads_triaged_reject.csv")
            default_quarantine_path = IMPORTANT_LEADS_OUTPUT.with_name("leads_triaged_quarantine.csv")
            current_keep = current_paths["keep_path"]
        else:
            current_paths = important_leads_verify_path_state()
            default_keep_path = IMPORTANT_LEADS_OUTPUT.with_name("leads_verified.csv")
            default_rejected_path = IMPORTANT_LEADS_OUTPUT.with_name("leads_verify_rejected.csv")
            default_quarantine_path = IMPORTANT_LEADS_OUTPUT.with_name("leads_quarantine.csv")
            current_keep = current_paths["verified_path"]
        input_path = _resolve_dashboard_csv_path(
            payload.input_path if payload else current_paths["input_path"],
            IMPORTANT_LEADS_OUTPUT,
        )
        if mode != TRIAGE_MODE_STRICT:
            verified_path = _resolve_dashboard_csv_path(
                payload.verified_path if payload else current_keep,
                default_keep_path,
            )
            rejected_path = _resolve_dashboard_csv_path(
                payload.rejected_path if payload else current_paths["rejected_path"],
                default_rejected_path,
            )
            quarantine_path = _resolve_dashboard_csv_path(
                payload.quarantine_path if payload else current_paths["quarantine_path"],
                default_quarantine_path,
            )
        else:
            verified_path = _resolve_dashboard_csv_path(
                payload.verified_path if payload else current_keep,
                default_keep_path,
            )
            rejected_path = _resolve_dashboard_csv_path(
                payload.rejected_path if payload else current_paths["rejected_path"],
                default_rejected_path,
            )
            quarantine_path = _resolve_dashboard_csv_path(
                payload.quarantine_path if payload else current_paths["quarantine_path"],
                default_quarantine_path,
            )
        if mode == TRIAGE_MODE_STRICT:
            save_state(
                important_leads_verify_paths=_important_verify_path_labels_for_state(
                    input_path,
                    verified_path,
                    rejected_path,
                    quarantine_path,
                )
            )
        else:
            save_state(
                important_leads_triage_paths=_important_triage_path_labels_for_state(
                    input_path,
                    verified_path,
                    rejected_path,
                    quarantine_path,
                )
            )
        job = _start_important_verify_job(
            input_path=input_path,
            verified_path=verified_path,
            rejected_path=rejected_path,
            quarantine_path=quarantine_path,
            mode=mode,
        )
    except FileNotFoundError as exc:
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=404)
    except ValueError as exc:
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Lead verify failed: {exc}"}, status_code=500)

    return JSONResponse(
        {
            "ok": True,
            "message": f"Queued {_triage_mode_label(mode)} job {job['job_id']}.",
            "job": job,
            "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
        },
        status_code=202,
    )


@app.get("/api/leads/verify-important/job/{job_id}")
def get_verify_important_leads_job(job_id: str) -> JSONResponse:
    try:
        job = _load_important_verify_job(job_id)
    except FileNotFoundError:
        return JSONResponse({"ok": False, "message": f"Verify job not found: {job_id}"}, status_code=404)
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Failed to load verify job: {exc}"}, status_code=500)
    return JSONResponse({"ok": True, "job": _job_progress_payload(job), "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()}})


@app.get("/api/leads/verify-important/active")
def get_active_verify_important_leads_job() -> JSONResponse:
    return JSONResponse(
        {
            "ok": True,
            "job": _find_active_dashboard_job(IMPORTANT_LEADS_VERIFY_JOBS),
            "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
        }
    )


@app.post("/api/leads/verify-important/job/{job_id}/cancel")
def cancel_verify_important_leads_job(job_id: str) -> JSONResponse:
    try:
        job = _load_important_verify_job(job_id)
    except FileNotFoundError:
        return JSONResponse({"ok": False, "message": f"Verify job not found: {job_id}"}, status_code=404)
    status = str(job.get("status") or "").lower()
    if status in {"completed", "failed", "canceled", "cancelled"}:
        return JSONResponse({"ok": True, "message": f"Verify job already terminal: {status}.", "job": _job_progress_payload(job)})
    job["cancel_requested"] = True
    job["stage"] = "cancel_requested"
    job["phase"] = "cancel_requested"
    job["message"] = "Stop requested. Verify will stop after the current row/checkpoint is saved."
    _save_important_verify_job(job)
    return JSONResponse({"ok": True, "message": "Stop requested for Verify Leads.", "job": _job_progress_payload(job)})


@app.post("/api/leads/dispatch-important/preview")
def preview_dispatch_important_leads(payload: ImportantLeadDispatchPayload | None = None) -> JSONResponse:
    try:
        preflight_block = _dispatch_preflight_block_response(snapshot=_build_live_snapshot())
        if preflight_block is not None:
            return preflight_block
        current_paths = important_leads_path_state()
        verify_paths = important_leads_verify_path_state()
        triage_paths = important_leads_triage_path_state()
        input_path = _resolve_dashboard_csv_path(
            getattr(payload, "input_path", current_paths["input_path"]) if payload else current_paths["input_path"],
            IMPORTANT_LEADS_INPUT,
        )
        output_path = _resolve_dashboard_csv_path(
            getattr(payload, "output_path", current_paths["output_path"]) if payload else current_paths["output_path"],
            IMPORTANT_LEADS_OUTPUT,
        )
        rejected_path = _resolve_dashboard_csv_path(
            getattr(payload, "rejected_path", current_paths["rejected_path"]) if payload else current_paths["rejected_path"],
            IMPORTANT_LEADS_REJECTED,
        )
        dispatch_source_mode = str(getattr(payload, "dispatch_source_mode", DISPATCH_SOURCE_TRIAGED_KEEP) if payload else DISPATCH_SOURCE_TRIAGED_KEEP).strip().lower() or DISPATCH_SOURCE_TRIAGED_KEEP
        dispatch_cap = str(getattr(payload, "dispatch_cap", DISPATCH_CAP_ALL) if payload else DISPATCH_CAP_ALL).strip().lower() or DISPATCH_CAP_ALL
        campaign_type = normalize_campaign_type(getattr(payload, "campaign_type", CAMPAIGN_TYPE_COLD) if payload else CAMPAIGN_TYPE_COLD)
        aliases = {
            "verified": DISPATCH_SOURCE_TRIAGED_KEEP,
            "fast_triage": DISPATCH_SOURCE_TRIAGED_KEEP,
            "fast_triage_keep": DISPATCH_SOURCE_TRIAGED_KEEP,
            "strict": DISPATCH_SOURCE_STRICT_VERIFIED,
            "strict_public_proof": DISPATCH_SOURCE_STRICT_VERIFIED,
        }
        dispatch_source_mode = aliases.get(dispatch_source_mode, dispatch_source_mode)
        if dispatch_source_mode not in {DISPATCH_SOURCE_TRIAGED_KEEP, DISPATCH_SOURCE_STRICT_VERIFIED, DISPATCH_SOURCE_CLEANED}:
            raise ValueError("Dispatch source mode must be triaged_keep, strict_verified, or cleaned.")
        if is_recontact_cold_campaign(campaign_type):
            dispatch_source_mode = DISPATCH_SOURCE_CLEANED
        save_state(important_leads_paths=_important_path_labels_for_state(input_path, output_path, rejected_path))
        save_state(important_leads_dispatch_source=_important_dispatch_source_labels_for_state(dispatch_source_mode))
        fast_triage_source = _latest_fast_triage_keep_source()
        if fast_triage_source.get("source_resolution") == "latest_completed_staged_run":
            triaged_keep_source_path = Path(str(fast_triage_source["path"]))
            triaged_keep_source_resolution = str(fast_triage_source.get("source_resolution") or "")
            staged_paths = fast_triage_source.get("paths") if isinstance(fast_triage_source.get("paths"), dict) else {}
            if dispatch_source_mode == DISPATCH_SOURCE_TRIAGED_KEEP and staged_paths:
                output_path = Path(staged_paths.get("input") or output_path)
                rejected_path = Path(staged_paths.get("rejected") or rejected_path)
        else:
            triaged_keep_source_path = _resolve_dashboard_csv_path(
                triage_paths["keep_path"],
                TRIAGED_KEEP_PATH,
            )
            triaged_keep_source_resolution = str(fast_triage_source.get("source_resolution") or "")
        verified_source_path = _resolve_dashboard_csv_path(
            verify_paths["verified_path"],
            STRICT_VERIFIED_PATH,
        )
        source_path_for_mode = {
            DISPATCH_SOURCE_TRIAGED_KEEP: triaged_keep_source_path,
            DISPATCH_SOURCE_STRICT_VERIFIED: verified_source_path,
            DISPATCH_SOURCE_CLEANED: output_path,
        }[dispatch_source_mode]
        source_block = _dispatch_source_readiness_block(
            dispatch_source_mode,
            source_path_for_mode,
            source_resolution=triaged_keep_source_resolution if dispatch_source_mode == DISPATCH_SOURCE_TRIAGED_KEEP else "",
        )
        if source_block is not None:
            return JSONResponse(
                {
                    "ok": False,
                    "blocked": True,
                    **source_block,
                    "source_path": _dashboard_path_label(source_path_for_mode),
                    "dispatch_source_mode": dispatch_source_mode,
                    "campaign_type": campaign_type,
                    "snapshot": _build_live_snapshot(),
                },
                status_code=409,
            )
        preview = preview_dispatch_master_leads(
            master_path=output_path,
            rejected_path=rejected_path,
            verified_path=verified_source_path,
            triaged_keep_path=triaged_keep_source_path,
            dispatch_source_mode=dispatch_source_mode,
            dispatch_cap=dispatch_cap,
            campaign_type=campaign_type,
            preview_dir=IMPORTANT_LEADS_DISPATCH_PREVIEWS,
        )
    except FileNotFoundError as exc:
        return JSONResponse({"ok": False, "error": "missing_source", "message": str(exc)}, status_code=404)
    except ValueError as exc:
        return JSONResponse({"ok": False, "error": "invalid_dispatch_request", "message": str(exc)}, status_code=400)
    except RuntimeError as exc:
        return JSONResponse({"ok": False, "error": "dispatch_preview_blocked", "message": str(exc)}, status_code=409)
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Lead dispatch preview failed: {exc}"}, status_code=500)

    return JSONResponse(
        {
            "ok": True,
            "message": f"Preview ready for {preview['dispatch_source_name']}.",
            "preview": preview,
            "status": _combined_leads_status(),
            "snapshot": _build_live_snapshot(),
        }
    )


@app.post("/api/leads/dispatch-important/safer-recontact-pool")
def create_safer_recontact_pool(payload: ImportantLeadDispatchPayload | None = None) -> JSONResponse:
    try:
        preview_id = str(getattr(payload, "preview_id", "") if payload else "").strip()
        summary = create_safer_recontact_pool_from_preview(
            preview_id,
            preview_dir=IMPORTANT_LEADS_DISPATCH_PREVIEWS,
        )
    except FileNotFoundError as exc:
        return JSONResponse({"ok": False, "error": "missing_recontact_preview", "message": str(exc)}, status_code=404)
    except ValueError as exc:
        return JSONResponse({"ok": False, "error": "invalid_recontact_preview", "message": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"ok": False, "error": "safer_recontact_failed", "message": f"Safer recontact pool failed: {exc}"}, status_code=500)
    return JSONResponse(
        {
            "ok": True,
            "message": f"Safer recontact pool created with {int(summary.get('safer_rows_written') or 0)} row(s).",
            "summary": summary,
            "status": _combined_leads_status(),
            "snapshot": _build_live_snapshot(),
        }
    )


def _dispatch_preflight_block_response(*, snapshot: dict[str, object] | None = None) -> JSONResponse | None:
    active_profiles = runtime_control.list_active_sender_snapshots(tail_lines=12)
    if not active_profiles:
        return None
    states = {str(item.name): str(item.runtime_state) for item in active_profiles}
    active_names = list(states.keys())
    return JSONResponse(
        {
            "ok": False,
            "blocked": True,
            "reason": "senders_active",
            "error": "senders_active",
            "active_profiles": active_names,
            "active_sender_count": len(active_names),
            "states": states,
            "message": f"Dispatch blocked: stop active senders first. Active: {', '.join(sorted(states))}",
            "snapshot": snapshot or _build_live_snapshot(),
        },
        status_code=409,
    )


def _dispatch_confirm_response(payload: ImportantLeadDispatchPayload | None = None) -> JSONResponse:
    try:
        preflight_block = _dispatch_preflight_block_response(snapshot=_build_live_snapshot())
        if preflight_block is not None:
            return preflight_block
        preview_id = str(getattr(payload, "preview_id", "") if payload else "").strip()
        if not preview_id:
            raise ValueError("Run Preview Dispatch first.")
        preview = validate_dispatch_preview(preview_id, preview_dir=IMPORTANT_LEADS_DISPATCH_PREVIEWS)
        if not str(preview.get("campaign_type") or "").strip():
            raise RuntimeError("Dispatch preview is missing campaign type. Re-run Preview Dispatch.")
        if not str(preview.get("dispatch_source_mode") or "").strip():
            raise RuntimeError("Dispatch preview is missing dispatch source mode. Re-run Preview Dispatch.")
        requested_campaign_type = normalize_campaign_type(
            getattr(payload, "campaign_type", preview.get("campaign_type"))
            if payload
            else preview.get("campaign_type")
        )
        preview_campaign_type = normalize_campaign_type(preview.get("campaign_type"))
        if requested_campaign_type != preview_campaign_type:
            raise RuntimeError("Dispatch preview does not match the selected campaign type. Re-run Preview Dispatch.")
        requested_mode = str(getattr(payload, "dispatch_source_mode", preview.get("dispatch_source_mode") or "") if payload else preview.get("dispatch_source_mode") or "").strip().lower()
        if requested_mode:
            aliases = {
                "verified": DISPATCH_SOURCE_TRIAGED_KEEP,
                "fast_triage": DISPATCH_SOURCE_TRIAGED_KEEP,
                "fast_triage_keep": DISPATCH_SOURCE_TRIAGED_KEEP,
                "strict": DISPATCH_SOURCE_STRICT_VERIFIED,
                "strict_public_proof": DISPATCH_SOURCE_STRICT_VERIFIED,
            }
            requested_mode = aliases.get(requested_mode, requested_mode)
            if is_recontact_cold_campaign(requested_campaign_type):
                requested_mode = DISPATCH_SOURCE_CLEANED
            if requested_mode != str(preview.get("dispatch_source_mode") or ""):
                raise RuntimeError("Dispatch preview does not match the selected source. Re-run Preview Dispatch.")
        current_status = _combined_leads_status()
        current_source_path = str(current_status.get("dispatch_source_path") or "").strip()
        if current_source_path and not _dashboard_paths_match(preview.get("dispatch_source_path"), current_source_path):
            raise RuntimeError("Dispatch preview source path does not match the current selected source. Re-run Preview Dispatch.")
        requested_cap = str(getattr(payload, "dispatch_cap", preview.get("dispatch_cap") or DISPATCH_CAP_ALL) if payload else preview.get("dispatch_cap") or DISPATCH_CAP_ALL).strip().lower()
        if requested_cap and requested_cap != str(preview.get("dispatch_cap") or DISPATCH_CAP_ALL):
            raise RuntimeError("Dispatch preview does not match the selected cap. Re-run Preview Dispatch.")
        recontact_recency = preview.get("recontact_recency") if isinstance(preview.get("recontact_recency"), dict) else {}
        recontact_override = bool(getattr(payload, "recontact_recency_override", False) if payload else False)
        if is_recontact_cold_campaign(preview_campaign_type) and bool(recontact_recency.get("high_risk")) and not recontact_override:
            raise RuntimeError("Recontact preview has high recent-contact overlap. Confirm requires explicit override.")
        job = _start_important_dispatch_job(
            preview_id=preview_id,
            campaign_type=preview_campaign_type,
            dispatch_source_mode=str(preview.get("dispatch_source_mode") or DISPATCH_SOURCE_TRIAGED_KEEP),
            dispatch_source_name=str(preview.get("dispatch_source_name") or ""),
            dispatch_source_path=str(preview.get("dispatch_source_path") or ""),
            dispatch_cap=str(preview.get("dispatch_cap") or DISPATCH_CAP_ALL),
            total_source_rows=int(preview.get("dispatch_source_row_count") or 0),
            eligible_rows=int(preview.get("dispatch_eligible_row_count") or 0),
            selected_rows=int(preview.get("dispatch_selected_row_count") or 0),
            total_rows_would_write=int(preview.get("total_rows_would_write") or 0),
            recontact_recency_override=recontact_override,
        )
    except FileNotFoundError as exc:
        return JSONResponse({"ok": False, "error": "missing_source", "message": str(exc)}, status_code=404)
    except ValueError as exc:
        return JSONResponse({"ok": False, "error": "invalid_dispatch_request", "message": str(exc)}, status_code=400)
    except RuntimeError as exc:
        error_code = "stale_preview" if "stale" in str(exc).lower() else "dispatch_blocked"
        return JSONResponse({"ok": False, "error": error_code, "message": str(exc)}, status_code=409)
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Lead dispatch failed: {exc}"}, status_code=500)

    return JSONResponse(
        {
            "ok": True,
            "message": f"Queued dispatch confirm job {job['job_id']}.",
            "job": job,
            "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
            "snapshot": _build_live_snapshot(),
        },
        status_code=202,
    )


def _write_private_jc_queue_repair_csv(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    with tmp_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({header: str(row.get(header, "") or "").strip() for header in headers})
    tmp_path.replace(path)
    settings.secure_private_file(path)


def _private_jc_repair_email_set(path: Path) -> set[str]:
    if not path.exists():
        return set()
    emails: set[str] = set()
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            email = ""
            for key in ("Email", "email", "AuthorEmail", "author_email"):
                value = str(row.get(key) or "").strip().lower()
                if value:
                    email = value
                    break
            if email:
                emails.add(email)
    return emails


def _confirmed_dispatch_archived_path(latest_dispatch: dict[str, object], key: str) -> Path | None:
    cleanup = latest_dispatch.get("staged_batch_cleanup") if isinstance(latest_dispatch, dict) else {}
    files = cleanup.get("files") if isinstance(cleanup, dict) else []
    if isinstance(files, list):
        for entry in files:
            if not isinstance(entry, dict) or str(entry.get("key") or "") != key:
                continue
            archive_path = Path(str(entry.get("archive_path") or ""))
            if archive_path.exists():
                return archive_path
    return None


def _latest_confirmed_dispatch_preview() -> tuple[dict[str, object], dict[str, object]]:
    state = load_state()
    latest_dispatch = state.get(MASTER_DISPATCH_STATE_KEY) if isinstance(state, dict) else {}
    if not isinstance(latest_dispatch, dict) or not latest_dispatch:
        raise RuntimeError("Run Preview Dispatch and Confirm Dispatch before repairing the Private JC queue.")
    if str(latest_dispatch.get("status") or "").strip().lower() not in {"completed", "confirmed"} and not latest_dispatch.get("generated_at_utc"):
        raise RuntimeError("Confirm Dispatch has not completed. Repair is blocked.")
    preview_id = str(latest_dispatch.get("preview_id") or "").strip()
    if not preview_id:
        raise RuntimeError("Latest dispatch does not reference a preview. Re-run Preview Dispatch and Confirm Dispatch.")
    preview = load_dispatch_preview(preview_id, preview_dir=IMPORTANT_LEADS_DISPATCH_PREVIEWS)
    if str(preview.get("status") or "").strip().lower() != "confirmed":
        raise RuntimeError("Latest dispatch preview is not confirmed. Run Confirm Dispatch before repairing the queue.")
    confirmed_run_id = str(preview.get("confirmed_run_id") or "").strip()
    dispatch_run_id = str(latest_dispatch.get("run_id") or "").strip()
    if confirmed_run_id and dispatch_run_id and confirmed_run_id != dispatch_run_id:
        raise RuntimeError("Latest dispatch preview does not match the confirmed dispatch run. Re-run Preview Dispatch and Confirm Dispatch.")
    return latest_dispatch, preview


@app.post("/api/profiles/private_jc/repair-queue")
def repair_private_jc_queue() -> JSONResponse:
    snapshot = _build_live_snapshot()
    preflight_block = _dispatch_preflight_block_response(snapshot=snapshot)
    if preflight_block is not None:
        return preflight_block
    try:
        before = build_dashboard_queue_safety_report("private_jc")
        if bool(before.get("safe")):
            return JSONResponse(
                {
                    "ok": True,
                    "repaired": False,
                    "message": "Private JC queue is already safe.",
                    "summary": {
                        "unsafe_rows_archived": 0,
                        "reject_overlap_rows_removed": 0,
                        "outside_source_rows_removed": 0,
                        "rebuilt_queue_rows": int(before.get("shard_row_count_total") or 0),
                        "backup_path": "",
                    },
                    "queue_safety": before,
                    "snapshot": snapshot,
                }
            )
        latest_dispatch, preview = _latest_confirmed_dispatch_preview()
        queue_headers = [str(value or "").strip() for value in (preview.get("queue_headers") or []) if str(value or "").strip()]
        if not queue_headers:
            raise RuntimeError("Confirmed dispatch preview is missing queue headers. Re-run Preview Dispatch and Confirm Dispatch.")
        queue_paths = preview.get("queue_paths") if isinstance(preview.get("queue_paths"), dict) else {}
        preview_private_path = Path(str(queue_paths.get("private_jc") or settings.SHARDS_DIR / "recipients_private_jc.csv"))
        private_queue_path = settings.SHARDS_DIR / "recipients_private_jc.csv"
        if preview_private_path.name != private_queue_path.name:
            raise RuntimeError("Confirmed dispatch preview does not target the Private JC queue. Re-run Preview Dispatch and Confirm Dispatch.")
        plan_rows_by_queue = preview.get("plan_rows_by_queue") if isinstance(preview.get("plan_rows_by_queue"), dict) else {}
        planned_rows = [
            {header: str(row.get(header, "") or "").strip() for header in queue_headers}
            for row in (plan_rows_by_queue.get("private_jc") or [])
            if isinstance(row, dict)
        ]
        live_queue_emails = _private_jc_repair_email_set(private_queue_path)
        confirmed_source_path = _confirmed_dispatch_archived_path(latest_dispatch, "triaged_keep")
        if confirmed_source_path is None:
            source_text = str(latest_dispatch.get("dispatch_source_path") or preview.get("dispatch_source_path") or "").strip()
            confirmed_source_path = Path(source_text) if source_text else None
        confirmed_source_emails = _private_jc_repair_email_set(confirmed_source_path) if confirmed_source_path else set()
        matching_current_source_reviewed = len(live_queue_emails & confirmed_source_emails) if confirmed_source_emails else 0
        outside_current_source_removed = (
            len(live_queue_emails - confirmed_source_emails)
            if confirmed_source_emails
            else max(
                int(before.get("outside_intended_source_count") or 0),
                int(before.get("outside_checked_output_count") or 0),
            )
        )
        confirmed_reject_path = _confirmed_dispatch_archived_path(latest_dispatch, "triaged_reject")
        confirmed_quarantine_path = _confirmed_dispatch_archived_path(latest_dispatch, "triaged_quarantine")
        reject_overlap_removed = len(live_queue_emails & _private_jc_repair_email_set(confirmed_reject_path)) if confirmed_reject_path else int(before.get("overlap_with_triaged_reject") or 0)
        quarantine_overlap_removed = len(live_queue_emails & _private_jc_repair_email_set(confirmed_quarantine_path)) if confirmed_quarantine_path else 0
        repair_dir = settings.BACKUPS_DIR / "private_jc_queue_repair" / timestamp_slug()
        repair_dir.mkdir(parents=True, exist_ok=True)
        settings.secure_private_dir(repair_dir)
        backup_path = repair_dir / private_queue_path.name
        if private_queue_path.exists():
            shutil.copy2(private_queue_path, backup_path)
        else:
            backup_path.write_text("", encoding="utf-8")
        settings.secure_private_file(backup_path)
        manifest = {
            "created_at_utc": iso_utc(),
            "purpose": "private_jc_queue_repair",
            "source": "confirmed_dispatch_preview",
            "preview_id": str(preview.get("preview_id") or latest_dispatch.get("preview_id") or ""),
            "dispatch_run_id": str(latest_dispatch.get("run_id") or ""),
            "confirmed_source_path": str(confirmed_source_path or ""),
            "live_queue_path": str(private_queue_path),
            "backup_path": str(backup_path),
            "unsafe_queue_rows_archived": int(before.get("shard_row_count_total") or 0),
            "reject_overlap_rows_removed": reject_overlap_removed,
            "quarantine_overlap_rows_removed": quarantine_overlap_removed,
            "outside_source_rows_removed": outside_current_source_removed,
            "matching_current_source_reviewed": matching_current_source_reviewed,
            "rebuilt_queue_rows": len(planned_rows),
        }
        manifest_path = repair_dir / "manifest.json"
        manifest_path.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        settings.secure_private_file(manifest_path)

        _write_private_jc_queue_repair_csv(private_queue_path, queue_headers, planned_rows)
        after = build_dashboard_queue_safety_report("private_jc")
        summary = {
            "unsafe_queue_rows_archived": manifest["unsafe_queue_rows_archived"],
            "reject_overlap_rows_removed": manifest["reject_overlap_rows_removed"],
            "quarantine_overlap_rows_removed": manifest["quarantine_overlap_rows_removed"],
            "outside_source_rows_removed": manifest["outside_source_rows_removed"],
            "matching_current_source_reviewed": manifest["matching_current_source_reviewed"],
            "rebuilt_queue_rows": manifest["rebuilt_queue_rows"],
            "backup_path": str(backup_path),
        }
        return JSONResponse(
            {
                "ok": True,
                "repaired": True,
                "message": "Private JC queue repaired from the latest confirmed dispatch preview. JC was not started.",
                "summary": summary,
                "queue_safety_before": before,
                "queue_safety": after,
                "snapshot": _build_live_snapshot(),
            }
        )
    except RuntimeError as exc:
        return JSONResponse({"ok": False, "blocked": True, "error": "private_jc_repair_blocked", "message": str(exc), "snapshot": snapshot}, status_code=409)
    except Exception as exc:
        return JSONResponse({"ok": False, "error": "private_jc_repair_failed", "message": f"Private JC queue repair failed: {exc}", "snapshot": snapshot}, status_code=500)


@app.post("/api/leads/dispatch-important/confirm")
def confirm_dispatch_important_leads(payload: ImportantLeadDispatchPayload | None = None) -> JSONResponse:
    return _dispatch_confirm_response(payload)


@app.post("/api/leads/dispatch-important")
def dispatch_important_leads(payload: ImportantLeadDispatchPayload | None = None) -> JSONResponse:
    return _dispatch_confirm_response(payload)


@app.get("/api/leads/dispatch-important/preflight")
def dispatch_important_leads_preflight() -> JSONResponse:
    snapshot = _build_live_snapshot()
    block = _dispatch_preflight_block_response(snapshot=snapshot)
    if block is not None:
        return block
    return JSONResponse(
        {
            "ok": True,
            "blocked": False,
            "reason": "",
            "error": "",
            "active_profiles": [],
            "active_sender_count": 0,
            "states": {},
            "message": "Dispatch preflight clear.",
            "snapshot": snapshot,
        }
    )


@app.get("/api/leads/dispatch-important/job/{job_id}")
def get_dispatch_important_leads_job(job_id: str) -> JSONResponse:
    try:
        job = _load_important_dispatch_job(job_id)
    except FileNotFoundError:
        return JSONResponse({"ok": False, "message": f"Dispatch job not found: {job_id}"}, status_code=404)
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Failed to load dispatch job: {exc}"}, status_code=500)
    return JSONResponse({"ok": True, "job": _job_progress_payload(job), "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()}})


@app.get("/api/leads/dispatch-important/active")
def get_active_dispatch_important_leads_job() -> JSONResponse:
    return JSONResponse(
        {
            "ok": True,
            "job": _find_active_dashboard_job(IMPORTANT_LEADS_DISPATCH_JOBS),
            "status": {**shard_status(), **important_leads_status(), **important_leads_verify_status()},
        }
    )


@app.get("/api/leads/quarantine-review")
def quarantine_review_list(
    reason_code: str = Query(default=""),
    stage: str = Query(default=""),
    status: str = Query(default="QUARANTINE"),
    sort: str = Query(default="score_desc"),
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
) -> JSONResponse:
    try:
        conn = connect_lead_ledger(_lead_ledger_db_path())
        try:
            review = list_quarantine_review_leads(
                conn,
                reason_code=reason_code,
                stage=stage,
                status=status,
                sort=sort,
                limit=limit,
                offset=offset,
            )
            review["recent_actions"] = load_recent_quarantine_review_actions(conn, limit=12)
        finally:
            conn.close()
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Failed to load quarantine review inbox: {exc}"}, status_code=500)
    return JSONResponse({"ok": True, "review": review})


@app.get("/api/leads/quarantine-review/{lead_id}")
def quarantine_review_detail(lead_id: str) -> JSONResponse:
    try:
        conn = connect_lead_ledger(_lead_ledger_db_path())
        try:
            lead = load_quarantine_review_lead(conn, lead_id)
        finally:
            conn.close()
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Failed to load quarantine review lead: {exc}"}, status_code=500)
    if lead is None:
        return JSONResponse({"ok": False, "message": f"Lead not found: {lead_id}"}, status_code=404)
    return JSONResponse({"ok": True, "lead": lead})


@app.post("/api/leads/quarantine-review/action")
def quarantine_review_action(payload: QuarantineReviewActionPayload) -> JSONResponse:
    try:
        conn = connect_lead_ledger(_lead_ledger_db_path())
        try:
            resolved_lead_ids = payload.lead_ids
            if payload.select_all_filtered:
                resolved_lead_ids = list_quarantine_review_lead_ids(
                    conn,
                    reason_code=payload.reason_code,
                    stage=payload.stage,
                    status=payload.status,
                    sort=payload.sort,
                    exclude_lead_ids=payload.excluded_lead_ids,
                )
            result = apply_quarantine_review_action(
                conn,
                lead_ids=resolved_lead_ids,
                action=payload.action,
                operator_note=payload.operator_note,
                run_id=f"quarantine_review_{timestamp_slug()}",
            )
            review = list_quarantine_review_leads(conn)
            review["recent_actions"] = load_recent_quarantine_review_actions(conn, limit=12)
        finally:
            conn.close()
    except ValueError as exc:
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Failed to apply quarantine review action: {exc}"}, status_code=500)
    return JSONResponse(
        {
            "ok": True,
            "message": f"Applied {result['action']} to {int(result['updated'] or 0)} lead(s).",
            "result": result,
            "review": review,
        }
    )


@app.post("/api/leads/shard")
def shard_leads(payload: ShardLeadsPayload) -> JSONResponse:
    active_snapshots = runtime_control.list_active_sender_snapshots(tail_lines=12)
    if active_snapshots:
        states = {snapshot.name: snapshot.runtime_state for snapshot in active_snapshots}
        return JSONResponse(
            {
                "ok": False,
                "error": "senders_active",
                "active_profiles": list(states.keys()),
                "states": states,
                "message": "Stop all senders before overwriting shards.",
            },
            status_code=409,
        )
    try:
        report = shard_cleaned_leads(
            cleaned_filename=payload.cleaned_filename,
            shard_count=payload.shard_count,
            strategy=payload.strategy,
        )
    except FileNotFoundError as exc:
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=404)
    except ValueError as exc:
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Sharding failed: {exc}"}, status_code=500)
    return JSONResponse(
        {
            "ok": True,
            "message": f"Shards updated from {report['source_cleaned_filename']} using {report['strategy']}.",
            "shard": report,
            "status": shard_status(),
            "snapshot": _build_live_snapshot(),
        }
    )


@app.post("/api/leads/shard/preview")
def preview_shard(payload: ShardLeadsPayload) -> JSONResponse:
    try:
        preview = preview_shard_cleaned_leads(
            cleaned_filename=payload.cleaned_filename,
            shard_count=payload.shard_count,
            strategy=payload.strategy,
        )
    except FileNotFoundError as exc:
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=404)
    except ValueError as exc:
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Preview failed: {exc}"}, status_code=500)
    return JSONResponse(
        {
            "ok": True,
            "message": f"Preview ready for {preview['source_cleaned_filename']}.",
            "preview": preview,
            "status": shard_status(),
        }
    )


@app.get("/api/leads/status")
def leads_status() -> JSONResponse:
    return JSONResponse({"ok": True, "status": _combined_leads_status()})


@app.post("/webhooks/sendgrid/events")
async def sendgrid_event_webhook(request: Request) -> JSONResponse:
    raw_body = await request.body()
    if SENDGRID_EVENT_PUBLIC_KEY:
        signature = request.headers.get(SENDGRID_SIG_HEADER, "")
        timestamp = request.headers.get(SENDGRID_TS_HEADER, "")
        if not signature or not timestamp:
            return JSONResponse({"ok": False, "message": "Missing SendGrid signature headers."}, status_code=401)
        try:
            verified = _verify_sendgrid_signature(raw_body, signature, timestamp)
        except Exception:
            return JSONResponse({"ok": False, "message": "Invalid SendGrid verification key or signature payload."}, status_code=400)
        if not verified:
            return JSONResponse({"ok": False, "message": "Invalid SendGrid webhook signature."}, status_code=401)
    try:
        payload = json.loads(raw_body.decode("utf-8"))
    except Exception:
        return JSONResponse({"ok": False, "message": "Invalid JSON payload."}, status_code=400)

    if not isinstance(payload, list):
        return JSONResponse({"ok": False, "message": "Expected JSON array."}, status_code=400)

    received_at = datetime.now(timezone.utc)
    normalized = normalize_webhook_events(
        [item for item in payload if isinstance(item, dict)],
        source_log=WEBHOOK_EVENTS_JSONL,
        received_at_utc=received_at,
    )
    dedupe_result = dedupe_webhook_events(normalized, WEBHOOK_DEDUPE_PATH, reference_utc=received_at)
    unique_events = list(dedupe_result.get("unique_events", []))
    appended = append_events_jsonl(unique_events, WEBHOOK_EVENTS_PATH)
    suppression_summary = update_suppressions_from_events(unique_events, SUPPRESSION_CSV)
    ledger_summary = ingest_send_outcome_events(unique_events, db_path=settings.LEAD_LEDGER_DB_PATH)
    auto_stops = runtime_control.apply_delivery_guards()
    json_safe_summary = {
        "updated_events": int(suppression_summary.get("updated_events", 0) or 0),
        "records_total": int(suppression_summary.get("records_total", 0) or 0),
        "total_perm": int(suppression_summary.get("total_perm", 0) or 0),
        "total_temp_active": int(suppression_summary.get("total_temp_active", 0) or 0),
    }
    json_safe_ledger_summary = {
        "processed_events": int(ledger_summary.get("processed_events", 0) or 0),
        "matched_events": int(ledger_summary.get("matched_events", 0) or 0),
        "unmatched_events": int(ledger_summary.get("unmatched_events", 0) or 0),
        "ignored_events": int(ledger_summary.get("ignored_events", 0) or 0),
        "dispatch_rows_updated": int(ledger_summary.get("dispatch_rows_updated", 0) or 0),
        "lead_rows_updated": int(ledger_summary.get("lead_rows_updated", 0) or 0),
        "suppressed_events": int(ledger_summary.get("suppressed_events", 0) or 0),
        "outcome_counts": {
            str(key): int(value or 0)
            for key, value in dict(ledger_summary.get("outcome_counts", {})).items()
        },
    }
    return JSONResponse(
        {
            "ok": True,
            "received": len(payload),
            "normalized": len(normalized),
            "stored": appended,
            "duplicates_ignored": int(dedupe_result.get("duplicates", 0) or 0),
            "suppression_summary": json_safe_summary,
            "ledger_summary": json_safe_ledger_summary,
            "auto_stops": auto_stops,
        }
    )


@app.websocket("/ws")
async def websocket_snapshot_stream(
    websocket: WebSocket,
    hours: int = Query(default=24, ge=1, le=168),
    tail_lines: int = Query(default=12, ge=4, le=50),
) -> None:
    if not _dashboard_auth_enabled() or not _dashboard_is_authenticated(websocket):
        await websocket.close(code=4401)
        return
    await websocket.accept()
    try:
        while True:
            await websocket.send_json(_build_live_snapshot(activity_hours=hours, tail_lines=tail_lines))
            await asyncio.sleep(1)
    except WebSocketDisconnect:
        return

# --- LOCAL DASHBOARD NO-AUTH OVERRIDE ---
# When DASHBOARD_AUTH_PASSWORD is empty, this local dashboard should behave as
# already authenticated instead of showing the Sign in / Auth unavailable panel.
def _dashboard_auth_response():
    auth_enabled = _dashboard_auth_enabled()
    return {
        "ok": True,
        "auth_enabled": auth_enabled,
        "authenticated": True if not auth_enabled else False,
        "username": str(settings.DASHBOARD_AUTH_USERNAME or "admin"),
        "local_mode": True if not auth_enabled else False,
    }


def _dashboard_is_authenticated(scope):
    if not _dashboard_auth_enabled():
        return True
    session = getattr(scope, "session", None)
    if not isinstance(session, dict):
        session = {}
    return bool(session.get(_AUTH_SESSION_KEY))
